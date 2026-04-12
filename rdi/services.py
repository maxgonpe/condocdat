import csv
import re
from datetime import date, datetime, timedelta, timezone as py_timezone

import xlrd
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    PLANOS_INICIALES_SHEET_SLUGS,
    PlanosImport,
    PlanosInicialesImport,
    PlanosInicialesRecord,
    PlanosRecord,
    RDIImport,
    RDIRecord,
    RDI_STATUS_CHOICES,
    RDI_STATUS_ABIERTA,
    RDI_STATUS_BORRADOR,
    RDI_STATUS_CERRADA,
    RDI_STATUS_NULA,
    RDI_STATUS_RECHAZADA,
    RDI_STATUS_REMITIDA,
    RDI_STATUS_RESPONDIDA,
)


_FILENAME_SNAPSHOT_RE = re.compile(
    r"SDI\s*-\s*(?P<date>\d{4}-\d{2}-\d{2})\s+(?P<h>\d{2})_(?P<m>\d{2})_(?P<s>\d{2})",
    re.IGNORECASE,
)
_PLANOS_FILENAME_SNAPSHOT_RE = re.compile(r"(?P<date>\d{4}-\d{2}-\d{2}).*?(?P<stamp>\d{12})(?=\.xlsx?$)", re.IGNORECASE)
_PLANOS_FILENAME_FALLBACK_RE = re.compile(
    r"(?P<date>\d{4}-\d{2}-\d{2})\s+(?P<h>\d{1,2})_(?P<m>\d{1,2})",
    re.IGNORECASE,
)


def parse_snapshot_datetime_from_filename(filename: str):
    """
    Espera algo como:
      ... - SDI - 2026-03-20 08_41_83
    Retorna datetime timezone-aware o None.
    """
    if not filename:
        return None

    m = _FILENAME_SNAPSHOT_RE.search(filename)
    if not m:
        return None

    try:
        year = int(m.group("date")[:4])
        month = int(m.group("date")[5:7])
        day = int(m.group("date")[8:10])
        hh = int(m.group("h"))
        mm = int(m.group("m"))
        ss = int(m.group("s"))

        # El CSV/archivo puede venir con "segundos" >= 60 (ej. 08_41_83).
        # Usamos timedelta para normalizar overflow hacia minutos/horas.
        dt = datetime(year, month, day, hh, 0, 0) + timedelta(minutes=mm, seconds=ss)
    except Exception:
        return None

    # Usa el timezone actual de Django.
    tz = timezone.get_current_timezone()
    return timezone.make_aware(dt, tz)


def parse_planos_snapshot_datetime_from_filename(filename: str):
    if not filename:
        return None

    m = _PLANOS_FILENAME_SNAPSHOT_RE.search(filename)
    if m:
        try:
            date_s = m.group("date")
            stamp = m.group("stamp")
            dt = datetime(
                year=int(date_s[:4]),
                month=int(date_s[5:7]),
                day=int(date_s[8:10]),
                hour=int(stamp[8:10]),
                minute=int(stamp[10:12]),
                second=0,
            )
            return timezone.make_aware(dt, timezone.get_current_timezone())
        except Exception:
            pass

    m2 = _PLANOS_FILENAME_FALLBACK_RE.search(filename)
    if m2:
        try:
            date_s = m2.group("date")
            dt = datetime(
                year=int(date_s[:4]),
                month=int(date_s[5:7]),
                day=int(date_s[8:10]),
                hour=int(m2.group("h")),
                minute=int(m2.group("m")),
                second=0,
            )
            return timezone.make_aware(dt, timezone.get_current_timezone())
        except Exception:
            return None
    return None


def map_csv_status_to_choice(raw_status: str) -> str:
    if not raw_status:
        return RDI_STATUS_NULA
    s = str(raw_status).strip().lower()
    if not s:
        return RDI_STATUS_NULA

    mapping = {
        # lo que trae el CSV que encontramos
        "open": RDI_STATUS_ABIERTA,
        "answered": RDI_STATUS_RESPONDIDA,
        "closed": RDI_STATUS_CERRADA,
        # posibles variantes
        "draft": RDI_STATUS_BORRADOR,
        "borrador": RDI_STATUS_BORRADOR,
        "preliminar": RDI_STATUS_BORRADOR,
        "remitida": RDI_STATUS_REMITIDA,
        "remitido": RDI_STATUS_REMITIDA,
        "enviada": RDI_STATUS_REMITIDA,
        "rejected": RDI_STATUS_RECHAZADA,
        "rechazada": RDI_STATUS_RECHAZADA,
        "rechazado": RDI_STATUS_RECHAZADA,
        "abierta": RDI_STATUS_ABIERTA,
        "respondida": RDI_STATUS_RESPONDIDA,
        "cerrada": RDI_STATUS_CERRADA,
        "nula": RDI_STATUS_NULA,
        "": RDI_STATUS_NULA,
    }
    return mapping.get(s, RDI_STATUS_NULA)


def _parse_csv_datetime(value):
    """
    Parse robusto para valores tipo:
      03/19/2026 12:31 PM (UTC)
      03/23/2026 (UTC)
    """
    if value is None:
        return None
    v = str(value).strip()
    if not v or v == '""':
        return None

    # Normaliza espacios extra.
    v = re.sub(r"\s+", " ", v)

    # Formato con hora 12h
    for fmt in ("%m/%d/%Y %I:%M %p (UTC)", "%m/%d/%Y %H:%M (UTC)"):
        try:
            dt = datetime.strptime(v, fmt)
            return timezone.make_aware(dt, py_timezone.utc)
        except Exception:
            pass

    # Formato solo fecha
    for fmt in ("%m/%d/%Y (UTC)",):
        try:
            d = datetime.strptime(v, fmt).date()
            dt = datetime(d.year, d.month, d.day, 0, 0, 0)
            return timezone.make_aware(dt, py_timezone.utc)
        except Exception:
            pass

    return None


def _parse_csv_bool(value):
    if value is None:
        return None
    s = str(value).strip().lower()
    if not s or s == '""':
        return None
    if s in ("yes", "y", "si", "sí"):
        return True
    if s in ("no", "n"):
        return False
    return None


def _norm_str(x) -> str:
    if x is None:
        return ""
    s = str(x)
    s = s.replace("\r\n", "\n").replace("\r", "\n").strip()
    return s


def _parse_planos_datetime_es(value):
    """
    Parse fechas en español tipo:
      "26 de jul. de 2023 22:45"
      "8 de ago. de 2023 18:40"
      "8 de ago de 2023"
    """
    if value is None:
        return None
    v = _norm_str(value)
    if not v:
        return None

    months = {
        "ene": 1,
        "feb": 2,
        "mar": 3,
        "abr": 4,
        "may": 5,
        "jun": 6,
        "jul": 7,
        "ago": 8,
        "sep": 9,
        "oct": 10,
        "nov": 11,
        "dic": 12,
    }
    m = re.match(
        r"^\s*(?P<d>\d{1,2})\s+de\s+(?P<mon>[a-zA-Záéíóúñ\.]+)\s+de\s+(?P<y>\d{4})(?:\s+(?P<h>\d{1,2}):(?P<mi>\d{2}))?\s*$",
        v,
        flags=re.IGNORECASE,
    )
    if not m:
        return None

    mon_raw = (m.group("mon") or "").lower().replace(".", "")
    mon_raw = (
        mon_raw.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )
    month = months.get(mon_raw[:3])
    if not month:
        return None

    try:
        day = int(m.group("d"))
        year = int(m.group("y"))
        hh = int(m.group("h")) if m.group("h") is not None else 0
        mm = int(m.group("mi")) if m.group("mi") is not None else 0
        dt = datetime(year, month, day, hh, mm, 0)
        return timezone.make_aware(dt, timezone.get_current_timezone())
    except Exception:
        return None


def _parse_planos_date_es(value):
    dt = _parse_planos_datetime_es(value)
    if dt:
        return dt.date()
    return None


@transaction.atomic
def attach_rdi_csv(uploaded_file, original_filename: str):
    """
    1) Crea un `RDIImport` a partir del archivo adjunto.
    2) Importa el CSV y actualiza/crea `RDIRecord` por `Id`.
    3) Si llega un CSV posterior, actualiza los campos que cambiaron.
    """
    snapshot_dt = parse_snapshot_datetime_from_filename(original_filename)
    imp = RDIImport.objects.create(
        file=uploaded_file,
        original_filename=original_filename,
        snapshot_datetime=snapshot_dt,
    )

    _import_rdi_csv_file(imp)
    return imp


def _import_rdi_csv_file(imp: RDIImport):
    file_path = imp.file.path
    snapshot_dt = imp.snapshot_datetime

    with open(file_path, "r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.DictReader(f)

        for row in reader:
            raw_id = (row.get("Id") or "").strip()
            try:
                csv_id = int(raw_id)
            except Exception:
                continue

            new_fields = {
                "csv_id": csv_id,
                "title": _norm_str(row.get("Title")),
                "question": _norm_str(row.get("Question")),
                "suggested_answer": _norm_str(row.get("Suggested answer")),
                "location_details": _norm_str(row.get("Location details")),
                "status": map_csv_status_to_choice(row.get("Status")),
                "response": _norm_str(row.get("Response")),
                "assigned_to": _norm_str(row.get("Assigned to")),
                "assignee_type": _norm_str(row.get("Assignee type")),
                "company": _norm_str(row.get("Company")),
                "due_date": _parse_csv_datetime(row.get("Due date")),
                "associated_to_document": _parse_csv_bool(row.get("Associated to document?")),
                "created_at": _parse_csv_datetime(row.get("Created at")),
                "created_by": _norm_str(row.get("Created by")),
                "updated_at": _parse_csv_datetime(row.get("Updated at")),
                "updated_by": _norm_str(row.get("Updated by")),
                "distribution_list": _norm_str(row.get("Distribution list")),
                "cost_impact": _norm_str(row.get("Cost impact")),
                "schedule_impact": _norm_str(row.get("Schedule impact")),
                "priority": _norm_str(row.get("Priority")),
                "discipline": _norm_str(row.get("Discipline")),
                "category": _norm_str(row.get("Category")),
                "reference": _norm_str(row.get("Reference")),
            }

            existing = RDIRecord.objects.filter(csv_id=csv_id).first()
            if not existing:
                new_fields["last_snapshot_datetime"] = snapshot_dt
                new_fields["last_diff_fields"] = "CREATED"
                new_fields["last_import"] = imp
                RDIRecord.objects.create(**new_fields)
                continue

            changed = []
            for field, new_val in new_fields.items():
                if field == "csv_id":
                    continue
                old_val = getattr(existing, field)
                if old_val != new_val:
                    changed.append(field)
                    setattr(existing, field, new_val)

            # Siempre registramos la última snapshot (aunque no haya cambios de contenido).
            if existing.last_snapshot_datetime != snapshot_dt:
                changed.append("last_snapshot_datetime")
            existing.last_snapshot_datetime = snapshot_dt
            existing.last_import = imp

            existing.last_diff_fields = ", ".join(changed)[:2000]

            # Si no hubo cambios de contenido y el snapshot es igual, evitamos escrituras.
            if changed and existing.last_diff_fields:
                existing.save()
            else:
                # Aun así, el FK last_import debe quedar consistente si era distinto.
                if existing.last_import_id != imp.id:
                    existing.save()


def get_rdi_records_for_ajax(q: str = "", limit: int = 200):
    """
    Devuelve el listado actual (último estado) para renderizar con AJAX.
    """
    qs = RDIRecord.objects.select_related("last_import")

    query = (q or "").strip()
    if query:
        terms = [t for t in re.split(r"\s+", query) if t]
        if terms:
            # Búsqueda por "términos": cada término debe estar en al menos
            # uno de los campos relevantes (AND global).
            from django.db.models import Q

            status_label_to_code = {label.lower(): code for code, label in RDI_STATUS_CHOICES}
            # también permitimos que el usuario busque por el código interno
            status_code_values = {code.lower(): code for code, _ in RDI_STATUS_CHOICES}

            fields = [
                # Nota: csv_id es entero, por eso se filtra aparte si el término es numérico.
                "title",
                "question",
                "suggested_answer",
                "location_details",
                "response",
                "assigned_to",
                "assignee_type",
                "company",
                "priority",
                "discipline",
                "category",
                "reference",
            ]

            combined = None
            for t in terms:
                t_l = t.lower()
                term_q = Q()

                # Si el término es numérico, permite buscar por csv_id exacto.
                if t_l.isdigit():
                    try:
                        term_q |= Q(csv_id=int(t_l))
                    except Exception:
                        pass

                for f in fields:
                    term_q |= Q(**{f + "__icontains": t_l})

                # status por label o por código
                if t_l in status_label_to_code:
                    term_q |= Q(status=status_label_to_code[t_l])
                if t_l in status_code_values:
                    term_q |= Q(status=status_code_values[t_l])

                combined = term_q if combined is None else (combined & term_q)

            if combined is not None:
                qs = qs.filter(combined)

    records = qs.order_by("-last_snapshot_datetime", "-csv_id")[:limit]

    out = []
    for r in records:
        status_label = dict(RDIRecord._meta.get_field("status").choices).get(r.status, r.status)
        informado_label = dict(RDIRecord._meta.get_field("informado").choices).get(
            r.informado, r.informado
        )
        out.append(
            {
                "id": r.id,
                "csv_id": r.csv_id,
                "title": r.title,
                "status": r.status,
                "status_label": status_label,
                "informado": r.informado,
                "informado_label": informado_label,
                "question": r.question,
                "suggested_answer": r.suggested_answer,
                "location_details": r.location_details,
                "response": r.response,
                "assigned_to": r.assigned_to,
                "assignee_type": r.assignee_type,
                "company": r.company,
                "due_date": r.due_date.isoformat() if r.due_date else None,
                "associated_to_document": r.associated_to_document,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "created_by": r.created_by,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
                "updated_by": r.updated_by,
                "distribution_list": r.distribution_list,
                "cost_impact": r.cost_impact,
                "schedule_impact": r.schedule_impact,
                "priority": r.priority,
                "discipline": r.discipline,
                "category": r.category,
                "reference": r.reference,
                "last_snapshot_datetime": (
                    r.last_snapshot_datetime.isoformat() if r.last_snapshot_datetime else None
                ),
                "last_diff_fields": r.last_diff_fields,
                "last_import_id": r.last_import_id,
            }
        )
    return out


def get_rdi_cost_schedule_impacts_for_ajax(q: str = "", limit: int = 300):
    """
    Devuelve registros RDI filtrados por impacto (costo o plazo = yes/si/sí),
    para el listado "Aumentos/disminuciones".
    """
    from django.db.models import Q

    qs = RDIRecord.objects.select_related("last_import").filter(
        Q(cost_impact__iregex=r"^\s*(yes|si|sí)\s*$")
        | Q(schedule_impact__iregex=r"^\s*(yes|si|sí)\s*$")
    )

    query = (q or "").strip()
    if query:
        terms = [t for t in re.split(r"\s+", query) if t]
        if terms:
            combined = None
            for t in terms:
                term_q = (
                    Q(title__icontains=t)
                    | Q(location_details__icontains=t)
                    | Q(cost_impact__icontains=t)
                    | Q(schedule_impact__icontains=t)
                    | Q(discipline__icontains=t)
                    | Q(priority__icontains=t)
                    | Q(status__icontains=t)
                )
                if t.isdigit():
                    try:
                        term_q |= Q(csv_id=int(t))
                    except Exception:
                        pass
                combined = term_q if combined is None else (combined & term_q)
            if combined is not None:
                qs = qs.filter(combined)

    records = qs.order_by("-last_snapshot_datetime", "-csv_id")[:limit]
    status_label_map = dict(RDIRecord._meta.get_field("status").choices)

    def _translate_term_es(value: str) -> str:
        v = (value or "").strip()
        if not v:
            return ""
        key = v.lower()
        translations = {
            "yes": "Sí",
            "no": "No",
            "high": "Alta",
            "hight": "Alta",
            "medium": "Media",
            "low": "Baja",
            "critical": "Crítica",
            "open": "Abierta",
            "closed": "Cerrada",
            "answered": "Respondida",
            "draft": "Borrador",
        }
        return translations.get(key, v)

    out = []
    for r in records:
        out.append(
            {
                "csv_id": r.csv_id,
                "title": r.title,
                "status": r.status,
                "status_label": status_label_map.get(r.status, r.status),
                "location_details": r.location_details,
                "cost_impact": r.cost_impact,
                "cost_impact_label": _translate_term_es(r.cost_impact),
                "schedule_impact": r.schedule_impact,
                "schedule_impact_label": _translate_term_es(r.schedule_impact),
                "discipline": r.discipline,
                "priority": r.priority,
                "priority_label": _translate_term_es(r.priority),
                "question": r.question,
                "response": r.response,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            }
        )
    return out


@transaction.atomic
def attach_planos_xlsx(uploaded_file, original_filename: str):
    snapshot_dt = parse_planos_snapshot_datetime_from_filename(original_filename)
    imp = PlanosImport.objects.create(
        file=uploaded_file,
        original_filename=original_filename,
        snapshot_datetime=snapshot_dt,
    )
    _import_planos_xlsx_file(imp)
    return imp


def _import_planos_xlsx_file(imp: PlanosImport):
    file_path = imp.file.path
    snapshot_dt = imp.snapshot_datetime

    wb = load_workbook(file_path, data_only=True)
    ws = wb["Contenido del informe"] if "Contenido del informe" in wb.sheetnames else wb[wb.sheetnames[0]]

    headers = [_norm_str(ws.cell(row=1, column=i).value) for i in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}

    def val(row_num, header):
        col = idx.get(header)
        if not col:
            return ""
        return _norm_str(ws.cell(row=row_num, column=col).value)

    for row_num in range(2, ws.max_row + 1):
        folder_path = val(row_num, "Ruta y nombre de la carpeta")
        name = val(row_num, "Nombre")
        if not folder_path and not name:
            continue

        new_fields = {
            "folder_path": folder_path,
            "name": name,
            "description": val(row_num, "Descripción"),
            "version": val(row_num, "Versión"),
            "size": val(row_num, "Tamaño"),
            "last_update_raw": val(row_num, "Última actualización"),
            "last_update_at": _parse_planos_datetime_es(val(row_num, "Última actualización")),
            "updated_by": val(row_num, "Actualizado por"),
            "last_upload_raw": val(row_num, "Última carga"),
            "last_upload_at": _parse_planos_datetime_es(val(row_num, "Última carga")),
            "uploaded_by": val(row_num, "Cargado por"),
            "review_mark": val(row_num, "Marca de revisión"),
            "incidence": val(row_num, "Incidencia"),
            "sdi": val(row_num, "SDI"),
            "review_status": val(row_num, "Estado de revisión"),
            "set_name": val(row_num, "Conjunto"),
            "issue_date_raw": val(row_num, "Fecha de emisión"),
            "issue_date": _parse_planos_date_es(val(row_num, "Fecha de emisión")),
            "sheet_number": val(row_num, "Sheet number"),
            "title": val(row_num, "Title"),
            "revision": val(row_num, "Revisión"),
        }

        existing = PlanosRecord.objects.filter(folder_path=folder_path, name=name).first()
        if not existing:
            new_fields["last_snapshot_datetime"] = snapshot_dt
            new_fields["last_diff_fields"] = "CREATED"
            new_fields["last_import"] = imp
            PlanosRecord.objects.create(**new_fields)
            continue

        changed = []
        for field, new_val in new_fields.items():
            old_val = getattr(existing, field)
            if old_val != new_val:
                changed.append(field)
                setattr(existing, field, new_val)

        if existing.last_snapshot_datetime != snapshot_dt:
            changed.append("last_snapshot_datetime")
        existing.last_snapshot_datetime = snapshot_dt
        existing.last_import = imp
        existing.last_diff_fields = ", ".join(changed)[:2000]

        if changed:
            existing.save()
        else:
            if existing.last_import_id != imp.id:
                existing.save()


def get_planos_records_for_ajax(q: str = "", limit: int = 300):
    from django.db.models import Q

    qs = PlanosRecord.objects.select_related("last_import")
    query = (q or "").strip()
    if query:
        terms = [t for t in re.split(r"\s+", query) if t]
        if terms:
            combined = None
            fields = [
                "folder_path",
                "name",
                "description",
                "version",
                "updated_by",
                "uploaded_by",
                "review_mark",
                "incidence",
                "sdi",
                "review_status",
                "set_name",
                "sheet_number",
                "title",
                "revision",
            ]
            for t in terms:
                term_q = Q()
                for f in fields:
                    term_q |= Q(**{f + "__icontains": t})
                combined = term_q if combined is None else (combined & term_q)
            if combined is not None:
                qs = qs.filter(combined)

    records = qs.order_by("-last_snapshot_datetime", "name")[:limit]
    out = []
    for r in records:
        out.append(
            {
                "id": r.id,
                "folder_path": r.folder_path,
                "name": r.name,
                "description": r.description,
                "version": r.version,
                "size": r.size,
                "last_update_raw": r.last_update_raw,
                "last_update_at": r.last_update_at.isoformat() if r.last_update_at else None,
                "updated_by": r.updated_by,
                "last_upload_raw": r.last_upload_raw,
                "last_upload_at": r.last_upload_at.isoformat() if r.last_upload_at else None,
                "uploaded_by": r.uploaded_by,
                "review_mark": r.review_mark,
                "incidence": r.incidence,
                "sdi": r.sdi,
                "review_status": r.review_status,
                "set_name": r.set_name,
                "issue_date_raw": r.issue_date_raw,
                "issue_date": r.issue_date.isoformat() if r.issue_date else None,
                "sheet_number": r.sheet_number,
                "title": r.title,
                "revision": r.revision,
                "last_snapshot_datetime": (
                    r.last_snapshot_datetime.isoformat() if r.last_snapshot_datetime else None
                ),
                "last_diff_fields": r.last_diff_fields,
                "last_import_id": r.last_import_id,
            }
        )
    return out


def _dedupe_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    out: list[str] = []
    for h in headers:
        c = counts.get(h, 0) + 1
        counts[h] = c
        out.append(h if c == 1 else f"{h}_{c}")
    return out


def _xlrd_cell_to_str(book, sheet, rowx: int, colx: int) -> str:
    try:
        ctype = sheet.cell_type(rowx, colx)
        val = sheet.cell_value(rowx, colx)
    except Exception:
        return ""
    if ctype == xlrd.XL_CELL_DATE:
        try:
            tup = xlrd.xldate_as_tuple(val, book.datemode)
            if tup[0] == 0 and tup[1] == 0 and tup[2] == 0:
                return f"{tup[3]:02d}:{tup[4]:02d}:{tup[5]:02d}"
            return f"{tup[0]:04d}-{tup[1]:02d}-{tup[2]:02d}"
        except Exception:
            return _norm_str(val)
    if ctype == xlrd.XL_CELL_NUMBER:
        if val == int(val):
            return str(int(val))
        return str(val)
    if ctype == xlrd.XL_CELL_BOOLEAN:
        return "TRUE" if val else "FALSE"
    if val is None:
        return ""
    return _norm_str(val)


def _planos_iniciales_search_blob(specialty: str, row_data: dict) -> str:
    parts = [specialty.lower()]
    for k, v in row_data.items():
        parts.append(str(k).lower())
        parts.append(str(v).lower())
    return " ".join(parts)


def _openpyxl_cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        if val.hour == 0 and val.minute == 0 and val.second == 0:
            return val.date().isoformat()
        return val.strftime("%Y-%m-%d %H:%M")
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (int, float)):
        if val == int(val):
            return str(int(val))
        return str(val)
    return _norm_str(val)


def _import_planos_iniciales_rows(
    imp: PlanosInicialesImport,
    slug: str,
    headers: list[str],
    data_rows: list[list],
    row_to_str,
):
    snapshot_dt = imp.snapshot_datetime
    for row_idx, row_vals in enumerate(data_rows, start=1):
        row_data = {}
        empty = True
        for c, header in enumerate(headers):
            raw = row_vals[c] if c < len(row_vals) else None
            cell_s = row_to_str(raw)
            row_data[header] = cell_s
            if cell_s:
                empty = False
        if empty:
            continue

        excel_row = row_idx + 1
        search_text = _planos_iniciales_search_blob(slug, row_data)
        new_fields = {
            "specialty": slug.lower(),
            "excel_row": excel_row,
            "columns_json": row_data,
            "column_headers_order": list(headers),
            "search_text": search_text[:8000],
        }

        existing = PlanosInicialesRecord.objects.filter(
            specialty=slug.lower(), excel_row=excel_row
        ).first()
        if not existing:
            new_fields["last_snapshot_datetime"] = snapshot_dt
            new_fields["last_diff_fields"] = "CREATED"
            new_fields["last_import"] = imp
            PlanosInicialesRecord.objects.create(**new_fields)
            continue

        changed = []
        for field, new_val in new_fields.items():
            old_val = getattr(existing, field)
            if old_val != new_val:
                changed.append(field)
                setattr(existing, field, new_val)

        if existing.last_snapshot_datetime != snapshot_dt:
            changed.append("last_snapshot_datetime")
        existing.last_snapshot_datetime = snapshot_dt
        existing.last_import = imp
        existing.last_diff_fields = ", ".join(changed)[:2000]

        if changed:
            existing.save()
        elif existing.last_import_id != imp.id:
            existing.save()


@transaction.atomic
def attach_planos_iniciales(uploaded_file, original_filename: str):
    snapshot_dt = parse_planos_snapshot_datetime_from_filename(original_filename)
    imp = PlanosInicialesImport.objects.create(
        file=uploaded_file,
        original_filename=original_filename,
        snapshot_datetime=snapshot_dt,
    )
    lower = (original_filename or "").lower()
    if lower.endswith(".xlsx"):
        _import_planos_iniciales_xlsx_file(imp)
    else:
        _import_planos_iniciales_xls_file(imp)
    return imp


def _import_planos_iniciales_xls_file(imp: PlanosInicialesImport):
    file_path = imp.file.path
    book = xlrd.open_workbook(file_path)
    name_by_lower = {sh.name.strip().lower(): sh.name for sh in book.sheets()}

    for slug in PLANOS_INICIALES_SHEET_SLUGS:
        actual_name = name_by_lower.get(slug.lower())
        if not actual_name:
            continue

        sh = book.sheet_by_name(actual_name)
        if sh.nrows < 2:
            continue

        headers_raw = []
        for c in range(sh.ncols):
            h = _norm_str(sh.cell_value(0, c))
            if not h:
                h = f"Column_{c + 1}"
            headers_raw.append(h)
        headers = _dedupe_headers(headers_raw)

        data_rows = []
        for rowx in range(1, sh.nrows):
            row_list = [_xlrd_cell_to_str(book, sh, rowx, c) for c in range(len(headers))]
            data_rows.append(row_list)
        _import_planos_iniciales_rows(imp, slug, headers, data_rows, lambda v: v if isinstance(v, str) else _norm_str(v))


def _import_planos_iniciales_xlsx_file(imp: PlanosInicialesImport):
    file_path = imp.file.path
    wb = load_workbook(file_path, data_only=True)
    name_by_lower = {str(ws.title).strip().lower(): ws.title for ws in wb.worksheets}

    for slug in PLANOS_INICIALES_SHEET_SLUGS:
        actual_title = name_by_lower.get(slug.lower())
        if not actual_title:
            continue
        ws = wb[actual_title]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers_raw = []
        for c, cell in enumerate(rows[0]):
            h = _norm_str(cell)
            if not h:
                h = f"Column_{c + 1}"
            headers_raw.append(h)
        headers = _dedupe_headers(headers_raw)
        data_rows = []
        for row_tuple in rows[1:]:
            row_list = list(row_tuple) if row_tuple else []
            while len(row_list) < len(headers):
                row_list.append(None)
            data_rows.append(row_list[: len(headers)])
        _import_planos_iniciales_rows(imp, slug, headers, data_rows, _openpyxl_cell_to_str)


def get_planos_iniciales_records_for_ajax(q: str = "", limit: int = 400):
    from django.db.models import Q

    qs = PlanosInicialesRecord.objects.select_related("last_import")
    query = (q or "").strip()
    if query:
        terms = [t for t in re.split(r"\s+", query) if t]
        if terms:
            combined = None
            for t in terms:
                term_q = Q(search_text__icontains=t.lower()) | Q(specialty__icontains=t.lower())
                combined = term_q if combined is None else (combined & term_q)
            qs = qs.filter(combined)

    records = qs.order_by("-last_snapshot_datetime", "specialty", "excel_row")[:limit]
    out = []
    for r in records:
        out.append(
            {
                "id": r.id,
                "specialty": r.specialty,
                "excel_row": r.excel_row,
                "columns": r.columns_json or {},
                "column_order": r.column_headers_order or [],
                "last_snapshot_datetime": (
                    r.last_snapshot_datetime.isoformat() if r.last_snapshot_datetime else None
                ),
                "last_diff_fields": r.last_diff_fields,
                "last_import_id": r.last_import_id,
            }
        )
    return out


def ordered_column_keys_for_planos_iniciales_export(rows: list[dict]) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()
    for r in rows:
        for k in r.get("column_order") or []:
            if k not in seen:
                seen.add(k)
                keys.append(k)
    for r in rows:
        for k in (r.get("columns") or {}).keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    return keys


_PLAN_CODE_RE = re.compile(r"\b[A-Z]{2,5}-\d{3,5}-[A-Z0-9-]+\b")
_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _normalize_plan_code(value: str) -> str:
    s = _norm_str(value).upper()
    if not s:
        return ""
    if "." in s:
        s = s.rsplit(".", 1)[0]
    return s.strip()


def _extract_plan_code_from_iniciales_row(columns: dict) -> str:
    if not columns:
        return ""
    priority_keys = ("codigo", "código", "plano", "nombre", "documento")
    for k, v in columns.items():
        kl = _norm_str(k).lower()
        if any(p in kl for p in priority_keys):
            cand = _normalize_plan_code(v)
            if _PLAN_CODE_RE.search(cand):
                return _PLAN_CODE_RE.search(cand).group(0)
    for v in columns.values():
        cand = _normalize_plan_code(v)
        m = _PLAN_CODE_RE.search(cand)
        if m:
            return m.group(0)
    return ""


def _parse_any_date_text(value: str):
    s = _norm_str(value)
    if not s:
        return None
    if _ISO_DATE_RE.match(s):
        try:
            y, m, d = s.split("-")
            return date(int(y), int(m), int(d))
        except Exception:
            return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.date()
        except Exception:
            pass
    return None


def _latest_date_from_iniciales_row(columns: dict):
    if not columns:
        return None
    best = None
    for v in columns.values():
        d = _parse_any_date_text(v)
        if d and (best is None or d > best):
            best = d
    return best


def _latest_date_key_in_row(columns: dict, column_order: list | None) -> str | None:
    """Clave de columna donde está la fecha máxima (orden de hoja para desempatar a la derecha)."""
    if not columns:
        return None
    order = column_order if column_order else list(columns.keys())
    best = None
    best_key = None
    best_idx = -1
    for idx, k in enumerate(order):
        v = columns.get(k)
        d = _parse_any_date_text(v)
        if d is None:
            continue
        if best is None or d > best or (d == best and idx > best_idx):
            best = d
            best_key = k
            best_idx = idx
    return best_key


def _cell_before_date_key(column_order: list | None, columns: dict, date_key: str | None) -> str:
    """Valor en la columna inmediatamente anterior a la de la fecha máxima (p. ej. última rev.)."""
    if not date_key or not column_order:
        return ""
    try:
        idx = column_order.index(date_key)
    except ValueError:
        return ""
    if idx <= 0:
        return ""
    prev_k = column_order[idx - 1]
    return _norm_str(columns.get(prev_k, ""))


def _norm_rev_label(s: str) -> str:
    t = _norm_str(s)
    if not t:
        return ""
    if t.isdigit():
        return f"r{t}"
    tl = t.lower().replace(" ", "")
    if re.match(r"^r\d+$", tl):
        return tl
    if len(t) <= 6:
        return f"r{t}" if not tl.startswith("r") else tl
    return f"r{t[:6]}"


def _norm_planos_revision(s: str) -> str:
    """Revisión en listado Planos → r3, rC, etc."""
    t = _norm_str(s)
    if not t:
        return ""
    m = re.search(r"(\d+)", t)
    if m:
        return "r" + m.group(1)
    if len(t) == 1 and t.isalpha():
        return "r" + t.upper()
    tl = t.lower().replace(" ", "")
    if re.match(r"^r\d+$", tl):
        return tl
    return "r" + t[:4] if len(t) <= 4 else t[:6]


def _extract_version_from_iniciales_row(columns: dict) -> str:
    if not columns:
        return ""
    for key in ("Versión", "Version", "versión", "version", "REV", "Rev", "Revisión", "Revision"):
        if key in columns and _norm_str(columns[key]):
            return _norm_str(columns[key])
    for k, v in columns.items():
        kl = _norm_str(k).lower()
        if ("versión" in kl or "version" in kl or kl in ("rev", "revisión", "revision")) and _norm_str(v):
            return _norm_str(v)
    return ""


def _normalize_version_label(s: str) -> str:
    t = _norm_str(s)
    if not t:
        return "—"
    u = t.upper()
    m = re.match(r"^V\s*(\d+)$", u)
    if m:
        return "v" + m.group(1)
    m = re.match(r"^(\d+)$", u)
    if m:
        return "v" + m.group(1)
    if len(t) <= 12:
        return t
    return t[:10] + "…"


def _specialty_hint_from_plan_code(code: str) -> str:
    m = re.match(r"^[A-Z]+-\d+-([A-Z]{2,5})-", code.upper())
    if m:
        return m.group(1).lower()
    return ""


def _filter_planos_actualizados_rows(rows: list[dict], q: str = "", specialty: str = "") -> list[dict]:
    q = (q or "").strip().lower()
    sp = (specialty or "").strip().lower()
    out = rows
    if sp:
        out = [r for r in out if (r.get("specialty") or "").lower() == sp]
    if q:
        terms = [t for t in re.split(r"\s+", q) if t]
        if terms:
            for t in terms:
                out = [
                    r
                    for r in out
                    if t in (r.get("code") or "").lower()
                    or t in (r.get("folder_path") or "").lower()
                    or t in (r.get("version") or "").lower()
                    or t in (r.get("revision") or "").lower()
                    or t in (r.get("version_transition") or "").lower()
                    or t in (r.get("version_matriz") or "").lower()
                    or t in (r.get("version_planos") or "").lower()
                    or t in (r.get("iniciales_version") or "").lower()
                    or t in (r.get("iniciales_rev_raw") or "").lower()
                    or t in (r.get("specialty") or "").lower()
                ]
    return out


def get_planos_updated_vs_iniciales(q: str = "", specialty: str = "", limit: int = 500):
    """
    Compara Planos.last_update_at vs última fecha detectada en PlanosIniciales
    por código de plano. Devuelve sólo diferencias, con versión en matriz vs Planos.
    """
    iniciales_rows = PlanosInicialesRecord.objects.all().only(
        "columns_json",
        "column_headers_order",
        "specialty",
        "excel_row",
        "last_snapshot_datetime",
    )
    latest_by_code: dict[str, dict] = {}

    for row in iniciales_rows:
        cols = row.columns_json or {}
        code = _extract_plan_code_from_iniciales_row(cols)
        if not code:
            continue
        d = _latest_date_from_iniciales_row(cols)
        if not d:
            continue
        order = list(row.column_headers_order or []) or list(cols.keys())
        date_key = _latest_date_key_in_row(cols, order)
        ini_rev_before_date = _cell_before_date_key(order, cols, date_key)
        ini_ver = _extract_version_from_iniciales_row(cols)
        snap = row.last_snapshot_datetime
        cur = latest_by_code.get(code)
        better = False
        if cur is None:
            better = True
        elif d > cur["date"]:
            better = True
        elif d == cur["date"] and snap and cur.get("snapshot"):
            if snap > cur["snapshot"]:
                better = True
        elif d == cur["date"] and snap and not cur.get("snapshot"):
            better = True
        if better:
            latest_by_code[code] = {
                "date": d,
                "specialty": (row.specialty or "").lower(),
                "ini_version": ini_ver,
                "ini_rev_before_date": ini_rev_before_date,
                "snapshot": snap,
            }

    out = []
    planos_qs = PlanosRecord.objects.all().only(
        "name", "last_update_at", "folder_path", "version", "revision"
    )
    for p in planos_qs:
        code = _normalize_plan_code(p.name)
        if not code:
            continue
        meta = latest_by_code.get(code)
        if not meta:
            continue
        ini_date = meta["date"]
        spec = meta.get("specialty") or ""
        if not spec:
            spec = _specialty_hint_from_plan_code(code)
        plano_date = p.last_update_at.date() if p.last_update_at else None
        if plano_date == ini_date:
            continue
        plan_ver_raw = _norm_str(p.version)
        plan_rev_raw = _norm_str(p.revision)
        ini_ver_raw = _norm_str(meta.get("ini_version") or "")
        ini_rev_cell_raw = _norm_str(meta.get("ini_rev_before_date") or "")

        rev_m = _norm_rev_label(ini_rev_cell_raw)
        v_m_label = _normalize_version_label(ini_ver_raw)
        if v_m_label == "—":
            v_m_label = ""
        matriz_parts = []
        if rev_m:
            matriz_parts.append(rev_m)
        if v_m_label:
            matriz_parts.append(v_m_label)
        elif ini_ver_raw and not rev_m:
            matriz_parts.append(ini_ver_raw[:20])
        version_matriz = " ".join(matriz_parts) if matriz_parts else "—"

        rev_p = _norm_planos_revision(plan_rev_raw)
        v_p_label = _normalize_version_label(plan_ver_raw)
        if v_p_label == "—":
            v_p_label = ""
        planos_parts = []
        if v_p_label:
            planos_parts.append(v_p_label)
        elif plan_ver_raw:
            planos_parts.append(plan_ver_raw)
        if rev_p:
            planos_parts.append(rev_p)
        version_planos = " ".join(planos_parts) if planos_parts else "—"

        version_transition = f"{version_matriz} → {version_planos}"
        out.append(
            {
                "code": code,
                "specialty": spec,
                "planos_last_update": plano_date.isoformat() if plano_date else "",
                "iniciales_last_date": ini_date.isoformat(),
                "iniciales_version": ini_ver_raw,
                "iniciales_rev_raw": ini_rev_cell_raw,
                "folder_path": p.folder_path or "",
                "version": plan_ver_raw,
                "revision": plan_rev_raw,
                "version_matriz": version_matriz,
                "version_planos": version_planos,
                "version_transition": version_transition,
            }
        )

    out.sort(key=lambda r: (r.get("specialty") or "", r["code"]))
    out = _filter_planos_actualizados_rows(out, q=q, specialty=specialty)
    return out[:limit]


from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_GET, require_POST

from .models import PLANOS_INICIALES_SHEET_SLUGS
from .services import (
    attach_planos_iniciales,
    attach_planos_xlsx,
    attach_rdi_csv,
    get_planos_updated_vs_iniciales,
    get_planos_iniciales_records_for_ajax,
    get_planos_records_for_ajax,
    get_rdi_cost_schedule_impacts_for_ajax,
    get_rdi_records_for_ajax,
    ordered_column_keys_for_planos_iniciales_export,
)


def _date_short(value):
    # Espera ISO "YYYY-MM-DD..." (o None)
    if not value:
        return ""
    return str(value)[:10]


def _escape_html_for_paragraph(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    # En ReportLab, una fila de Table no se parte entre páginas.
    # Quitamos saltos para evitar celdas con altura extrema.
    s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    # Colapsa espacios múltiples.
    s = " ".join(s.split())
    return s


@login_required
@require_GET
def rdi_export_excel(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    q = request.GET.get("q", "").strip()
    # Debe coincidir con lo que muestra la tabla (AJAX usa limit=200 por defecto).
    records = get_rdi_records_for_ajax(q=q, limit=200)

    headers = [
        "CSV ID",
        "Fecha de archivo (versión)",
        "Titulo",
        "Estado",
        "Consulta",
        "Respuesta",
        "Fecha vencimiento",
        "Fecha creación",
        "Fecha actualización",
        "Disciplina",
        "Empresa",
        "Asignado a",
        "Tipo asignación",
        "Respuesta sugerida",
        "Ubicación detalle",
        "Creado por",
        "Actualizado por",
        "Lista distribución",
        "Impacto costo",
        "Impacto plazo",
        "Prioridad",
        "Categoría",
        "Referencia",
        "Asociado a documento",
        "Informado (código)",
        "Informado",
        "ID interno",
        "Last diff fields",
        "Last import ID",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RDI"

    from django.utils import timezone
    fecha_str = timezone.now().strftime("%d-%m-%Y")
    filtro_str = (f"Filtro: {q}" if q else "Filtro: (vacío)")
    ws["A1"] = filtro_str + f" | Generado: {fecha_str}"
    # Fusiona la fila del filtro sobre todas las columnas exportadas.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Fila 2: headers
    ws.append(headers)

    # Datos desde fila 3
    for r in records:
        ws.append([
            r.get("csv_id", ""),
            _date_short(r.get("last_snapshot_datetime")),
            r.get("title", ""),
            r.get("status_label", r.get("status", "")),
            r.get("question", ""),
            r.get("response", ""),
            _date_short(r.get("due_date")),
            _date_short(r.get("created_at")),
            _date_short(r.get("updated_at")),
            r.get("discipline", ""),
            r.get("company", ""),
            r.get("assigned_to", ""),
            r.get("assignee_type", ""),
            r.get("suggested_answer", ""),
            r.get("location_details", ""),
            r.get("created_by", ""),
            r.get("updated_by", ""),
            r.get("distribution_list", ""),
            r.get("cost_impact", ""),
            r.get("schedule_impact", ""),
            r.get("priority", ""),
            r.get("category", ""),
            r.get("reference", ""),
            "Sí" if r.get("associated_to_document") is True else ("No" if r.get("associated_to_document") is False else ""),
            r.get("informado", ""),
            r.get("informado_label", ""),
            r.get("id", ""),
            r.get("last_diff_fields", ""),
            r.get("last_import_id", ""),
        ])

    # Estilos y ancho simple + wrap
    wrap_text = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        # Anchos aproximados
        ws.column_dimensions[letter].width = 18 if col in (1, 2, 4, 7, 8, 9, 24, 25, 26, 27, 28, 29) else 28

    # Alinea y activa wrap en todas las filas excepto A1 (fusionado)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap_text

    # Congelar panes: mantener headers visibles
    ws.freeze_panes = "A3"

    # Filtros automáticos en la fila de encabezados (fila 2)
    last_col_letter = get_column_letter(len(headers))
    n_data = len(records) + 2
    ws.auto_filter.ref = f"A2:{last_col_letter}{n_data}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"RDI_export-{fecha_str}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_GET
def rdi_export_pdf(request):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    q = request.GET.get("q", "").strip()
    # Para evitar LayoutError (filas demasiado altas), limitamos.
    records = get_rdi_records_for_ajax(q=q, limit=200)

    headers = ["CSV ID", "Titulo", "Estado", "Consulta", "Respuesta", "Fecha vencimiento", "Fecha creación", "Fecha actualización"]

    styles = getSampleStyleSheet()
    style_title = styles["Title"]
    style_body = styles["BodyText"]
    style_cell = ParagraphStyle(
        "cell",
        parent=style_body,
        fontSize=9,
        leading=11,
        wordWrap="CJK",
    )
    style_cell_bold = ParagraphStyle(
        "cell_bold",
        parent=style_body,
        fontSize=9,
        leading=11,
        textColor=colors.white,
    )

    def cell_paragraph(txt, bold=False, max_chars=600):
        if txt is None:
            txt = ""
        txt = str(txt)
        if len(txt) > max_chars:
            txt = txt[:max_chars] + " …"
        txt = _escape_html_for_paragraph(txt)
        return Paragraph(txt, style_cell_bold if bold else style_cell)

    pdf_buf = BytesIO()
    doc = SimpleDocTemplate(pdf_buf, pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)

    title = Paragraph("RDI - Exportación", style_title)
    filtro = Paragraph(f"Filtro: {q}" if q else "Filtro: (vacío)", style_body)

    data = []
    header_row = [cell_paragraph(h, bold=True, max_chars=200) for h in headers]
    data.append(header_row)

    for r in records:
        data.append([
            cell_paragraph(r.get("csv_id", "")),
            cell_paragraph(r.get("title", ""), max_chars=220),
            cell_paragraph(r.get("status_label", r.get("status", "")), max_chars=60),
            cell_paragraph(r.get("question", ""), max_chars=450),
            cell_paragraph(r.get("response", ""), max_chars=450),
            cell_paragraph(_date_short(r.get("due_date", ""))),
            cell_paragraph(_date_short(r.get("created_at", ""))),
            cell_paragraph(_date_short(r.get("updated_at", ""))),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5090")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ]
        )
    )

    doc.build([title, Spacer(1, 10), filtro, Spacer(1, 12), table])
    pdf_buf.seek(0)

    resp = HttpResponse(pdf_buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="RDI_export.pdf"'
    return resp


@login_required
@require_GET
def rdi_list_view(request):
    return render(request, "rdi/rdi_list.html")


@login_required
@require_GET
def informar_bim_list_view(request):
    """Listado tipo RDI para «Informar desde BIM» (misma fuente AJAX que RDI)."""
    return render(request, "rdi/informar_bim_list.html")


@login_required
@require_GET
def rdi_increments_decrements_view(request):
    """
    Listado de RDI con impacto en costo o plazo (yes/si).
    """
    return render(request, "rdi/rdi_increments_decrements.html")


@login_required
@require_GET
def rdi_increments_decrements_export_excel(request):
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from django.utils import timezone

    q = request.GET.get("q", "").strip()
    records = get_rdi_cost_schedule_impacts_for_ajax(q=q, limit=1000)
    headers = [
        "CSV ID",
        "Estado",
        "Ubicacion",
        "Impacto de costo",
        "Impacto de plazo",
        "Disciplina",
        "Prioridad",
        "Pregunta",
        "Respuesta",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aumentos-Disminuciones"
    fecha_str = timezone.now().strftime("%d-%m-%Y")
    filtro_str = (f"Filtro: {q}" if q else "Filtro: (vacío)")
    ws["A1"] = filtro_str + f" | Generado: {fecha_str}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append(headers)

    for r in records:
        ws.append(
            [
                r.get("csv_id", ""),
                r.get("status_label", r.get("status", "")),
                r.get("location_details", ""),
                r.get("cost_impact_label", r.get("cost_impact", "")),
                r.get("schedule_impact_label", r.get("schedule_impact", "")),
                r.get("discipline", ""),
                r.get("priority_label", r.get("priority", "")),
                r.get("question", ""),
                r.get("response", ""),
            ]
        )

    wrap_text = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18 if col in (1, 2, 4, 5, 7) else 38

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap_text

    ws.freeze_panes = "A3"
    last_col_letter = get_column_letter(len(headers))
    n_data = len(records) + 2
    ws.auto_filter.ref = f"A2:{last_col_letter}{n_data}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Aumentos-disminuciones-{fecha_str}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_GET
def rdi_increments_decrements_export_pdf(request):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    q = request.GET.get("q", "").strip()
    records = get_rdi_cost_schedule_impacts_for_ajax(q=q, limit=600)
    headers = [
        "CSV ID",
        "Estado",
        "Ubicacion",
        "Impacto costo",
        "Impacto plazo",
        "Disciplina",
        "Prioridad",
    ]

    styles = getSampleStyleSheet()
    style_title = styles["Title"]
    style_body = styles["BodyText"]
    style_cell = ParagraphStyle("cell", parent=style_body, fontSize=8.5, leading=10.5, wordWrap="CJK")
    style_cell_bold = ParagraphStyle("cell_bold", parent=style_body, fontSize=8.5, leading=10.5, textColor=colors.white)

    def cell_paragraph(txt, bold=False, max_chars=500):
        if txt is None:
            txt = ""
        txt = str(txt)
        if len(txt) > max_chars:
            txt = txt[:max_chars] + " …"
        txt = _escape_html_for_paragraph(txt)
        return Paragraph(txt, style_cell_bold if bold else style_cell)

    pdf_buf = BytesIO()
    doc = SimpleDocTemplate(
        pdf_buf,
        pagesize=landscape(A4),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    title = Paragraph("Aumentos/disminuciones - RDI", style_title)
    filtro = Paragraph(f"Filtro: {q}" if q else "Filtro: (vacío)", style_body)
    data = [[cell_paragraph(h, bold=True, max_chars=180) for h in headers]]
    for r in records:
        data.append(
            [
                cell_paragraph(r.get("csv_id", "")),
                cell_paragraph(r.get("status_label", r.get("status", "")), max_chars=60),
                cell_paragraph(r.get("location_details", ""), max_chars=280),
                cell_paragraph(r.get("cost_impact_label", r.get("cost_impact", "")), max_chars=60),
                cell_paragraph(r.get("schedule_impact_label", r.get("schedule_impact", "")), max_chars=60),
                cell_paragraph(r.get("discipline", ""), max_chars=90),
                cell_paragraph(r.get("priority_label", r.get("priority", "")), max_chars=60),
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5090")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ]
        )
    )

    doc.build([title, Spacer(1, 8), filtro, Spacer(1, 10), table])
    pdf_buf.seek(0)
    resp = HttpResponse(pdf_buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="Aumentos-disminuciones-RDI.pdf"'
    return resp


@login_required
@require_POST
def rdi_import_view(request):
    uploaded = request.FILES.get("file")
    if not uploaded:
        messages.error(request, "No se recibió ningún archivo CSV.")
        return redirect("rdi_list")

    original_filename = uploaded.name

    try:
        attach_rdi_csv(uploaded, original_filename=original_filename)
        messages.success(
            request,
            f"CSV importado correctamente: {original_filename}.",
        )
    except Exception as e:
        messages.error(request, f"Error importando CSV: {e}")

    return redirect("rdi_list")


@login_required
@require_GET
def rdi_records_json(request):
    q = request.GET.get("q", "").strip()
    data = get_rdi_records_for_ajax(q=q)
    return JsonResponse({"records": data})


@login_required
@require_GET
def rdi_increments_decrements_json(request):
    q = request.GET.get("q", "").strip()
    data = get_rdi_cost_schedule_impacts_for_ajax(q=q)
    return JsonResponse({"records": data})


@login_required
@require_GET
def planos_list_view(request):
    return render(request, "rdi/planos_list.html")


@login_required
@require_POST
def planos_import_view(request):
    uploaded = request.FILES.get("file")
    if not uploaded:
        messages.error(request, "No se recibió ningún archivo XLSX.")
        return redirect("planos_list")

    original_filename = uploaded.name
    if not original_filename.lower().endswith(".xlsx"):
        messages.error(request, "El archivo debe ser .xlsx")
        return redirect("planos_list")

    try:
        attach_planos_xlsx(uploaded, original_filename=original_filename)
        messages.success(request, f"Planos importados correctamente: {original_filename}.")
    except Exception as e:
        messages.error(request, f"Error importando XLSX: {e}")

    return redirect("planos_list")


@login_required
@require_GET
def planos_records_json(request):
    q = request.GET.get("q", "").strip()
    data = get_planos_records_for_ajax(q=q)
    return JsonResponse({"records": data})


@login_required
@require_GET
def planos_export_excel(request):
    import openpyxl
    from io import BytesIO
    from django.utils import timezone
    from openpyxl.utils import get_column_letter

    q = request.GET.get("q", "").strip()
    records = get_planos_records_for_ajax(q=q, limit=2000)
    headers = [
        "Ruta y nombre de la carpeta",
        "Nombre",
        "Descripción",
        "Versión",
        "Tamaño",
        "Última actualización",
        "Actualizado por",
        "Última carga",
        "Cargado por",
        "Marca de revisión",
        "Incidencia",
        "SDI",
        "Estado de revisión",
        "Conjunto",
        "Fecha de emisión",
        "Sheet number",
        "Title",
        "Revisión",
        "Fecha snapshot",
        "ID interno",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planos"

    fecha_str = timezone.now().strftime("%d-%m-%Y")
    ws["A1"] = (f"Filtro: {q}" if q else "Filtro: (vacío)") + f" | Generado: {fecha_str}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append(headers)

    for r in records:
        ws.append(
            [
                r.get("folder_path", ""),
                r.get("name", ""),
                r.get("description", ""),
                r.get("version", ""),
                r.get("size", ""),
                r.get("last_update_raw", ""),
                r.get("updated_by", ""),
                r.get("last_upload_raw", ""),
                r.get("uploaded_by", ""),
                r.get("review_mark", ""),
                r.get("incidence", ""),
                r.get("sdi", ""),
                r.get("review_status", ""),
                r.get("set_name", ""),
                r.get("issue_date_raw", ""),
                r.get("sheet_number", ""),
                r.get("title", ""),
                r.get("revision", ""),
                _date_short(r.get("last_snapshot_datetime")),
                r.get("id", ""),
            ]
        )

    wrap_text = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 26 if col in (1, 2, 3, 17) else 18
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap_text
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{len(records) + 2}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Planos_export-{fecha_str}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_GET
def planos_export_pdf(request):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    q = request.GET.get("q", "").strip()
    records = get_planos_records_for_ajax(q=q, limit=600)
    headers = [
        "Nombre",
        "Versión",
        "Ruta",
        "Última actualización",
        "Actualizado por",
        "Revisión",
    ]

    styles = getSampleStyleSheet()
    style_title = styles["Title"]
    style_body = styles["BodyText"]
    style_cell = ParagraphStyle("cell", parent=style_body, fontSize=8.5, leading=10.5, wordWrap="CJK")
    style_cell_bold = ParagraphStyle("cell_bold", parent=style_body, fontSize=8.5, leading=10.5, textColor=colors.white)

    def cell_paragraph(txt, bold=False, max_chars=450):
        txt = _escape_html_for_paragraph("" if txt is None else str(txt))
        if len(txt) > max_chars:
            txt = txt[:max_chars] + " …"
        return Paragraph(txt, style_cell_bold if bold else style_cell)

    pdf_buf = BytesIO()
    doc = SimpleDocTemplate(
        pdf_buf,
        pagesize=landscape(A4),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    title = Paragraph("Planos - Exportación", style_title)
    filtro = Paragraph(f"Filtro: {q}" if q else "Filtro: (vacío)", style_body)
    data = [[cell_paragraph(h, bold=True, max_chars=120) for h in headers]]

    for r in records:
        data.append(
            [
                cell_paragraph(r.get("name", ""), max_chars=120),
                cell_paragraph(r.get("version", ""), max_chars=50),
                cell_paragraph(r.get("folder_path", ""), max_chars=180),
                cell_paragraph(r.get("last_update_raw", ""), max_chars=80),
                cell_paragraph(r.get("updated_by", ""), max_chars=90),
                cell_paragraph(r.get("revision", ""), max_chars=80),
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5090")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ]
        )
    )

    doc.build([title, Spacer(1, 8), filtro, Spacer(1, 10), table])
    pdf_buf.seek(0)
    resp = HttpResponse(pdf_buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="Planos_export.pdf"'
    return resp


@login_required
@require_GET
def planos_iniciales_list_view(request):
    return render(request, "rdi/planos_iniciales_list.html")


@login_required
@require_POST
def planos_iniciales_import_view(request):
    uploaded = request.FILES.get("file")
    if not uploaded:
        messages.error(request, "No se recibió ningún archivo.")
        return redirect("planos_iniciales_list")

    original_filename = uploaded.name
    lower = original_filename.lower()
    if not (lower.endswith(".xls") or lower.endswith(".xlsx")):
        messages.error(request, "El archivo debe ser .xls o .xlsx (planos iniciales).")
        return redirect("planos_iniciales_list")

    try:
        attach_planos_iniciales(uploaded, original_filename=original_filename)
        messages.success(
            request,
            f"Planos iniciales importados: {original_filename}.",
        )
    except Exception as e:
        messages.error(request, f"Error importando: {e}")

    return redirect("planos_iniciales_list")


@login_required
@require_GET
def planos_iniciales_records_json(request):
    q = request.GET.get("q", "").strip()
    data = get_planos_iniciales_records_for_ajax(q=q)
    return JsonResponse({"records": data})


@login_required
@require_GET
def planos_iniciales_export_excel(request):
    import openpyxl
    from io import BytesIO
    from django.utils import timezone
    from openpyxl.utils import get_column_letter

    q = request.GET.get("q", "").strip()
    records = get_planos_iniciales_records_for_ajax(q=q, limit=5000)
    col_keys = ordered_column_keys_for_planos_iniciales_export(records)
    headers = ["Especialidad", "Fila Excel"] + col_keys + ["Fecha snapshot", "ID interno"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planos iniciales"

    fecha_str = timezone.now().strftime("%d-%m-%Y")
    ws["A1"] = (f"Filtro: {q}" if q else "Filtro: (vacío)") + f" | Generado: {fecha_str}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append(headers)

    for r in records:
        cols = r.get("columns") or {}
        row_vals = [r.get("specialty", ""), r.get("excel_row", "")]
        row_vals += [cols.get(k, "") for k in col_keys]
        row_vals.append(_date_short(r.get("last_snapshot_datetime")))
        row_vals.append(r.get("id", ""))
        ws.append(row_vals)

    wrap_text = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 14 if col <= 2 else 22
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap_text
    ws.freeze_panes = "A3"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A2:{last_col}{len(records) + 2}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Planos_iniciales_export-{fecha_str}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_GET
def planos_iniciales_export_pdf(request):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    q = request.GET.get("q", "").strip()
    records = get_planos_iniciales_records_for_ajax(q=q, limit=500)
    headers = ["Especialidad", "Fila", "Resumen (primeras columnas)"]

    styles = getSampleStyleSheet()
    style_title = styles["Title"]
    style_body = styles["BodyText"]
    style_cell = ParagraphStyle("cell", parent=style_body, fontSize=8.5, leading=10.5, wordWrap="CJK")
    style_cell_bold = ParagraphStyle(
        "cell_bold", parent=style_body, fontSize=8.5, leading=10.5, textColor=colors.white
    )

    def cell_paragraph(txt, bold=False, max_chars=500):
        txt = _escape_html_for_paragraph("" if txt is None else str(txt))
        if len(txt) > max_chars:
            txt = txt[:max_chars] + " …"
        return Paragraph(txt, style_cell_bold if bold else style_cell)

    def row_summary(r):
        order = r.get("column_order") or []
        cols = r.get("columns") or {}
        parts = []
        for k in order[:5]:
            if k in cols and cols[k]:
                parts.append(f"{k}: {cols[k]}")
        return " | ".join(parts)

    pdf_buf = BytesIO()
    doc = SimpleDocTemplate(
        pdf_buf,
        pagesize=landscape(A4),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    title = Paragraph("Planos iniciales - Exportación", style_title)
    filtro = Paragraph(f"Filtro: {q}" if q else "Filtro: (vacío)", style_body)
    data = [[cell_paragraph(h, bold=True, max_chars=120) for h in headers]]

    for r in records:
        data.append(
            [
                cell_paragraph(r.get("specialty", ""), max_chars=40),
                cell_paragraph(r.get("excel_row", ""), max_chars=20),
                cell_paragraph(row_summary(r), max_chars=400),
            ]
        )

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5090")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ]
        )
    )

    doc.build([title, Spacer(1, 8), filtro, Spacer(1, 10), table])
    pdf_buf.seek(0)
    resp = HttpResponse(pdf_buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="Planos_iniciales_export.pdf"'
    return resp


@login_required
@require_GET
def planos_actualizados_view(request):
    return render(
        request,
        "rdi/planos_actualizados_list.html",
        {
            "specialty_slugs": PLANOS_INICIALES_SHEET_SLUGS,
        },
    )


@login_required
@require_GET
def planos_actualizados_json(request):
    q = request.GET.get("q", "").strip()
    specialty = request.GET.get("specialty", "").strip()
    rows = get_planos_updated_vs_iniciales(q=q, specialty=specialty, limit=800)
    return JsonResponse({"records": rows, "count": len(rows)})


@login_required
@require_GET
def planos_actualizados_export_excel(request):
    import openpyxl
    from io import BytesIO
    from django.utils import timezone
    from openpyxl.utils import get_column_letter

    q = request.GET.get("q", "").strip()
    specialty = request.GET.get("specialty", "").strip()
    rows = get_planos_updated_vs_iniciales(q=q, specialty=specialty, limit=5000)
    headers = [
        "Especialidad",
        "Plano",
        "Versión matriz (iniciales)",
        "Versión Planos",
        "Fecha Planos",
        "Fecha matriz (iniciales)",
        "Ruta",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planos actualizados"
    fecha_str = timezone.now().strftime("%d-%m-%Y")
    filtros = []
    if specialty:
        filtros.append(f"Especialidad={specialty}")
    if q:
        filtros.append(f"Buscar={q}")
    ws["A1"] = (" | ".join(filtros) if filtros else "Sin filtro") + f" | Generado: {fecha_str}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r.get("specialty", ""),
                r.get("code", ""),
                r.get("version_matriz", ""),
                r.get("version_planos", ""),
                r.get("planos_last_update", ""),
                r.get("iniciales_last_date", ""),
                r.get("folder_path", ""),
            ]
        )
    wrap_text = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16 if col != 7 else 36
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = wrap_text
    ws.freeze_panes = "A3"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A2:{last_col}{len(rows) + 2}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Planos_actualizados-{fecha_str}.xlsx"
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
@require_GET
def planos_actualizados_export_pdf(request):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    q = request.GET.get("q", "").strip()
    specialty = request.GET.get("specialty", "").strip()
    rows = get_planos_updated_vs_iniciales(q=q, specialty=specialty, limit=400)
    headers = ["Esp.", "Plano", "Matriz / Planos", "F. Planos", "F. matriz", "Ruta"]

    styles = getSampleStyleSheet()
    style_title = styles["Title"]
    style_body = styles["BodyText"]
    style_cell = ParagraphStyle("cell", parent=style_body, fontSize=8, leading=10, wordWrap="CJK")
    style_cell_bold = ParagraphStyle(
        "cell_bold", parent=style_body, fontSize=8, leading=10, textColor=colors.white
    )

    def cell_paragraph(txt, bold=False, max_chars=320):
        txt = _escape_html_for_paragraph("" if txt is None else str(txt))
        if len(txt) > max_chars:
            txt = txt[:max_chars] + " …"
        return Paragraph(txt, style_cell_bold if bold else style_cell)

    pdf_buf = BytesIO()
    doc = SimpleDocTemplate(
        pdf_buf,
        pagesize=landscape(A4),
        rightMargin=16,
        leftMargin=16,
        topMargin=16,
        bottomMargin=16,
    )
    title = Paragraph("Planos actualizados", style_title)
    filtro_parts = []
    if specialty:
        filtro_parts.append(f"Especialidad: {specialty}")
    if q:
        filtro_parts.append(f"Búsqueda: {q}")
    filtro = Paragraph(" | ".join(filtro_parts) if filtro_parts else "Sin filtros", style_body)
    data = [[cell_paragraph(h, bold=True, max_chars=80) for h in headers]]
    for r in rows:
        data.append(
            [
                cell_paragraph(r.get("specialty", ""), max_chars=20),
                cell_paragraph(r.get("code", ""), max_chars=120),
                cell_paragraph(
                    f"Mat.: {r.get('version_matriz', '')} | Plan.: {r.get('version_planos', '')}",
                    max_chars=56,
                ),
                cell_paragraph(r.get("planos_last_update", ""), max_chars=24),
                cell_paragraph(r.get("iniciales_last_date", ""), max_chars=24),
                cell_paragraph(r.get("folder_path", ""), max_chars=200),
            ]
        )
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5090")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    doc.build([title, Spacer(1, 6), filtro, Spacer(1, 10), table])
    pdf_buf.seek(0)
    resp = HttpResponse(pdf_buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="Planos_actualizados.pdf"'
    return resp


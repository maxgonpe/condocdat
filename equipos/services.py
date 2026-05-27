from __future__ import annotations

import io
import re
import shutil
import zipfile
from datetime import date, datetime
from xml.sax.saxutils import escape
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from django.core.exceptions import FieldDoesNotExist
from django.core.files.storage import default_storage
from django.db import models as dj_models
from django.db import transaction
from django.db.models import Max, Q
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel as xl_from_excel
from openpyxl.utils.datetime import to_excel as xl_to_excel

from .models import (
    EQUIPOS_DOWNLOAD_BASE_STEM,
    EQUIPOS_LIBRO_REL_PATH,
    EQUIPOS_MASTER_DIR_REL,
    EQUIPOS_WORKBOOK_FILENAME,
    EquiposAsset,
    EquiposCambioLog,
    EquiposLibro,
    EquiposLocation,
    EquiposOtro,
    EquiposResumenFila,
    EquiposSignificadoFila,
)

SHEET_RESUMEN = "Resumen - TD"
SHEET_SIGNIFICADO = "Significado status"
SHEET_LOCATIONS = "Locations"
SHEET_ASSET = "Asset"
SHEET_OTROS = "Otros equipos"

# «Resumen - TD» filas 6–12: la columna A puede editarse vía formularios; B/C deben quedar como
# en el Excel (cabeceras + fórmulas SUMPRODUCT / IFERROR / SUM enlazadas a filtros B3:B4 y Asset).
# Si se escriben cuenta/fracción desde la BD (valores congelados del import data_only), Excel pierde
# las fórmulas y el gráfico deja de ser dinámico.
RESUMEN_EXCEL_ROWS_PRESERVE_BC = frozenset(range(6, 13))

# Copias nombradas por semana (p. ej. SEMANA 14) bajo MEDIA_ROOT
EQUIPOS_SEMANALES_REL_DIR = "equipos/semanales"
# Importaciones antiguas (antes de centralizar en 2026/04).
LEGACY_EQUIPOS_LIBRO_REL_PATH = "equipos/libro_actual.xlsx"
# Libro de referencia con chart pivot + título intactos (p. ej. export previo a sync).
# Si existe, se reutiliza el cuerpo de xl/charts/chart1.xml y luego se corrige el nombre del libro.
EQUIPOS_CHART_REFERENCE_REL = (
    "equipos/2026/04/ST01-EXP_F5-E2_Control_de_equipos_2026-10-04-antes.xlsx"
)
CHART_PART_PATH = "xl/charts/chart1.xml"


def ensure_equipos_workbook_parent_dir() -> None:
    """Crea ``media/equipos/2026/04`` para poder guardar el maestro sin Errno 2."""
    from django.conf import settings

    (Path(settings.MEDIA_ROOT) / EQUIPOS_MASTER_DIR_REL).mkdir(parents=True, exist_ok=True)


def resolve_equipos_master_xlsx_path() -> Path | None:
    """
    Ruta absoluta al .xlsx maestro en MEDIA (el mismo fichero que ``EQUIPOS_LIBRO_REL_PATH``).

    Orden: ``settings.EQUIPOS_MASTER_XLSX`` si existe; si no, el workbook canónico bajo
    ``equipos/2026/04/``; si aún no existe, el primer ``*.xlsx`` de esa carpeta (migración).
    """
    from django.conf import settings

    custom = getattr(settings, "EQUIPOS_MASTER_XLSX", None)
    if custom:
        p = Path(custom)
        if not p.is_absolute():
            cand = Path(settings.MEDIA_ROOT) / custom
            if cand.is_file():
                return cand.resolve()
            p = Path(settings.BASE_DIR) / custom
        if p.is_file():
            return p.resolve()

    canonical = Path(settings.MEDIA_ROOT) / EQUIPOS_LIBRO_REL_PATH
    if canonical.is_file():
        return canonical.resolve()

    d = Path(settings.MEDIA_ROOT) / EQUIPOS_MASTER_DIR_REL
    if not d.is_dir():
        return None
    names = sorted(d.glob("*.xlsx"))
    if not names:
        return None
    non_antes = [x for x in names if "antes" not in x.name.lower()]
    return (non_antes[0] if non_antes else names[0]).resolve()


def _equipos_chart_structure_reference_path() -> str | None:
    from django.conf import settings

    custom = getattr(settings, "EQUIPOS_CHART_STRUCTURE_REFERENCE", None)
    if custom:
        p = Path(custom)
        if not p.is_absolute():
            p = Path(settings.BASE_DIR) / p
        if not p.is_file():
            p = Path(settings.MEDIA_ROOT) / custom
        if p.is_file():
            return str(p)
    mp = resolve_equipos_master_xlsx_path()
    if mp and mp.is_file():
        return str(mp)
    fallback = Path(settings.MEDIA_ROOT) / EQUIPOS_CHART_REFERENCE_REL
    if fallback.is_file():
        return str(fallback)
    return None


def _patch_pivot_chart_workbook_in_xml(xml_text: str, workbook_basename: str) -> str:
    """
    Excel enlaza el gráfico dinámico a la tabla vía ``<c:pivotSource><c:name>`` usando
    ``[nombreLibro.xlsx]Hoja!Pivot`` (prefijo ``c:`` u otro en OOXML).

    El regex antiguo buscaba ``<pivotSource><name>`` sin prefijo y no coincidía con el XML
    real, de modo que el vínculo quedaba desalineado tras ``openpyxl`` o al renombrar al descargar.
    """
    wb_bracket = "[" + workbook_basename + "]"
    leading_file_bracket = re.compile(r"^\s*\[[^\]]*\.xls[xm]\]", re.IGNORECASE)

    def patch_block(m: re.Match[str]) -> str:
        open_tag, inner, close_tag = m.group(1), m.group(2), m.group(3)

        def patch_name(nm: re.Match[str]) -> str:
            o, body, c = nm.group(1), nm.group(2), nm.group(3)
            if leading_file_bracket.match(body):
                new_body = leading_file_bracket.sub(wb_bracket, body, count=1)
                return o + new_body + c
            return nm.group(0)

        inner2 = re.sub(
            r"(<(?:[\w-]+:)?name\b[^>]*>)([^<]*)(</(?:[\w-]+:)?name\b[^>]*>)",
            patch_name,
            inner,
            flags=re.IGNORECASE,
        )
        return open_tag + inner2 + close_tag

    return re.sub(
        r"(<(?:[\w-]+:)?pivotSource\b[^>]*>)([\s\S]*?)(</(?:[\w-]+:)?pivotSource\b[^>]*>)",
        patch_block,
        xml_text,
        flags=re.IGNORECASE,
    )


def _xlsx_bytes_with_chart_postprocess(
    src: Path,
    workbook_display_basename: str,
    chart_structure_reference: str | None,
) -> bytes:
    ref_chart_xml: bytes | None = None
    if chart_structure_reference and Path(chart_structure_reference).is_file():
        with zipfile.ZipFile(chart_structure_reference, "r") as zr:
            if CHART_PART_PATH in zr.namelist():
                ref_chart_xml = zr.read(CHART_PART_PATH)

    out_buf = io.BytesIO()
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(
        out_buf, "w", compression=zipfile.ZIP_DEFLATED
    ) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename == CHART_PART_PATH and ref_chart_xml is not None:
                data = ref_chart_xml
            if info.filename.startswith("xl/charts/chart") and info.filename.endswith(
                ".xml"
            ):
                text = data.decode("utf-8")
                text = _patch_pivot_chart_workbook_in_xml(text, workbook_display_basename)
                data = text.encode("utf-8")
            zout.writestr(info, data)

    return out_buf.getvalue()


def rewrite_equipos_xlsx_pivot_chart_workbook(
    xlsx_path: str | Path,
    workbook_display_basename: str,
    *,
    chart_structure_reference: str | None = None,
) -> None:
    """
    Tras guardar con openpyxl, corrige pivotSource en charts y opcionalmente restaura
    chart1.xml desde un .xlsx de referencia (p. ej. media/.../...-antes.xlsx).

    chart_structure_reference: None = usar settings.EQUIPOS_CHART_STRUCTURE_REFERENCE
    o la ruta por defecto bajo MEDIA_ROOT si el fichero existe.
    """
    path = Path(xlsx_path)
    ref = chart_structure_reference
    if ref is None:
        ref = _equipos_chart_structure_reference_path()
    path.write_bytes(
        _xlsx_bytes_with_chart_postprocess(path, workbook_display_basename, ref)
    )


def equipos_xlsx_bytes_for_download(
    xlsx_path: str | Path,
    attachment_basename: str,
    *,
    chart_structure_reference: str | None = None,
) -> io.BytesIO:
    """Copia en memoria con pivotSource del chart alineado al nombre del adjunto."""
    path = Path(xlsx_path)
    ref = chart_structure_reference
    if ref is None:
        ref = _equipos_chart_structure_reference_path()
    return io.BytesIO(
        _xlsx_bytes_with_chart_postprocess(path, attachment_basename, ref)
    )


def resolve_libro_xlsx_path(libro: EquiposLibro) -> str:
    """
    Ruta absoluta al .xlsx del libro (maestro en ``EQUIPOS_LIBRO_REL_PATH``).

    Si la BD apunta a ``equipos/libro_actual.xlsx``, se alinea al canónico en 2026/04
    (copiando el legado si hace falta). Si solo existe el legado, lo copia al canónico.
    """
    from django.conf import settings

    media_root = Path(settings.MEDIA_ROOT)
    canonical = media_root / EQUIPOS_LIBRO_REL_PATH
    legacy = media_root / LEGACY_EQUIPOS_LIBRO_REL_PATH

    path_str: str | None = None
    if libro.file:
        try:
            path_str = libro.file.path
        except (NotImplementedError, ValueError):
            path_str = None

    def _repair_file_field() -> None:
        if libro.file.name != EQUIPOS_LIBRO_REL_PATH:
            libro.file.name = EQUIPOS_LIBRO_REL_PATH
            libro.save(update_fields=["file"])

    if path_str and Path(path_str).is_file():
        real = Path(path_str).resolve()
        try:
            leg = legacy.resolve()
        except OSError:
            leg = None
        try:
            can = canonical.resolve()
        except OSError:
            can = None
        if leg is not None and real == leg:
            ensure_equipos_workbook_parent_dir()
            if not canonical.is_file():
                shutil.copy2(legacy, canonical)
            _repair_file_field()
            return str(canonical.resolve())
        if can is not None and real == can:
            _repair_file_field()
            return str(canonical)
        return path_str

    if canonical.is_file():
        _repair_file_field()
        return str(canonical)

    if legacy.is_file():
        ensure_equipos_workbook_parent_dir()
        shutil.copy2(legacy, canonical)
        _repair_file_field()
        return str(canonical)

    bd = repr(libro.file.name) if (libro.file and libro.file.name) else "ninguna"
    raise FileNotFoundError(
        f"No se encontró el Excel (ruta en BD: {bd}; se buscó {canonical} y {legacy}). "
        "Importá el libro desde el hub o colocá el maestro en media/equipos/2026/04/."
    )


def build_equipos_download_filename(
    _libro: EquiposLibro, when: date | None = None
) -> str:
    """
    Nombre al descargar: ``ST01-EXP_F5-E2_Control_de_equipos_YYYY-MM-DD.xlsx``
    (fecha = día local de la descarga).
    """
    when = when or timezone.localdate()
    return f"{EQUIPOS_DOWNLOAD_BASE_STEM}_{when.isoformat()}.xlsx"


def build_equipos_pdf_download_filename(
    libro: EquiposLibro, when: date | None = None
) -> str:
    xlsx = build_equipos_download_filename(libro, when=when)
    return xlsx[:-5] + ".pdf" if xlsx.lower().endswith(".xlsx") else xlsx + ".pdf"


def build_equipos_semana_xlsx_filename(libro: EquiposLibro, semana: int | str) -> str:
    """Nombre para copia semanal: ``ST01-EXP_F5-E2_Control_de_equipos_SEMANA_<n>.xlsx``."""
    lab = str(semana).strip()
    if not lab:
        lab = "0"
    if not lab.isdigit():
        lab = re.sub(r"[^a-zA-Z0-9_.\-]", "_", lab).strip("_")[:32] or "0"
    return f"{EQUIPOS_DOWNLOAD_BASE_STEM}_SEMANA_{lab}.xlsx"


def _str_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def _cell_to_date(v: Any):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            dt = xl_from_excel(float(v))
            return dt.date() if isinstance(dt, datetime) else None
        except Exception:
            return None
    return None


def _date_to_excel(d: date | None) -> Any:
    if d is None:
        return None
    return xl_to_excel(datetime.combine(d, datetime.min.time()))


def _otro_fecha_char_to_excel_value(raw: Any) -> Any:
    """Fechas Otros guardadas como texto ISO (desde el formulario) → serial Excel; si no aplica, el valor tal cual."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return _date_to_excel(raw.date())
    if isinstance(raw, date):
        return _date_to_excel(raw)
    s = str(raw).strip()
    if not s:
        return None
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        try:
            return _date_to_excel(date.fromisoformat(s[:10]))
        except ValueError:
            return raw
    return raw


def _cell_to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def latest_libro() -> EquiposLibro | None:
    return EquiposLibro.objects.order_by("-imported_at").first()


def ultima_cambio_formulario_map(
    libro: EquiposLibro | None, modelo: str, record_ids: list[int]
) -> dict[int, datetime]:
    """Última fecha de cambio desde formulario (EquiposCambioLog) por record_id."""
    if not libro or not record_ids:
        return {}
    rows = (
        EquiposCambioLog.objects.filter(
            libro=libro, modelo=modelo, record_id__in=record_ids
        )
        .values("record_id")
        .annotate(last=Max("created_at"))
    )
    return {r["record_id"]: r["last"] for r in rows}


def ultima_cambio_un_registro(
    libro: EquiposLibro | None, modelo: str, record_id: int
) -> datetime | None:
    if not libro:
        return None
    row = (
        EquiposCambioLog.objects.filter(
            libro=libro, modelo=modelo, record_id=record_id
        )
        .aggregate(last=Max("created_at"))
    )
    return row["last"]


def format_ultima_cambio_para_json(dt: datetime | None) -> tuple[str | None, str]:
    """(iso o None, texto dd/mm/aaaa HH:MM en zona local) para APIs y tablas."""
    if dt is None:
        return None, ""
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    local = timezone.localtime(dt)
    return (local.isoformat(), local.strftime("%d/%m/%Y %H:%M"))


@transaction.atomic
def replace_libro_with_import(uploaded_file, original_filename: str) -> EquiposLibro:
    """
    Sustituye datos y el fichero en disco en ``EQUIPOS_LIBRO_REL_PATH`` (maestro en 2026/04).
    Elimina filas en cascada y borra el archivo anterior para no acumular copias.
    """
    ensure_equipos_workbook_parent_dir()
    for old in EquiposLibro.objects.all():
        if old.file:
            old.file.delete(save=False)
    EquiposLibro.objects.all().delete()

    if default_storage.exists(EQUIPOS_LIBRO_REL_PATH):
        default_storage.delete(EQUIPOS_LIBRO_REL_PATH)
    if default_storage.exists(LEGACY_EQUIPOS_LIBRO_REL_PATH):
        default_storage.delete(LEGACY_EQUIPOS_LIBRO_REL_PATH)

    libro = EquiposLibro(
        original_filename=(original_filename or "equipos.xlsx").strip(),
    )
    uploaded_file.seek(0)
    libro.file.save(EQUIPOS_WORKBOOK_FILENAME, uploaded_file, save=True)
    _import_into_libro(libro)
    return libro


@transaction.atomic
def refresh_libro_from_excel_file(
    source_xlsx: Path | str,
    original_filename: str | None = None,
) -> EquiposLibro:
    """
    Copia un .xlsx base al maestro canónico y reimporta filas a la BD **sin** borrar
    EquiposLibro ni EquiposCambioLog (mismo libro_id).
    """
    from django.core.files import File

    master_path = Path(source_xlsx).resolve()
    if not master_path.is_file():
        raise FileNotFoundError(str(master_path))
    ofn = (original_filename or master_path.name).strip()

    libro = latest_libro()
    if libro is None:
        ensure_equipos_workbook_parent_dir()
        if default_storage.exists(EQUIPOS_LIBRO_REL_PATH):
            default_storage.delete(EQUIPOS_LIBRO_REL_PATH)
        if default_storage.exists(LEGACY_EQUIPOS_LIBRO_REL_PATH):
            default_storage.delete(LEGACY_EQUIPOS_LIBRO_REL_PATH)
        libro = EquiposLibro(original_filename=ofn)
        with master_path.open("rb") as fh:
            libro.file.save(EQUIPOS_WORKBOOK_FILENAME, File(fh), save=True)
        _import_into_libro(libro)
        return libro

    libro.resumen_filas.all().delete()
    libro.significado_filas.all().delete()
    libro.locations.all().delete()
    libro.assets.all().delete()
    libro.otros.all().delete()

    from django.conf import settings

    ensure_equipos_workbook_parent_dir()
    dest = Path(settings.MEDIA_ROOT) / EQUIPOS_LIBRO_REL_PATH
    shutil.copy2(master_path, dest)
    if libro.file.name != EQUIPOS_LIBRO_REL_PATH:
        libro.file.name = EQUIPOS_LIBRO_REL_PATH
        libro.save(update_fields=["file"])
    update_fields: list[str] = []
    if ofn and libro.original_filename != ofn:
        libro.original_filename = ofn
        update_fields.append("original_filename")
    if update_fields:
        libro.save(update_fields=update_fields)
    _import_into_libro(libro)
    EquiposLibro.objects.filter(pk=libro.pk).update(updated_at=timezone.now())
    return libro


def reseed_libro_datos_from_semana_excel(
    libro: EquiposLibro, semana_excel_path: Path | str
) -> None:
    """
    Sustituye en la BD las filas importadas desde las hojas (Resumen, Significado,
    Locations, Asset, Otros) leyendo **otro** .xlsx con la misma estructura (p. ej. semana 13).

    No borra ``EquiposLibro`` ni ``EquiposCambioLog``. No modifica el .xlsx en disco hasta
    que llames a ``sync_libro_to_excel``.
    """
    src = Path(semana_excel_path).resolve()
    if not src.is_file():
        raise FileNotFoundError(str(src))
    libro.resumen_filas.all().delete()
    libro.significado_filas.all().delete()
    libro.locations.all().delete()
    libro.assets.all().delete()
    libro.otros.all().delete()
    _import_into_libro(libro, source_path=str(src))
    EquiposLibro.objects.filter(pk=libro.pk).update(updated_at=timezone.now())


def equipos_reseed_semana_replay_logs_y_sync(
    libro: EquiposLibro,
    semana_excel_path: Path | str,
    log_since: datetime,
) -> tuple[int, int, list[str]]:
    """
    1) BD desde ``semana_excel_path`` (mismas filas que el maestro).
    2) Reaplica ``EquiposCambioLog`` con ``created_at >= log_since`` (los logs no se borran).
    3) ``sync_libro_to_excel``: escribe celdas en el maestro canónico.

    Devuelve lo mismo que ``replay_equipos_logs_on_libro`` (aplicados, omitidos, notas).
    """
    with transaction.atomic():
        reseed_libro_datos_from_semana_excel(libro, semana_excel_path)
        log_rows = list(
            EquiposCambioLog.objects.filter(libro=libro, created_at__gte=log_since)
            .order_by("created_at")
            .values("modelo", "excel_row", "campo", "valor_nuevo")
        )
        applied, skipped, notes = replay_equipos_logs_on_libro(libro, log_rows)
    sync_libro_to_excel(libro)
    return applied, skipped, notes


def _row_empty_asset(ws, row: int, cmin: int = 2, cmax: int = 22) -> bool:
    for c in range(cmin, cmax + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip() != "":
            return False
    return True


def _import_into_libro(
    libro: EquiposLibro, *, source_path: str | Path | None = None
) -> None:
    """Importa hojas a la BD. Si ``source_path`` se indica, lee ese .xlsx (p. ej. semana 13)."""
    path = str(Path(source_path).resolve()) if source_path is not None else libro.file.path
    wb = load_workbook(path, data_only=True)

    ws = wb[SHEET_RESUMEN]
    for r in range(6, 14):
        etiqueta = _str_cell(ws.cell(row=r, column=1).value)
        cuenta = ws.cell(row=r, column=2).value
        fr = ws.cell(row=r, column=3).value
        ci = None
        if cuenta is not None and str(cuenta).strip() != "":
            try:
                ci = int(float(cuenta))
            except (TypeError, ValueError):
                ci = None
        cf = None
        if fr is not None and str(fr).strip() != "":
            try:
                cf = float(fr)
            except (TypeError, ValueError):
                cf = None
        EquiposResumenFila.objects.create(
            libro=libro,
            excel_row=r,
            etiqueta=etiqueta,
            cuenta=ci,
            fraccion=cf,
        )

    ws = wb[SHEET_SIGNIFICADO]
    for r in range(4, (ws.max_row or 0) + 1):
        flujo = _str_cell(ws.cell(row=r, column=2).value)
        st = _str_cell(ws.cell(row=r, column=3).value)
        sig = _str_cell(ws.cell(row=r, column=4).value)
        if not flujo and not st and not sig:
            continue
        EquiposSignificadoFila.objects.create(
            libro=libro,
            excel_row=r,
            flujo=flujo,
            status=st,
            significado=sig,
        )

    ws = wb[SHEET_LOCATIONS]
    max_r = ws.max_row or 0
    for r in range(4, max_r + 1):
        if _row_empty_asset(ws, r, 2, 10):
            continue
        EquiposLocation.objects.create(
            libro=libro,
            excel_row=r,
            campus=_str_cell(ws.cell(row=r, column=2).value),
            building=_str_cell(ws.cell(row=r, column=3).value),
            zones=_str_cell(ws.cell(row=r, column=4).value),
            floors=_str_cell(ws.cell(row=r, column=5).value),
            space_name=_str_cell(ws.cell(row=r, column=6).value),
            fase=_str_cell(ws.cell(row=r, column=8).value),
            area_m2=_cell_to_decimal(ws.cell(row=r, column=9).value),
            code=_str_cell(ws.cell(row=r, column=10).value),
        )

    ws = wb[SHEET_ASSET]
    max_r = ws.max_row or 0
    for r in range(4, max_r + 1):
        if _row_empty_asset(ws, r, 2, 22):
            continue
        raw_tipe = _str_cell(ws.cell(row=r, column=2).value).upper()
        if raw_tipe == EquiposAsset.ROW_TITULO:
            rt = EquiposAsset.ROW_TITULO
        elif raw_tipe == EquiposAsset.ROW_SUBTITULO:
            rt = EquiposAsset.ROW_SUBTITULO
        else:
            rt = EquiposAsset.ROW_TAREA
        EquiposAsset.objects.create(
            libro=libro,
            excel_row=r,
            row_type=rt,
            tipe=_str_cell(ws.cell(row=r, column=2).value),
            especialidad=_str_cell(ws.cell(row=r, column=3).value),
            tag_number=_str_cell(ws.cell(row=r, column=4).value),
            asset_name=_str_cell(ws.cell(row=r, column=5).value),
            space_room=_str_cell(ws.cell(row=r, column=6).value),
            unit=_str_cell(ws.cell(row=r, column=7).value),
            quantity=_str_cell(ws.cell(row=r, column=8).value),
            phase=_str_cell(ws.cell(row=r, column=9).value),
            zones=_str_cell(ws.cell(row=r, column=10).value),
            proveedor=_str_cell(ws.cell(row=r, column=11).value),
            vendor=_str_cell(ws.cell(row=r, column=12).value),
            estado=_str_cell(ws.cell(row=r, column=13).value),
            con_oc=_str_cell(ws.cell(row=r, column=14).value),
            fecha_compra=_cell_to_date(ws.cell(row=r, column=15).value),
            rdi_ttal=_str_cell(ws.cell(row=r, column=16).value),
            fecha_llegada_obra=_cell_to_date(ws.cell(row=r, column=17).value),
            fecha_planificacion=_cell_to_date(ws.cell(row=r, column=18).value),
            cumple=_str_cell(ws.cell(row=r, column=19).value),
            dias=_str_cell(ws.cell(row=r, column=20).value),
            avance_montaje=_str_cell(ws.cell(row=r, column=21).value),
            avance_conexion=_str_cell(ws.cell(row=r, column=22).value),
        )

    ws = wb[SHEET_OTROS]
    max_r = ws.max_row or 0
    for r in range(2, max_r + 1):
        if _row_empty_asset(ws, r, 2, 10):
            continue
        esp = _str_cell(ws.cell(row=r, column=3).value)
        tipe = _str_cell(ws.cell(row=r, column=2).value)
        tag = _str_cell(ws.cell(row=r, column=4).value)
        asset = _str_cell(ws.cell(row=r, column=5).value)
        st = _str_cell(ws.cell(row=r, column=6).value)
        rdi = _str_cell(ws.cell(row=r, column=7).value)
        fe_c = ws.cell(row=r, column=8).value
        fr_c = ws.cell(row=r, column=9).value
        fe_dt = _cell_to_date(fe_c)
        fr_dt = _cell_to_date(fr_c)
        fe = fe_dt.isoformat() if fe_dt else ""
        fr = fr_dt.isoformat() if fr_dt else ""
        oc = _str_cell(ws.cell(row=r, column=10).value)
        is_section = (
            esp
            and not tipe
            and not tag
            and not asset
            and not st
            and not rdi
            and not fe
            and not fr
            and not oc
        )
        EquiposOtro.objects.create(
            libro=libro,
            excel_row=r,
            row_type=EquiposOtro.ROW_SECTION if is_section else EquiposOtro.ROW_DATA,
            tipe=tipe,
            especialidad=esp,
            tag_number=tag,
            asset_name=asset,
            estado=st,
            rdi_ttal=rdi,
            fecha_envio_rdi=fe,
            fecha_respuesta_rdi=fr,
            con_oc=oc,
        )

    wb.close()


def log_changes(
    libro: EquiposLibro,
    user,
    modelo: str,
    record_id: int,
    excel_row: int | None,
    before: dict[str, Any],
    after: dict[str, Any],
    fields: Iterable[str],
) -> None:
    for f in fields:
        ov = before.get(f)
        nv = after.get(f)
        ovs = "" if ov is None else str(ov)
        nvs = "" if nv is None else str(nv)
        if ovs == nvs:
            continue
        EquiposCambioLog.objects.create(
            libro=libro,
            user=user,
            modelo=modelo,
            record_id=record_id,
            excel_row=excel_row,
            campo=f,
            valor_anterior=ovs,
            valor_nuevo=nvs,
        )


def _parse_log_valor_nuevo(model_cls: type, campo: str, raw: str) -> Any:
    """Convierte valor_nuevo del log al tipo del campo Django."""
    try:
        field = model_cls._meta.get_field(campo)
    except FieldDoesNotExist as e:
        raise ValueError(f"Campo desconocido {campo} en {model_cls.__name__}") from e
    s = "" if raw is None else str(raw).strip()
    if isinstance(field, dj_models.DateField):
        if not s:
            return None
        return date.fromisoformat(s[:10])
    if isinstance(field, dj_models.DecimalField):
        if not s:
            return None
        return Decimal(s.replace(",", "."))
    if isinstance(field, dj_models.IntegerField):
        if not s:
            return None
        return int(float(s))
    if isinstance(field, dj_models.FloatField):
        if not s:
            return None
        return float(s.replace(",", "."))
    if s == "" and getattr(field, "null", False):
        return None
    return raw if raw is not None else ""


def replay_equipos_logs_on_libro(
    libro: EquiposLibro,
    log_rows: Iterable[dict[str, Any]],
) -> tuple[int, int, list[str]]:
    """
    Reaplica eventos de EquiposCambioLog (dicts con modelo, excel_row, campo, valor_nuevo)
    sobre el libro dado, en orden. Busca fila por excel_row + modelo.

    Returns:
        (aplicados, omitidos, mensajes_omitidos)
    """
    model_map = {
        "EquiposAsset": EquiposAsset,
        "EquiposOtro": EquiposOtro,
        "EquiposLocation": EquiposLocation,
        "EquiposResumenFila": EquiposResumenFila,
        "EquiposSignificadoFila": EquiposSignificadoFila,
    }
    applied = 0
    skipped = 0
    notes: list[str] = []

    for row in log_rows:
        modelo = row.get("modelo") or ""
        excel_row = row.get("excel_row")
        campo = row.get("campo") or ""
        valor_nuevo = row.get("valor_nuevo")
        if modelo not in model_map or excel_row is None or not campo:
            skipped += 1
            notes.append(f"Fila log incompleta: {row}")
            continue
        model_cls = model_map[modelo]
        try:
            obj = model_cls.objects.get(libro=libro, excel_row=int(excel_row))
        except model_cls.DoesNotExist:
            skipped += 1
            notes.append(f"{modelo} excel_row={excel_row} no existe en el libro actual")
            continue
        try:
            val = _parse_log_valor_nuevo(model_cls, campo, valor_nuevo if valor_nuevo is not None else "")
        except (ValueError, TypeError) as e:
            skipped += 1
            notes.append(f"{modelo} row={excel_row} {campo}={valor_nuevo!r}: {e}")
            continue
        setattr(obj, campo, val)
        obj.save(update_fields=[campo])
        applied += 1

    return applied, skipped, notes


def copy_libro_xlsx_to_semanales(libro: EquiposLibro, filename: str) -> str:
    """
    Copia el xlsx actual del libro (tras sync) a MEDIA_ROOT/equipos/semanales/<filename>.
    filename debe incluir .xlsx si corresponde.
    """
    from django.conf import settings

    src = Path(resolve_libro_xlsx_path(libro))
    dest_dir = Path(settings.MEDIA_ROOT) / EQUIPOS_SEMANALES_REL_DIR
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / filename
    shutil.copy2(src, dest)
    rewrite_equipos_xlsx_pivot_chart_workbook(dest, dest.name)
    return str(dest)


def sync_libro_to_excel(libro: EquiposLibro) -> None:
    """Escribe los valores del modelo en el maestro ``EQUIPOS_LIBRO_REL_PATH`` (mismo .xlsx del libro).

    Se abre el fichero canónico bajo ``media/equipos/2026/04/``, se actualizan celdas y se guarda
    in situ (formato, tablas y gráficos del libro se conservan en la medida de openpyxl).

    En «Resumen - TD» no se sobreescriben B/C en las filas 6–12 (fórmulas y cabeceras); ver
    ``RESUMEN_EXCEL_ROWS_PRESERVE_BC``.
    """
    ensure_equipos_workbook_parent_dir()
    path = resolve_libro_xlsx_path(libro)
    wb = load_workbook(path, keep_vba=False, keep_links=True)

    ws = wb[SHEET_RESUMEN]
    for row in libro.resumen_filas.all():
        r = row.excel_row
        ws.cell(row=r, column=1, value=row.etiqueta or None)
        if r in RESUMEN_EXCEL_ROWS_PRESERVE_BC:
            continue
        ws.cell(row=r, column=2, value=row.cuenta)
        ws.cell(row=r, column=3, value=row.fraccion)

    ws = wb[SHEET_SIGNIFICADO]
    for row in libro.significado_filas.all():
        r = row.excel_row
        ws.cell(row=r, column=2, value=row.flujo or None)
        ws.cell(row=r, column=3, value=row.status or None)
        ws.cell(row=r, column=4, value=row.significado or None)

    ws = wb[SHEET_LOCATIONS]
    for row in libro.locations.all():
        r = row.excel_row
        ws.cell(row=r, column=2, value=row.campus or None)
        ws.cell(row=r, column=3, value=row.building or None)
        ws.cell(row=r, column=4, value=row.zones or None)
        ws.cell(row=r, column=5, value=row.floors or None)
        ws.cell(row=r, column=6, value=row.space_name or None)
        ws.cell(row=r, column=8, value=row.fase or None)
        ws.cell(row=r, column=9, value=float(row.area_m2) if row.area_m2 is not None else None)
        ws.cell(row=r, column=10, value=row.code or None)

    ws = wb[SHEET_ASSET]
    for row in libro.assets.all():
        r = row.excel_row
        ws.cell(row=r, column=2, value=row.tipe or None)
        ws.cell(row=r, column=3, value=row.especialidad or None)
        ws.cell(row=r, column=4, value=row.tag_number or None)
        ws.cell(row=r, column=5, value=row.asset_name or None)
        ws.cell(row=r, column=6, value=row.space_room or None)
        ws.cell(row=r, column=7, value=row.unit or None)
        ws.cell(row=r, column=8, value=row.quantity or None)
        ws.cell(row=r, column=9, value=row.phase or None)
        ws.cell(row=r, column=10, value=row.zones or None)
        ws.cell(row=r, column=11, value=row.proveedor or None)
        ws.cell(row=r, column=12, value=row.vendor or None)
        ws.cell(row=r, column=13, value=row.estado or None)
        ws.cell(row=r, column=14, value=row.con_oc or None)
        ws.cell(row=r, column=15, value=_date_to_excel(row.fecha_compra))
        ws.cell(row=r, column=16, value=row.rdi_ttal or None)
        ws.cell(row=r, column=17, value=_date_to_excel(row.fecha_llegada_obra))
        ws.cell(row=r, column=18, value=_date_to_excel(row.fecha_planificacion))
        # Columnas 19–20 (Cumple / Días): no escribir; en el libro suelen ser fórmulas.
        ws.cell(row=r, column=21, value=row.avance_montaje or None)
        ws.cell(row=r, column=22, value=row.avance_conexion or None)

    ws = wb[SHEET_OTROS]
    for row in libro.otros.all():
        r = row.excel_row
        ws.cell(row=r, column=2, value=row.tipe or None)
        ws.cell(row=r, column=3, value=row.especialidad or None)
        ws.cell(row=r, column=4, value=row.tag_number or None)
        ws.cell(row=r, column=5, value=row.asset_name or None)
        ws.cell(row=r, column=6, value=row.estado or None)
        ws.cell(row=r, column=7, value=row.rdi_ttal or None)
        ws.cell(row=r, column=8, value=_otro_fecha_char_to_excel_value(row.fecha_envio_rdi))
        ws.cell(row=r, column=9, value=_otro_fecha_char_to_excel_value(row.fecha_respuesta_rdi))
        ws.cell(row=r, column=10, value=row.con_oc or None)

    wb.save(path)
    wb.close()
    rewrite_equipos_xlsx_pivot_chart_workbook(path, Path(path).name)
    EquiposLibro.objects.filter(pk=libro.pk).update(updated_at=timezone.now())


def equipos_cambios_logs_queryset(libro: EquiposLibro | None, q: str = ""):
    """
    Misma regla que la lista HTML: si hay ``libro`` activo, solo sus entradas; si no, todas.
    ``q`` aplica búsqueda insensible a mayúsculas en varios campos.
    """
    base = EquiposCambioLog.objects.select_related("user", "libro")
    if libro:
        base = base.filter(libro=libro)
    q = (q or "").strip()
    if not q:
        return base.order_by("-created_at")
    cond = (
        Q(modelo__icontains=q)
        | Q(campo__icontains=q)
        | Q(valor_anterior__icontains=q)
        | Q(valor_nuevo__icontains=q)
        | Q(user__username__icontains=q)
        | Q(libro__original_filename__icontains=q)
    )
    if q.isdigit():
        n = int(q)
        cond |= Q(record_id=n) | Q(excel_row=n)
    return base.filter(cond).order_by("-created_at")


def build_equipos_cambios_pdf_filename(libro: EquiposLibro | None, when: date | None = None) -> str:
    when = when or timezone.localdate()
    stem = EQUIPOS_DOWNLOAD_BASE_STEM
    if libro and (libro.original_filename or "").strip():
        safe = re.sub(r"[^\w.\-]+", "_", Path(libro.original_filename).stem)[:48]
        stem = f"{EQUIPOS_DOWNLOAD_BASE_STEM}_{safe}"
    return f"{stem}_cambios_{when.isoformat()}.pdf"


def build_equipos_cambios_log_pdf_buffer(
    rows: list[EquiposCambioLog],
    title_line: str,
) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="CambiosTitle",
        parent=styles["Heading2"],
        fontSize=10,
        spaceAfter=6,
    )
    story = [Paragraph(escape(title_line), title_style), Spacer(1, 8)]
    if not rows:
        story.append(Paragraph("No hay registros para los criterios actuales.", styles["Normal"]))
        doc.build(story)
        buf.seek(0)
        return buf

    def trunc(s: str, n: int) -> str:
        t = (s or "").replace("\n", " ").strip()
        return t if len(t) <= n else t[: n - 1] + "…"

    data = [
        ["Fecha", "Usuario", "Modelo", "ID", "Fila", "Campo", "Antes", "Después"],
    ]
    for c in rows:
        lt = timezone.localtime(c.created_at)
        data.append(
            [
                lt.strftime("%d/%m/%Y %H:%M"),
                trunc(c.user.get_username(), 24) if c.user else "—",
                trunc(c.modelo, 20),
                str(c.record_id),
                str(c.excel_row) if c.excel_row is not None else "—",
                trunc(c.campo, 22),
                trunc(c.valor_anterior, 48),
                trunc(c.valor_nuevo, 48),
            ]
        )
    t = Table(data, repeatRows=1, hAlign="LEFT")
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8e8e8")),
                ("FONTSIZE", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(t)
    doc.build(story)
    buf.seek(0)
    return buf


def build_pdf_buffer(libro: EquiposLibro) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="EqTitle",
        parent=styles["Heading2"],
        fontSize=11,
        spaceAfter=6,
    )
    story = []
    story.append(
        Paragraph(
            f"Control de equipos — {libro.original_filename} — {libro.imported_at:%d/%m/%Y %H:%M}",
            title_style,
        )
    )
    story.append(Spacer(1, 8))

    def add_table(headers: list[str], rows: list[list[str]], caption: str):
        story.append(Paragraph(caption, title_style))
        data = [headers] + rows[:400]
        t = Table(data, repeatRows=1, hAlign="LEFT")
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8e8e8")),
                    ("FONTSIZE", (0, 0), (-1, -1), 6),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(t)
        story.append(Spacer(1, 10))

    assets = libro.assets.filter(row_type=EquiposAsset.ROW_TAREA).order_by("excel_row")[:500]
    if assets.exists():
        add_table(
            ["Fila", "TAG", "Asset", "Estado", "Zona", "Proveedor"],
            [
                [
                    str(a.excel_row),
                    a.tag_number,
                    (a.asset_name or "")[:80],
                    (a.estado or "")[:40],
                    (a.zones or "")[:30],
                    (a.proveedor or "")[:30],
                ]
                for a in assets
            ],
            "Asset (solo filas TAREA)",
        )

    otros = libro.otros.filter(row_type=EquiposOtro.ROW_DATA).order_by("excel_row")[:400]
    if otros.exists():
        add_table(
            ["Fila", "TAG", "Asset", "Estado", "RDI"],
            [
                [
                    str(o.excel_row),
                    o.tag_number,
                    (o.asset_name or "")[:80],
                    (o.estado or "")[:40],
                    o.rdi_ttal,
                ]
                for o in otros
            ],
            "Otros equipos",
        )

    locs = libro.locations.order_by("excel_row")[:300]
    if locs.exists():
        add_table(
            ["Código", "Space", "Zona", "Área"],
            [
                [
                    l.code,
                    (l.space_name or "")[:60],
                    l.zones,
                    str(l.area_m2 or ""),
                ]
                for l in locs
            ],
            "Locations",
        )

    doc.build(story)
    buf.seek(0)
    return buf

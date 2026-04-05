from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib import messages
from django.contrib.auth.views import LoginView
from django.views.decorators.http import require_GET, require_POST
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.utils import timezone
from django.contrib.auth import get_user_model
import os
import re
from datetime import date, datetime
from urllib.parse import quote

from django.core.mail import EmailMessage
from django.conf import settings

from .models import (
    Document,
    Folder,
    FolderFile,
    DocumentAttachment,
    CorreoEnviado,
    GrupoCorreo,
    UserSessionLog,
    UserPresence,
)
from .search_backend import search_unified, _normalize_terms
from .snippets import extract_snippets, extract_snippets_multi_term
from .text_search_match import text_matches_all_terms_as_words, text_matches_single_query
from .traceability import build_journeys_for_query

from rdi.models import (
    RDIRecord,
    RDI_INFORMADO_NO,
    RDI_INFORMADO_SI,
    RDI_INFORMADO_OTRA,
    RDI_STATUS_BORRADOR,
    RDI_STATUS_NULA,
    RDI_STATUS_REMITIDA,
    RDI_STATUS_RESPONDIDA,
)

# Estados no mostrados en el desglose del panel RDI
RDI_PANEL_STATUS_EXCLUDE = frozenset(
    {RDI_STATUS_BORRADOR, RDI_STATUS_REMITIDA, RDI_STATUS_NULA}
)


def _rdi_discipline_rows(rdi_universe, now):
    """
    Por valor de `discipline` no vacío: totales, respuesta según el campo `response`
    (con texto vs vacío; siempre suman el total de la fila) y vencimiento.
    Registros sin disciplina se omiten.
    """
    from collections import defaultdict

    buckets = defaultdict(
        lambda: {
            "total": 0,
            "con_respuesta": 0,
            "sin_respuesta": 0,
            "vencidas": 0,
            "vigentes": 0,
            "sin_fecha": 0,
        }
    )
    for discipline, response, due_date in rdi_universe.values_list(
        "discipline", "response", "due_date"
    ):
        raw = (discipline or "").strip()
        if not raw:
            continue
        key = raw
        b = buckets[key]
        b["total"] += 1
        if (response or "").strip():
            b["con_respuesta"] += 1
        else:
            b["sin_respuesta"] += 1
        if due_date is None:
            b["sin_fecha"] += 1
        elif due_date < now:
            b["vencidas"] += 1
        else:
            b["vigentes"] += 1

    rows = []
    for key, b in buckets.items():
        short = key if len(key) <= 64 else key[:61] + "…"
        rows.append(
            {
                "label": short,
                "full": key,
                "total": b["total"],
                "con_respuesta": b["con_respuesta"],
                "sin_respuesta": b["sin_respuesta"],
                "vencidas": b["vencidas"],
                "vigentes": b["vigentes"],
                "sin_fecha": b["sin_fecha"],
            }
        )
    rows.sort(key=lambda r: (-r["total"], r["full"].lower()))
    return rows


def _rdi_panel_universe():
    """RDI con actividad en 2026 o posteriores (mismo criterio que panel staff y pizarra)."""
    start_2026 = timezone.make_aware(datetime(2026, 1, 1, 0, 0, 0))
    return RDIRecord.objects.filter(
        Q(created_at__gte=start_2026)
        | Q(created_at__isnull=True, updated_at__gte=start_2026)
        | Q(
            created_at__isnull=True,
            updated_at__isnull=True,
            last_snapshot_datetime__gte=start_2026,
        )
        | Q(
            created_at__isnull=True,
            updated_at__isnull=True,
            last_snapshot_datetime__isnull=True,
            due_date__gte=start_2026,
        )
    )


def _format_datetime_correo_es(dt):
    if not dt:
        return "—"
    try:
        if timezone.is_aware(dt):
            dt = timezone.localtime(dt)
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(dt)


def _bool_carta_es(val):
    if val is None:
        return "No indicado"
    return "Sí" if val else "No"


def _cuerpo_correo_desde_rdi_record(rdi: RDIRecord, request) -> str:
    """
    Texto tipo carta con todos los campos visibles del registro RDI (BIM).
    """
    status_labels = dict(RDIRecord._meta.get_field("status").choices)
    try:
        informado_txt = rdi.get_informado_display()
    except Exception:
        informado_txt = str(rdi.informado or "—")

    def t(s):
        return (s or "").strip()

    lines = [
        "Estimados,",
        "",
        "Por medio del presente se informa el siguiente registro RDI (origen BIM) para su conocimiento y gestión:",
        "",
        "─── Datos generales ───",
        f"Identificador (CSV): {rdi.csv_id}",
    ]
    if t(rdi.title):
        lines.append(f"Asunto / título: {t(rdi.title)}")
    lines.append(f"Estado del RDI: {status_labels.get(rdi.status, rdi.status)}")
    lines.append(f"Estado «Informar»: {informado_txt}")
    lines.append(f"Compañía: {t(rdi.company) or '—'}")
    lines.append(f"Prioridad: {t(rdi.priority) or '—'}")
    lines.append(f"Disciplina: {t(rdi.discipline) or '—'}")
    lines.append(f"Categoría: {t(rdi.category) or '—'}")
    lines.append(f"Referencia: {t(rdi.reference) or '—'}")
    lines.append(f"Impacto en costo: {t(rdi.cost_impact) or '—'}")
    lines.append(f"Impacto en cronograma: {t(rdi.schedule_impact) or '—'}")
    lines.append(f"¿Asociado a documento?: {_bool_carta_es(rdi.associated_to_document)}")
    lines.append(f"Fecha de vencimiento: {_format_datetime_correo_es(rdi.due_date)}")
    lines.extend(["", "─── Consulta ───", t(rdi.question) or "(Sin texto)", ""])
    lines.extend(["─── Respuesta sugerida ───", t(rdi.suggested_answer) or "(Sin texto)", ""])
    lines.extend(["─── Ubicación / detalle ───", t(rdi.location_details) or "(Sin texto)", ""])
    lines.extend(["─── Respuesta registrada ───", t(rdi.response) or "(Sin texto)", ""])
    lines.extend(
        [
            "─── Asignación ───",
            f"Asignado a: {t(rdi.assigned_to) or '—'}",
            f"Tipo de asignación: {t(rdi.assignee_type) or '—'}",
            f"Lista de distribución: {t(rdi.distribution_list) or '—'}",
            "",
            "─── Trazabilidad ───",
            f"Creado: {_format_datetime_correo_es(rdi.created_at)}   Por: {t(rdi.created_by) or '—'}",
            f"Actualizado: {_format_datetime_correo_es(rdi.updated_at)}   Por: {t(rdi.updated_by) or '—'}",
        ]
    )
    if rdi.last_snapshot_datetime:
        lines.append(
            f"Última versión de archivo (importación): {_format_datetime_correo_es(rdi.last_snapshot_datetime)}"
        )
    if getattr(rdi, "last_import_id", None) and rdi.last_import:
        lines.append(f"Archivo CSV de origen: {t(rdi.last_import.original_filename) or '—'}")
    if t(rdi.last_diff_fields):
        lines.append(f"Últimos campos modificados (sistema): {rdi.last_diff_fields}")
    lines.append("")
    try:
        admin_url = request.build_absolute_uri(reverse("admin:rdi_rdirecord_change", args=[rdi.pk]))
        lines.append(f"Ficha en administración Condocdat: {admin_url}")
    except Exception:
        pass
    try:
        bim_url = request.build_absolute_uri(reverse("informar_bim_list"))
        lines.append(f"Listado «Informar desde BIM»: {bim_url}")
    except Exception:
        pass
    lines.extend(["", "Atentamente,"])
    return "\n".join(lines)


def _parse_requiere_respuesta(content_extract):
    """
    Extrae «Requiere respuesta» SI o NO según la cadena 'SI  NO  X' o 'SI  X  NO' en el extracto.
    Esa cadena puede estar al final, tras "REQUIERE RESPUESTA", o en cualquier otra posición.
    Se consulta el extracto y se busca en todo el texto la primera zona con SI, NO y X.
    """
    if not content_extract or not isinstance(content_extract, str):
        return ""

    t = content_extract.upper()
    t = t.replace("\r\n", " ").replace("\n", " ")
    t = t.replace("\u00CD", "I").replace("Í", "I")
    # Algunos OCR marcan la casilla con símbolos distintos a "X"
    for char in ("\u2713", "\u2714", "\u2612", "\u2611", "\u25A0", "□", "■", "✗", "✕", "✖", "×", "✘"):
        t = t.replace(char, "X")
    # Normalizar espacios para hacer match más estable
    t = re.sub(r"\s+", " ", t).strip()

    # OCR común: "S I" / "N O" en vez de "SI" / "NO"
    t = re.sub(r"\bS\s*I\b", "SI", t)
    t = re.sub(r"\bN\s*O\b", "NO", t)

    def find_word(seg, word):
        i = 0
        while True:
            p = seg.find(word, i)
            if p < 0:
                return -1
            before_ok = p == 0 or not seg[p - 1].isalnum()
            end = p + len(word)
            after_ok = end >= len(seg) or not seg[end].isalnum()
            if before_ok and after_ok:
                return p
            i = p + 1

    def decide_from_segment(segment):
        if len(segment) < 5:
            return ""
        pos_si = find_word(segment, "SI")
        pos_no = find_word(segment, "NO")
        if pos_si < 0 or pos_no < 0:
            return ""

        # Heurística principal: patrones de casilla típicos
        # - SI X NO  => requiere "SI"
        # - SI NO X  => requiere "NO"
        # - (análogos invertidos)
        seg = segment
        try:
            if re.search(r"(?<!\w)SI(?!\w)\s*X\s*(?<!\w)NO(?!\w)", seg):
                return "SI"
            if re.search(r"(?<!\w)NO(?!\w)\s*X\s*(?<!\w)SI(?!\w)", seg):
                return "NO"
            if re.search(r"(?<!\w)SI(?!\w)\s*(?<!\w)NO(?!\w)\s*(?<!\w)X(?!\w)", seg):
                return "NO"
            if re.search(r"(?<!\w)NO(?!\w)\s*(?<!\w)SI(?!\w)\s*(?<!\w)X(?!\w)", seg):
                return "SI"
        except Exception:
            pass

        # Fallback: orden de apariciones (mantener lo existente pero más acotado)
        pos_x = -1
        for m in re.finditer(r"(?<!\w)X(?!\w)", seg):
            x_pos = m.start()
            # X cerca de la casilla: entre SI y NO o muy próximo
            if (pos_si <= x_pos <= pos_no) or (pos_no <= x_pos <= pos_si) or abs(x_pos - pos_si) < 200:
                pos_x = x_pos
                break
        if pos_x < 0:
            return ""

        orden = sorted([(pos_si, "SI"), (pos_no, "NO"), (pos_x, "X")], key=lambda x: x[0])
        secuencia = [eti for _, eti in orden]
        if secuencia == ["SI", "X", "NO"] or secuencia == ["NO", "X", "SI"]:
            return "SI" if secuencia[0] == "SI" else "NO"
        # Si queda X al final, generalmente marcó el último label
        if secuencia[-1] == "X":
            return "NO" if secuencia[0] == "SI" else "SI"
        if "X" in secuencia[1:2]:
            return "SI" if secuencia[0] == "SI" else "NO"
        return ""

    # 1) Tras "REQUIERE RESPUESTA" si existe
    if "REQUIERE RESPUESTA" in t:
        idx = t.find("REQUIERE RESPUESTA")
        r = decide_from_segment(t[idx : idx + 500])
        if r:
            return r

    # 2) Final del documento
    seg_end = t[-500:] if len(t) >= 500 else t
    r = decide_from_segment(seg_end)
    if r:
        return r

    # 3) Buscar en todo el extracto (la cadena puede estar en cualquier sitio)
    window = 220
    i = 0
    while i < len(t):
        p_si = t.find(" SI ", i)
        p_no = t.find(" NO ", i)
        p = -1
        if p_si >= 0 and p_no >= 0:
            p = min(p_si, p_no)
        elif p_si >= 0:
            p = p_si
        elif p_no >= 0:
            p = p_no
        if p < 0:
            break
        start = max(0, p - 30)
        end = min(len(t), p + window)
        r = decide_from_segment(t[start:end])
        if r:
            return r
        i = p + 1

    return ""


def _parse_asunto(content_extract):
    """
    Extrae del extracto (carta) lo que está después de «Asunto:» y antes de «Atención:».
    Puede ser varias líneas. Se usa en Estatus Cartas para la columna Asunto.
    """
    if not content_extract or not isinstance(content_extract, str):
        return ""
    t = content_extract
    tu = t.upper()
    idx = tu.find("ASUNTO:")
    if idx < 0:
        return ""
    start = idx + len("ASUNTO:")
    while start < len(t) and t[start] in " :\t":
        start += 1
    if start >= len(t):
        return ""
    # Cortar en "Atención:" o "Atencion:"
    end_pos = len(t)
    for mark in ("ATENCIÓN:", "ATENCION:"):
        pos = tu.find(mark, start)
        if pos >= 0:
            end_pos = min(end_pos, pos)
    block = t[start:end_pos].strip()
    return block[:500] if len(block) > 500 else block


def _parse_enviado_a_despues_senor(content_extract):
    """
    Extrae lo que está después de «Señor» (o «Señor») y antes de «ODATA» en el extracto.
    Ejemplo: «Señor\nPablo Bravo\nODATA CHILE SPA» → «Pablo Bravo».
    """
    if not content_extract or not isinstance(content_extract, str):
        return ""
    t = content_extract
    tu = t.upper()
    idx = tu.find("SEÑOR")
    if idx < 0:
        idx = tu.find("SENOR")
    if idx < 0:
        return ""
    start = idx + (5 if "SEÑOR" in t[idx : idx + 6].upper() else 5)
    while start < len(t) and (t[start] in " \n\r\t:" or t[start].isspace()):
        start += 1
    if start >= len(t):
        return ""
    pos_odata = tu.find("ODATA", start)
    if pos_odata >= 0:
        block = t[start:pos_odata].strip()
    else:
        block = t[start:].strip()
    return block[:300] if len(block) > 300 else block


def _parse_atencion(content_extract):
    """
    Extrae del extracto el texto que sigue a «Atención:» (ej. «Atención: Sr. Ignacio Bravo G.»).
    Devuelve solo la parte después de «Atención:», hasta fin de línea o un límite razonable.
    """
    if not content_extract or not isinstance(content_extract, str):
        return ""
    t = content_extract
    tu = t.upper()
    idx = tu.find("ATENCIÓN:")
    if idx < 0:
        idx = tu.find("ATENCION:")
    if idx < 0:
        return ""
    # Buscar los dos puntos y empezar después
    colon = t.find(":", idx)
    if colon < 0:
        return ""
    start = colon + 1
    while start < len(t) and t[start] in " \t":
        start += 1
    if start >= len(t):
        return ""
    end = start
    while end < len(t) and end < start + 200:
        if t[end] in "\r\n":
            break
        end += 1
    return t[start:end].strip()


def _parse_saluda_atentamente(content_extract):
    """
    Extrae del extracto el bloque tras «Saluda atentamente,» hasta «PROPAMAT».
    Corresponde a la firma: nombre y cargo (ej. Claudio Simonetti / Administrador de Contrato).
    Sirve para cartas de respuesta donde puede firmar otra persona.
    """
    if not content_extract or not isinstance(content_extract, str):
        return ""
    t = content_extract
    tu = t.upper()
    idx = tu.find("SALUDA ATENTAMENTE")
    if idx < 0:
        return ""
    # Empezar después de la frase (y coma si existe)
    start = idx + len("SALUDA ATENTAMENTE")
    while start < len(t) and t[start] in " ,\t\r\n":
        start += 1
    if start >= len(t):
        return ""
    # Buscar «PROPAMAT» como fin del bloque (nombre de empresa)
    end_mark = tu.find("PROPAMAT", start)
    if end_mark >= 0:
        end = end_mark
    else:
        end = len(t)
    block = t[start:end]
    # Limpiar: quitar líneas vacías al final, normalizar espacios internos
    lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
    return " ".join(lines) if lines else ""


# Meses en español (clave en mayúsculas sin tildes para comparar)
_MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


def _parse_fecha_envio(content_extract):
    """
    Extrae del extracto una fecha tipo «Santiago, 26 de FEBRERO de 2026» y la devuelve como date.
    Busca el patrón: día (1-31) + " de " + mes (nombre) + " de " + año (4 dígitos).
    Devuelve date o None si no se encuentra o no es válida.
    """
    if not content_extract or not isinstance(content_extract, str):
        return None
    t = content_extract.strip()
    # Patrón: opcional ciudad + coma + espacios, luego día, " de ", mes, " de ", año
    m = re.search(
        r"(\d{1,2})\s+de\s+([A-Za-zÁ-Úá-ú]+)\s+de\s+(\d{4})\b",
        t,
        re.IGNORECASE,
    )
    if not m:
        return None
    day = int(m.group(1))
    month_name = m.group(2).upper().replace("Í", "I").replace("É", "E").replace("Ú", "U").replace("Á", "A").replace("Ó", "O")
    year = int(m.group(3))
    month = _MESES.get(month_name)
    if not month:
        return None
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _normalize_car_code(nombre_archivo_or_code):
    """
    Extrae y normaliza el código CAR desde el nombre de archivo (ej. ODA-BUF-CC-CAR-0005.pdf).
    Devuelve algo como ODA-BUF-CC-CAR-0005 para poder hacer match con referencias en otros extractos.
    """
    if not nombre_archivo_or_code:
        return ""
    s = os.path.splitext(nombre_archivo_or_code)[0].strip().upper()
    # Si ya parece un código CAR (contiene CAR y dígitos), normalizar número a 4 dígitos
    m = re.search(r"(.*-CAR-)(\d+)$", s, re.IGNORECASE)
    if m:
        prefix, num = m.group(1).upper(), m.group(2)
        return prefix + num.zfill(4)
    return s


def _find_car_references_in_text(text):
    """
    Busca en el texto referencias a códigos de carta tipo ODA-BUF-CC-CAR-XXXX.
    Devuelve set de códigos normalizados (ODA-BUF-CC-CAR-0005, etc.) a los que se hace referencia.
    """
    if not text:
        return set()
    refs = set()
    # Coincidir ODA-BUF-CC-CAR-5, ODA-BUF-CC-CAR-0005, etc.
    for m in re.finditer(r"ODA-BUF-CC-CAR-0*(\d+)", text, re.IGNORECASE):
        refs.add("ODA-BUF-CC-CAR-" + m.group(1).zfill(4))
    return refs


def _build_respuesta_map():
    """
    Construye un mapa: código CAR referenciado -> extracto de la carta de respuesta.
    Solo documentos en carpetas de la contraparte ODATA: ODATA-ST01-F5-TTAL-PPT.
    En el extracto del adjunto se busca si referencian alguna CAR (ej. ODA-BUF-CC-CAR-0005).
    """
    from .models import Folder
    folders_odata = Folder.objects.filter(code__icontains="ODATA-ST01-F5-TTAL-PPT")
    docs_odata = (
        Document.objects
        .filter(folder__in=folders_odata, attachments__file__icontains="CAR")
        .distinct()
        .prefetch_related("attachments")
    )
    out = {}
    for d in docs_odata:
        for att in d.attachments.all():
            if "CAR" not in (att.file.name or "").upper():
                continue
            extract = (att.extracted_text or "") or (d.content_extract or "")
            for ref in _find_car_references_in_text(extract):
                if ref not in out:
                    out[ref] = extract
    return out


class CustomLoginView(LoginView):
    """Vista de login con template propio (misma lógica que mantto)."""
    template_name = 'login.html'
    redirect_authenticated_user = True

    def get_success_url(self):
        next_url = self.request.GET.get('next')
        if next_url and next_url != '/':
            return next_url
        return '/'


def custom_logout(request):
    """Cerrar sesión y redirigir al login."""
    logout(request)
    messages.success(request, 'Has cerrado sesión correctamente.')
    return redirect('login')


def _log_rows_stats(rows):
    """A partir de filas de logs (referencia, descripcion, documento_archivo), devuelve conteos por tipo."""
    total = len(rows)
    ultimo_transmittal = (rows[0].get("transmittal") or "").strip() if rows else ""
    text_key = lambda r: " ".join([
        str(r.get("referencia") or ""),
        str(r.get("descripcion") or ""),
        str(r.get("documento_archivo") or ""),
    ]).lower()
    acreditaciones = 0
    acreditacion_personal = 0
    acreditacion_equipos = 0
    procedimientos = 0
    for r in rows:
        t = text_key(r)
        if "acreditacion" in t or "acreditación" in t:
            acreditaciones += 1
            if "personal" in t or "trabajador" in t or "trabajadores" in t:
                acreditacion_personal += 1
            if "equipo" in t or "equipos" in t:
                acreditacion_equipos += 1
        if "procedimiento" in t:
            procedimientos += 1
    return {
        "total": total,
        "ultimo_transmittal": ultimo_transmittal[:50] if ultimo_transmittal else "—",
        "acreditaciones": acreditaciones,
        "acreditacion_personal": acreditacion_personal,
        "acreditacion_equipos": acreditacion_equipos,
        "procedimientos": procedimientos,
    }


@login_required
def pizarra(request):
    """Pizarra: sección por defecto al abrir el sitio. Aquí irán datos de logs, cartas y estadísticas."""
    rows_cartas = _get_cartas_status_rows()
    total = len(rows_cartas)
    propamat_a_odata = sum(1 for r in rows_cartas if "ODATA-ST01-F5-TTAL-PPT" in (r.get("transmittal") or ""))
    odata_a_propamat = sum(1 for r in rows_cartas if "TRN" in (r.get("transmittal") or "") and "ODATA-ST01-F5-TTAL-PPT" not in (r.get("transmittal") or ""))
    requiere_respuesta = sum(1 for r in rows_cartas if r.get("requiere_respuesta") == "SI")
    no_respondidas = sum(1 for r in rows_cartas if r.get("requiere_respuesta") == "SI" and not (r.get("respuesta") or "").strip())
    si_respondidas = sum(1 for r in rows_cartas if r.get("requiere_respuesta") == "SI" and (r.get("respuesta") or "").strip())

    rows_propamat = _get_logs_folder_rows("TRN")
    rows_odata = _get_logs_folder_rows("Odata")
    logs_propamat = _log_rows_stats(rows_propamat)
    logs_odata = _log_rows_stats(rows_odata)

    # Estadísticas RDI (mismo universo, estados y tabla por disciplina que el panel staff)
    try:
        now = timezone.now()
        rdi_universe = _rdi_panel_universe()
        rdi_total = rdi_universe.count()
        status_labels = dict(RDIRecord._meta.get_field("status").choices)
        rdi_stats = []
        for row in (
            rdi_universe.exclude(status__in=RDI_PANEL_STATUS_EXCLUDE)
            .values("status")
            .annotate(c=Count("id"))
            .order_by("-c", "status")
        ):
            code = row["status"]
            rdi_stats.append(
                {
                    "code": code,
                    "label": status_labels.get(code, code),
                    "count": row["c"],
                }
            )
        rdi_discipline_rows = _rdi_discipline_rows(rdi_universe, now)
    except Exception:
        # Si RDI aún no está instalado en este entorno, no rompemos la pizarra.
        rdi_total = 0
        rdi_stats = []
        rdi_discipline_rows = []

    return render(request, "pizarra.html", {
        "cartas_total": total,
        "cartas_propamat_a_odata": propamat_a_odata,
        "cartas_odata_a_propamat": odata_a_propamat,
        "cartas_requieren_respuesta": requiere_respuesta,
        "cartas_no_respondidas": no_respondidas,
        "cartas_si_respondidas": si_respondidas,
        "logs_propamat": logs_propamat,
        "logs_odata": logs_odata,
        "rdi_total": rdi_total,
        "rdi_stats": rdi_stats,
        "rdi_discipline_rows": rdi_discipline_rows,
    })


@login_required
def dashboard(request):
    """Panel principal del proyecto (solo para `staff`)."""
    if not request.user.is_staff:
        from django.core.exceptions import PermissionDenied

        raise PermissionDenied

    user_model = get_user_model()
    # Usuarios "en línea": actividad reciente (últimos 10 minutos)
    now = timezone.now()
    online_window_minutes = 10
    cutoff = now - timezone.timedelta(minutes=online_window_minutes)
    presences = (
        UserPresence.objects.select_related("user")
        .filter(last_seen__gte=cutoff)
        .order_by("-last_seen")
    )
    online_users_count = presences.count()
    online_staff_users_count = presences.filter(user__is_staff=True).count()
    online_users = [p.user for p in presences[:200]]

    session_logs = (
        UserSessionLog.objects.select_related("user")
        .order_by("-occurred_at")[:200]
    )

    # Estadísticas RDI (panel): universo = actividad en 2026 en adelante
    rdi_universe = _rdi_panel_universe()
    rdi_total = rdi_universe.count()
    status_labels = dict(RDIRecord._meta.get_field("status").choices)
    rdi_stats_status = []
    for row in (
        rdi_universe.exclude(status__in=RDI_PANEL_STATUS_EXCLUDE)
        .values("status")
        .annotate(c=Count("id"))
        .order_by("-c", "status")
    ):
        code = row["status"]
        rdi_stats_status.append(
            {"label": status_labels.get(code, code), "code": code, "count": row["c"]}
        )

    rdi_discipline_rows = _rdi_discipline_rows(rdi_universe, now)

    return render(
        request,
        "dashboard.html",
        {
            "online_users_count": online_users_count,
            "online_staff_users_count": online_staff_users_count,
            "online_users": online_users,
            "online_window_minutes": online_window_minutes,
            "session_logs": session_logs,
            "rdi_total": rdi_total,
            "rdi_stats_status": rdi_stats_status,
            "rdi_discipline_rows": rdi_discipline_rows,
        },
    )


@login_required
@require_GET
def trazabilidad_view(request):
    """
    Trazabilidad en carpetas ODATA-ST01-F5-TTAL-PPT-*, ODATA-BUF-* y TRN-PRO-CM-TRN-*.
    Agrupa por título base; orden del trazo por fecha/hora de creación del documento.
    """
    q = request.GET.get("q", "").strip()
    journeys, summary = build_journeys_for_query(q)
    return render(
        request,
        "documents/trazabilidad.html",
        {
            "q": q,
            "journeys": journeys,
            "summary": summary,
        },
    )


# ---------- Listado de documentos (CRUD: listar, ver, borrar) ----------

@login_required
@require_GET
def document_list(request):
    """Listado de documentos. Si piden JSON (AJAX), usa búsqueda unificada (FTS) y devuelve lista con hits/snippets."""
    q = request.GET.get('q', '').strip()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        # Búsqueda unificada (PostgreSQL FTS) cuando hay consulta; si no hay q, listar todos
        if q:
            result = search_unified(q, limit=500)
            docs_for_list = result['documents']
        else:
            docs_for_list = list(
                Document.objects.select_related('project', 'company', 'process', 'doc_type', 'folder')
                .order_by('-date', '-created_at')[:500]
            )
        doc_map = {d.pk: d for d in docs_for_list}
        qs_json = Document.objects.filter(pk__in=[d.pk for d in docs_for_list]).prefetch_related('attachments')
        terms = _normalize_terms(q)
        use_multi = len(terms) > 1
        docs = []
        for d in qs_json:
            hits = []
            if q and d.content_extract:
                if use_multi and text_matches_all_terms_as_words(d.content_extract, terms):
                    snippets = extract_snippets_multi_term(d.content_extract, terms, context_words=10, max_snippets=5)
                elif not use_multi and text_matches_single_query(d.content_extract, q):
                    snippets = extract_snippets(d.content_extract, q, context_words=10, max_snippets=5)
                else:
                    snippets = []
                if snippets:
                    file_name = d.file.name if d.file else 'Documento principal'
                    qa = _extract_qa_from_delimiters(
                        d.content_extract,
                        query=None if use_multi else q,
                        terms=terms if use_multi else [],
                    )
                    pregunta = qa.get("pregunta", "")
                    respuesta = qa.get("respuesta", "")
                    especialidad = ""
                    # Fallback por si el contenido NO viene en formato delimitado fijo
                    if not respuesta:
                        respuesta = _extract_respuesta_from_extracted_text("\n".join(snippets))
                    if not respuesta:
                        respuesta = _extract_respuesta_from_extracted_text(d.content_extract)
                    hits.append({
                        'source': 'document',
                        'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                        'file_url': d.file.url if d.file else None,
                        'snippets': snippets,
                        'pregunta': pregunta,
                        'respuesta': respuesta,
                        'especialidad': especialidad,
                    })
            for att in d.attachments.all():
                if q and att.extracted_text:
                    if use_multi and text_matches_all_terms_as_words(att.extracted_text, terms):
                        snippets = extract_snippets_multi_term(att.extracted_text, terms, context_words=10, max_snippets=5)
                    elif not use_multi and text_matches_single_query(att.extracted_text, q):
                        snippets = extract_snippets(att.extracted_text, q, context_words=10, max_snippets=5)
                    else:
                        snippets = []
                    if snippets:
                        name = att.file.name.split('/')[-1] if att.file.name else 'Adjunto'
                        hits.append({
                            'source': 'attachment',
                            'file_name': name,
                            'file_url': att.file.url if att.file else None,
                            'snippets': snippets,
                        })
            docs.append({
                'id': d.id,
                'code': d.code,
                'title': d.title or '',
                'status': d.status,
                'status_display': d.get_status_display(),
                'date': d.date.isoformat() if d.date else '',
                'revision': d.revision or '',
                'project': d.project.code,
                'company': d.company.code,
                'process': d.process.code,
                'doc_type': d.doc_type.code,
                'file_url': d.file.url if d.file else None,
                'has_file': bool(d.file),
                'folder_id': d.folder_id,
                'folder_code': d.folder.code if d.folder else None,
                'hits': hits,
            })
        return JsonResponse({'documents': docs})
    # HTML: listado inicial (sin búsqueda)
    qs = Document.objects.select_related('project', 'company', 'process', 'doc_type', 'folder').order_by('-date', '-created_at')
    return render(request, 'documents/document_list.html', {'document_list': qs[:100]})


@login_required
@require_GET
def contrato_view(request):
    """
    Sección Contrato: clon de Documentos pero restringido a la carpeta/documento 'contrato'.
    - HTML: muestra una tabla igual a Documentos pero solo con el documento de la carpeta 'contrato'.
    - JSON: búsqueda por contenido solo dentro del documento principal y sus adjuntos de esa carpeta.
    """
    q = request.GET.get('q', '').strip()

    def _folder_scoped_qs(folder_code: str):
        folder = Folder.objects.filter(code__iexact=folder_code).first()
        qs = Document.objects.select_related('project', 'company', 'process', 'doc_type', 'folder').prefetch_related('attachments')
        if folder:
            qs = qs.filter(folder=folder)
        else:
            qs = qs.none()
        return folder, qs

    def _folder_scoped_json(qs, query: str):
        docs_for_list = list(qs.order_by('-date', '-created_at')[:10])
        terms = _normalize_terms(query)
        use_multi = len(terms) > 1
        docs = []
        for d in docs_for_list:
            hits = []
            if query and d.content_extract:
                # 1) Modo simple Excel: buscar QA dentro de delimitadores ¿... ? @ ... @
                if use_multi:
                    qa_list = _extract_all_qa_from_delimiters(
                        d.content_extract,
                        query=None,
                        terms=terms,
                        max_pairs=8,
                    )
                    if qa_list:
                        file_name = d.file.name if d.file else 'Documento principal'
                        for qa in qa_list:
                            hits.append({
                                'source': 'document',
                                'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                                'file_url': d.file.url if d.file else None,
                                'snippets': [],
                                'pregunta': qa.get("pregunta", ""),
                                'respuesta': qa.get("respuesta", ""),
                                'especialidad': qa.get("especialidad", ""),
                                'gc': qa.get("gc", ""),
                            })
                    else:
                        # 2) Fallback: snippets tradicionales (PDF/DOCX u otros)
                        if text_matches_all_terms_as_words(d.content_extract, terms):
                            snippets = extract_snippets_multi_term(d.content_extract, terms, context_words=10, max_snippets=8)
                        else:
                            snippets = []
                        if snippets:
                            file_name = d.file.name if d.file else 'Documento principal'
                            hits.append({
                                'source': 'document',
                                'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                                'file_url': d.file.url if d.file else None,
                                'snippets': snippets,
                            })
                else:
                    qa = _extract_qa_from_delimiters(
                        d.content_extract,
                        query=query,
                        terms=[],
                    )
                    if qa.get("pregunta") and qa.get("respuesta"):
                        file_name = d.file.name if d.file else 'Documento principal'
                        hits.append({
                            'source': 'document',
                            'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                            'file_url': d.file.url if d.file else None,
                            'snippets': [],
                            'pregunta': qa.get("pregunta", ""),
                            'respuesta': qa.get("respuesta", ""),
                            'especialidad': qa.get("especialidad", ""),
                            'gc': qa.get("gc", ""),
                        })
                    else:
                        # 2) Fallback: snippets tradicionales (PDF/DOCX u otros)
                        if text_matches_single_query(d.content_extract, query):
                            snippets = extract_snippets(d.content_extract, query, context_words=10, max_snippets=8)
                        else:
                            snippets = []
                        if snippets:
                            file_name = d.file.name if d.file else 'Documento principal'
                            hits.append({
                                'source': 'document',
                                'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                                'file_url': d.file.url if d.file else None,
                                'snippets': snippets,
                            })
            for att in d.attachments.all():
                if query and att.extracted_text:
                    # 1) Modo simple Excel
                    if use_multi:
                        qa_list = _extract_all_qa_from_delimiters(
                            att.extracted_text,
                            query=None,
                            terms=terms,
                            max_pairs=8,
                        )
                        if qa_list:
                            name = att.file.name.split('/')[-1] if att.file and att.file.name else 'Adjunto'
                            for qa in qa_list:
                                hits.append({
                                    'source': 'attachment',
                                    'file_name': name,
                                    'file_url': att.file.url if att.file else None,
                                    'snippets': [],
                                    'pregunta': qa.get("pregunta", ""),
                                    'respuesta': qa.get("respuesta", ""),
                                    'especialidad': qa.get("especialidad", ""),
                                    'gc': qa.get("gc", ""),
                                })
                        else:
                            # 2) Fallback snippets
                            if text_matches_all_terms_as_words(att.extracted_text, terms):
                                snippets = extract_snippets_multi_term(att.extracted_text, terms, context_words=10, max_snippets=8)
                            else:
                                snippets = []
                            if snippets:
                                name = att.file.name.split('/')[-1] if att.file and att.file.name else 'Adjunto'
                                hits.append({
                                    'source': 'attachment',
                                    'file_name': name,
                                    'file_url': att.file.url if att.file else None,
                                    'snippets': snippets,
                                })
                    else:
                        qa = _extract_qa_from_delimiters(
                            att.extracted_text,
                            query=query,
                            terms=[],
                        )
                        if qa.get("pregunta") and qa.get("respuesta"):
                            name = att.file.name.split('/')[-1] if att.file and att.file.name else 'Adjunto'
                            hits.append({
                                'source': 'attachment',
                                'file_name': name,
                                'file_url': att.file.url if att.file else None,
                                'snippets': [],
                                'pregunta': qa.get("pregunta", ""),
                                'respuesta': qa.get("respuesta", ""),
                                'especialidad': qa.get("especialidad", ""),
                                'gc': qa.get("gc", ""),
                            })
                        else:
                            # 2) Fallback snippets
                            if text_matches_single_query(att.extracted_text, query):
                                snippets = extract_snippets(att.extracted_text, query, context_words=10, max_snippets=8)
                            else:
                                snippets = []
                            if snippets:
                                name = att.file.name.split('/')[-1] if att.file and att.file.name else 'Adjunto'
                                hits.append({
                                    'source': 'attachment',
                                    'file_name': name,
                                    'file_url': att.file.url if att.file else None,
                                    'snippets': snippets,
                                })
            docs.append({
                'id': d.id,
                'code': d.code,
                'title': d.title or '',
                'status': d.status,
                'status_display': d.get_status_display(),
                'date': d.date.isoformat() if d.date else '',
                'revision': d.revision or '',
                'project': d.project.code,
                'company': d.company.code,
                'process': d.process.code,
                'doc_type': d.doc_type.code,
                'file_url': d.file.url if d.file else None,
                'has_file': bool(d.file),
                'folder_id': d.folder_id,
                'folder_code': d.folder.code if d.folder else None,
                'hits': hits,
            })
        return JsonResponse({'documents': docs})

    contrato_folder, contrato_qs = _folder_scoped_qs("contrato")
    consolidado_folder, consolidado_qs = _folder_scoped_qs("consolidado")

    # JSON (AJAX) solo para contrato en esta ruta
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        return _folder_scoped_json(contrato_qs, q)

    # HTML inicial: mostrar documentos de contrato y consolidado
    return render(
        request,
        'documents/contrato.html',
        {
            'document_list': contrato_qs.order_by('-date', '-created_at')[:10],
            'contrato_folder': contrato_folder,
            'consolidado_document_list': consolidado_qs.order_by('-date', '-created_at')[:10],
            'consolidado_folder': consolidado_folder,
        }
    )


@login_required
@require_GET
def consolidado_view(request):
    """
    Endpoint JSON/HTML para la carpeta/documento 'consolidado'.
    Se usa desde la sección Contrato para buscar solo dentro de esa carpeta.
    """
    q = request.GET.get('q', '').strip()
    consolidado_folder = Folder.objects.filter(code__iexact="consolidado").first()
    base_qs = Document.objects.select_related('project', 'company', 'process', 'doc_type', 'folder').prefetch_related('attachments')
    if consolidado_folder:
        base_qs = base_qs.filter(folder=consolidado_folder)
    else:
        base_qs = base_qs.none()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        docs_for_list = list(base_qs.order_by('-date', '-created_at')[:10])
        terms = _normalize_terms(q)
        use_multi = len(terms) > 1
        docs = []
        for d in docs_for_list:
            hits = []
            if q and d.content_extract:
                if use_multi:
                    qa_list = _extract_all_qa_from_delimiters(
                        d.content_extract,
                        query=None,
                        terms=terms,
                        max_pairs=8,
                    )
                    if qa_list:
                        file_name = d.file.name if d.file else 'Documento principal'
                        for qa in qa_list:
                            hits.append({
                                'source': 'document',
                                'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                                'file_url': d.file.url if d.file else None,
                                'snippets': [],
                                'pregunta': qa.get("pregunta", ""),
                                'respuesta': qa.get("respuesta", ""),
                                'especialidad': qa.get("especialidad", ""),
                                'gc': qa.get("gc", ""),
                            })
                    else:
                        if text_matches_all_terms_as_words(d.content_extract, terms):
                            snippets = extract_snippets_multi_term(d.content_extract, terms, context_words=10, max_snippets=8)
                        else:
                            snippets = []
                        if snippets:
                            file_name = d.file.name if d.file else 'Documento principal'
                            hits.append({
                                'source': 'document',
                                'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                                'file_url': d.file.url if d.file else None,
                                'snippets': snippets,
                            })
                else:
                    qa = _extract_qa_from_delimiters(
                        d.content_extract,
                        query=q,
                        terms=[],
                    )
                    if qa.get("pregunta") and qa.get("respuesta"):
                        file_name = d.file.name if d.file else 'Documento principal'
                        hits.append({
                            'source': 'document',
                            'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                            'file_url': d.file.url if d.file else None,
                            'snippets': [],
                            'pregunta': qa.get("pregunta", ""),
                            'respuesta': qa.get("respuesta", ""),
                            'especialidad': qa.get("especialidad", ""),
                            'gc': qa.get("gc", ""),
                        })
                    else:
                        if text_matches_single_query(d.content_extract, q):
                            snippets = extract_snippets(d.content_extract, q, context_words=10, max_snippets=8)
                        else:
                            snippets = []
                        if snippets:
                            file_name = d.file.name if d.file else 'Documento principal'
                            hits.append({
                                'source': 'document',
                                'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                                'file_url': d.file.url if d.file else None,
                                'snippets': snippets,
                            })
            for att in d.attachments.all():
                if q and att.extracted_text:
                    if use_multi:
                        qa_list = _extract_all_qa_from_delimiters(
                            att.extracted_text,
                            query=None,
                            terms=terms,
                            max_pairs=8,
                        )
                        if qa_list:
                            name = att.file.name.split('/')[-1] if att.file and att.file.name else 'Adjunto'
                            for qa in qa_list:
                                hits.append({
                                    'source': 'attachment',
                                    'file_name': name,
                                    'file_url': att.file.url if att.file else None,
                                    'snippets': [],
                                    'pregunta': qa.get("pregunta", ""),
                                    'respuesta': qa.get("respuesta", ""),
                                    'especialidad': qa.get("especialidad", ""),
                                    'gc': qa.get("gc", ""),
                                })
                        else:
                            if text_matches_all_terms_as_words(att.extracted_text, terms):
                                snippets = extract_snippets_multi_term(att.extracted_text, terms, context_words=10, max_snippets=8)
                            else:
                                snippets = []
                            if snippets:
                                name = att.file.name.split('/')[-1] if att.file and att.file.name else 'Adjunto'
                                hits.append({
                                    'source': 'attachment',
                                    'file_name': name,
                                    'file_url': att.file.url if att.file else None,
                                    'snippets': snippets,
                                })
                    else:
                        qa = _extract_qa_from_delimiters(
                            att.extracted_text,
                            query=q,
                            terms=[],
                        )
                        if qa.get("pregunta") and qa.get("respuesta"):
                            name = att.file.name.split('/')[-1] if att.file and att.file.name else 'Adjunto'
                            hits.append({
                                'source': 'attachment',
                                'file_name': name,
                                'file_url': att.file.url if att.file else None,
                                'snippets': [],
                                'pregunta': qa.get("pregunta", ""),
                                'respuesta': qa.get("respuesta", ""),
                                'especialidad': qa.get("especialidad", ""),
                                'gc': qa.get("gc", ""),
                            })
                        else:
                            if text_matches_single_query(att.extracted_text, q):
                                snippets = extract_snippets(att.extracted_text, q, context_words=10, max_snippets=8)
                            else:
                                snippets = []
                            if snippets:
                                name = att.file.name.split('/')[-1] if att.file and att.file.name else 'Adjunto'
                                hits.append({
                                    'source': 'attachment',
                                    'file_name': name,
                                    'file_url': att.file.url if att.file else None,
                                    'snippets': snippets,
                                })
            docs.append({
                'id': d.id,
                'code': d.code,
                'title': d.title or '',
                'status': d.status,
                'status_display': d.get_status_display(),
                'date': d.date.isoformat() if d.date else '',
                'revision': d.revision or '',
                'project': d.project.code,
                'company': d.company.code,
                'process': d.process.code,
                'doc_type': d.doc_type.code,
                'file_url': d.file.url if d.file else None,
                'has_file': bool(d.file),
                'folder_id': d.folder_id,
                'folder_code': d.folder.code if d.folder else None,
                'hits': hits,
            })
        return JsonResponse({'documents': docs})

    # Si alguien abre /documentos/consolidado/ en navegador, reusar la página principal de contrato
    return redirect('contrato_view')

@login_required
@require_GET
def document_detail(request, pk):
    """Detalle de un documento, archivo principal y adjuntos."""
    doc = get_object_or_404(
        Document.objects.select_related('project', 'company', 'process', 'doc_type', 'folder').prefetch_related('attachments'),
        pk=pk
    )
    return render(request, 'documents/document_detail.html', {'document': doc})


@login_required
@require_POST
def document_delete(request, pk):
    """Eliminar documento. Esperado por AJAX; devuelve JSON."""
    doc = get_object_or_404(Document, pk=pk)
    code = doc.code
    doc.delete()
    return JsonResponse({'success': True, 'message': f'Documento {code} eliminado.', 'id': int(pk)})


def _parse_emails(raw):
    """Convierte una cadena 'a@x.com, b@y.com; c@z.com' en lista de emails sin espacios."""
    if not raw or not raw.strip():
        return []
    out = []
    for part in re.split(r'[,;\n]+', raw):
        email = part.strip()
        if email and '@' in email:
            out.append(email)
    return out


# --- Plantilla correo Transmittal (ODATA-ST01-F5-TTAL-PPT-?????) ---
TRANSMITTAL_ASUNTO_TEMPLATE = "ODATA ST01 EXP 05-E2 |   {{ referencia }}   | {{ transmittal }}"
# Plantilla SharePoint: la carpeta del transmittal existe siempre en este repositorio; solo cambia el código.
# Ojo: la ruta dentro de SharePoint DEBE coincidir exacto con el link que funciona.
# En la URL correcta aparece ".../PROPAMAT/PROP a ODATA/...".
SHAREPOINT_TRANSMITTAL_BASE_PATH = "/sites/GestorDocumentalProyectoODATA/Documentos compartidos/1. ODATA ST01/4. EXP05 - FASE 5.2/4.1 Comunicaciones/4.1.1 Transmittal/PROPAMAT/PROP a ODATA/"
SHAREPOINT_TRANSMITTAL_VIEWID = "d9a3f383-3cf1-43be-85e9-c0a8cafec845"
SHAREPOINT_TRANSMITTAL_BASE_URL = (
    "https://bufferconsultores.sharepoint.com/sites/GestorDocumentalProyectoODATA/Documentos%20compartidos/Forms/AllItems.aspx"
)
TRANSMITTAL_CUERPO_TEMPLATE = """Estimados,

Junto con saludar y mediante el presente, se adjunta ruta de la documentación del asunto de la referencia.

De acuerdo con lo anterior, se pone a vuestra disposición el Transmittal "{{ transmittal_bold }}", {{ referencia }}, los cuales han sido cargados en SharePoint.

{{ transmittal_link }}

Detalle Documentos Adjuntos:

{{ documentos }}

Saludos cordiales,

MAX GONZÁLEZ PEÑA
CONTROL DE DOCUMENTOS
PROPAMAT"""


def _extract_text_from_uploaded_file(file_obj):
    """
    Extrae texto de un archivo subido (PDF o DOCX) para parsear transmittal.
    file_obj: UploadedFile con .read() y .name.
    """
    content = file_obj.read()
    if not content:
        return ""
    name = (file_obj.name or "").lower()
    try:
        if name.endswith(".pdf"):
            from pypdf import PdfReader
            from io import BytesIO
            reader = PdfReader(BytesIO(content))
            return " ".join((p.extract_text() or "").replace("\r", "\n") for p in reader.pages)
        if name.endswith(".docx"):
            from docx import Document as DocxDocument
            from io import BytesIO
            doc = DocxDocument(BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        pass
    return ""


def _parse_transmittal_extract(text, filename_hint=""):
    """
    Parsea el extracto de una hoja de transmittal y devuelve:
    - transmittal: código tipo ODATA-ST01-F5-TTAL-PPT-00093
    - referencia: texto después de "Referencia:"
    - documentos: lista de códigos de documento (ej. JLG 193022)
    """
    result = {"transmittal": "", "referencia": "", "documentos": []}
    if not text:
        text = ""
    text = text.replace("\r", "\n")
    # Código transmittal: del contenido o del nombre de archivo
    trans_re = re.compile(r"ODATA-ST01-F5-TTAL-PPT-\d+", re.IGNORECASE)
    match = trans_re.search(text)
    if match:
        result["transmittal"] = match.group(0).upper()
    if not result["transmittal"] and filename_hint:
        match = trans_re.search(filename_hint)
        if match:
            result["transmittal"] = match.group(0).upper()
    # Referencia: después de "Referencia:" hasta fin de línea o doble salto
    ref_re = re.compile(r"Referencia:\s*(.+?)(?:\n\n|\n\s*\n|$)", re.IGNORECASE | re.DOTALL)
    match = ref_re.search(text)
    if match:
        result["referencia"] = match.group(1).strip().replace("\n", " ").strip()
        if len(result["referencia"]) > 200:
            result["referencia"] = result["referencia"][:200].rsplit(" ", 1)[0]
    # Documentos: columna "Título / Descripción Documento" (ej. JLG 193022, SRT-477)
    # 1) Reducir espacios múltiples para que el regex funcione aunque el PDF meta espacios de más
    text_norm = re.sub(r"\s+", " ", text)
    lines = text.splitlines()
    seen = set()
    # Patrones: "JLG 193022" (letras + espacio(s) + dígitos) y "SRT-477" (letras + guión + dígitos)
    doc_code_re = re.compile(r"\b([A-Za-z]{2,8}\s*\d{5,})\b")
    doc_code_hyphen_re = re.compile(r"\b([A-Za-z]{2,8}-\d{2,})\b")

    def _add_doc(c):
        c = re.sub(r"\s+", " ", (c or "").strip()).strip()
        if not c or c in seen:
            return
        if "ODATA" in c.upper():
            return
        # No incluir códigos tipo PPT-00095 o PPT-000 (parte del transmittal, no nombre de documento)
        if re.match(r"^PPT-\d+$", c, re.IGNORECASE):
            return
        # No incluir "omitido" ni "emitido" (issued) ni variantes como "Emitido ,"
        c_low = c.lower().strip()
        c_norm = re.sub(r"[,.\s]+$", "", c_low).strip()
        if c_norm in ("omitido", "emitido"):
            return
        docs_list = result["documentos"]
        # No duplicar: si c ya está contenido en un nombre más largo (ej. "TDDF-80" en "Minicargador TDDF-80"), no añadir c
        for d in docs_list:
            if c != d and c in d:
                return
        # Si un ítem ya en la lista es solo un fragmento del nuevo (ej. "TDDF-80" y llegamos a "Minicargador TDDF-80"), quitar el corto
        for d in list(docs_list):
            if d != c and d in c:
                docs_list.remove(d)
                seen.discard(d)
        seen.add(c)
        docs_list.append(c)

    # Buscar bloque de tabla: entre "Título / Descripción Documento" o "Estatus" y "Emitido" / "Unidad revisora"
    bloque_docs = ""
    lower_text = text.lower()
    idx_titulo = max(lower_text.find("título / descripción documento"), lower_text.find("descripción documento"), lower_text.find("titulo / descripcion documento"))
    idx_estatus = lower_text.find("estatus")
    idx_emitido = lower_text.find("emitido para")
    idx_unidad = lower_text.find("unidad revisora")
    start = idx_titulo if idx_titulo >= 0 else idx_estatus
    if start >= 0:
        end = len(text)
        for e in (idx_emitido, idx_unidad):
            if e > start:
                end = min(end, e)
        bloque_docs = text[start:end] + " " + text_norm

    # Primero: extraer títulos/descripciones de documento de cada línea "Para revisión" (ej. "Maquinaria", "Trabajador operador")
    # En el PDF suele verse como "0Maquinaria0Para revisión" o "0Trabajador operador0Para revisión"
    for line in lines:
        if "para revisión" not in line.lower() and "para revision" not in line.lower():
            continue
        antes = re.split(r"\bpara\s+revisi[oó]n\b", line, 1, re.IGNORECASE)[0].strip()
        # Quitar código ODATA-ST01-F5-TTAL-PPT-XXXXX si está
        antes = re.sub(r"ODATA-ST01-F5-TTAL-PPT-\d+", " ", antes, flags=re.IGNORECASE)
        # Quitar ruido al inicio (dígitos/espacios de tabla)
        antes = re.sub(r"^[\d\s\.]+", "", antes).strip()
        # Quitar solo ruido de columna al final (espacio + 1-3 dígitos), no códigos como TDDF-80 o JLG 193022
        antes = re.sub(r"\s+\d{1,3}\s*$", "", antes).strip()
        antes = re.sub(r"\s+", " ", antes).strip()
        if len(antes) >= 2 and any(c.isalpha() for c in antes):
            _add_doc(antes)

    search_text = bloque_docs if bloque_docs.strip() else text_norm + " " + text
    for regex in (doc_code_re, doc_code_hyphen_re):
        for m in regex.finditer(search_text):
            _add_doc(m.group(1))

    # Por línea: si la línea contiene "Para revisión", extraer también códigos tipo JLG 193022 / SRT-477
    for line in lines:
        if "para revisión" in line.lower() or "para revision" in line.lower():
            for m in doc_code_re.finditer(line):
                _add_doc(m.group(1))
            for m in doc_code_hyphen_re.finditer(line):
                _add_doc(m.group(1))

    # Último recurso: todos los códigos tipo WORD+NUMBER en todo el texto (excluir ODATA)
    if not result["documentos"]:
        for m in doc_code_re.finditer(text_norm):
            _add_doc(m.group(1))
        for m in doc_code_hyphen_re.finditer(text_norm):
            _add_doc(m.group(1))

    # Si el PDF metió espacio entre cada carácter (ej. "J L G 1 9 3 0 2 2"), buscar sin espacios
    if not result["documentos"]:
        sin_espacios = re.sub(r"\s+", "", bloque_docs if bloque_docs.strip() else text)
        for m in re.finditer(r"[A-Za-z]{2,8}\d{5,}", sin_espacios):
            cod = m.group(0)
            if "ODATA" in cod.upper():
                continue
            # Formatear con espacio: JLG193022 -> JLG 193022
            for i in range(len(cod) - 1, 0, -1):
                if cod[i].isdigit() and cod[i - 1].isalpha():
                    cod = cod[:i] + " " + cod[i:]
                    break
            _add_doc(cod)

    return result


def _limpiar_referencia_para_asunto(ref):
    """Quita de la referencia texto como 'REV . WWW.PROPAMAT.CL' para el asunto del correo."""
    if not ref:
        return ref
    # Quitar variantes de REV . WWW.PROPAMAT.CL (con o sin punto, espacios)
    ref = re.sub(r"\s*REV\s*\.?\s*WWW\.PROPAMAT\.CL\s*", " ", ref, flags=re.IGNORECASE)
    return ref.strip()


def _referencia_solo_para_asunto(ref):
    """
    Para el asunto: solo el texto antes de "Nota" (sin incluir Nota), o si no hay Nota,
    el texto hasta antes de la tabla que comienza con "Ítem" o "Item".
    """
    if not ref or not ref.strip():
        return ref
    ref = ref.strip()
    # Primero cortar en "nota" (palabra completa) si existe
    if re.search(r"\bnota\b", ref, re.IGNORECASE):
        ref = re.split(r"\bnota\b", ref, 1, re.IGNORECASE)[0].strip()
    # Luego cortar en "ítem" o "item" (inicio de tabla) si existe
    if re.search(r"\b(?:ítem|item)\b", ref, re.IGNORECASE):
        ref = re.split(r"\b(?:ítem|item)\b", ref, 1, re.IGNORECASE)[0].strip()
    return ref


def _build_asunto_transmittal(data):
    """Sustituye en la plantilla de asunto solo {{ referencia }} y {{ transmittal }}. Sin listado de documentos."""
    ref = (data.get("referencia") or "").strip()
    ref = _limpiar_referencia_para_asunto(ref)
    ref = _referencia_solo_para_asunto(ref)
    ref = ref.upper()
    # Una sola línea; límite de longitud
    if "\n" in ref:
        ref = ref.split("\n")[0].strip()
    if len(ref) > 120:
        ref = ref[:120].rsplit(" ", 1)[0]
    trans = (data.get("transmittal") or "").strip()
    return (
        TRANSMITTAL_ASUNTO_TEMPLATE.replace("{{ referencia }}", ref).replace("{{ transmittal }}", trans)
    )


def _build_cuerpo_transmittal(data, request=None):
    """Sustituye en la plantilla de cuerpo: {{ transmittal }}, {{ referencia }}, {{ documentos }}.
    Documentos como tabla de una columna. Quita " Rev . www.propamat.cl" de la referencia.
    Si request se pasa, devuelve HTML con el transmittal como link a la carpeta."""
    trans = (data.get("transmittal") or "").strip()
    ref = _limpiar_referencia_para_asunto((data.get("referencia") or "").strip())
    docs = data.get("documentos") or []
    if docs:
        documentos_bloque = "\n".join("• " + d for d in docs)
    else:
        documentos_bloque = "• (no detectados)"

    # En el párrafo: código en negrita y sin link. En la línea previa a Detalle: link a SharePoint de la carpeta del transmittal.
    if request and trans:
        transmittal_bold = "<strong>%s</strong>" % trans
        path = SHAREPOINT_TRANSMITTAL_BASE_PATH + trans
        id_param = quote(path, safe="")
        transmittal_link = '<a href="%s?id=%s&viewid=%s&p=true&startedResponseCatch=true">%s</a>' % (
            SHAREPOINT_TRANSMITTAL_BASE_URL,
            id_param,
            quote(SHAREPOINT_TRANSMITTAL_VIEWID, safe=""),
            trans,
        )
    else:
        transmittal_bold = trans
        transmittal_link = trans

    body = (
        TRANSMITTAL_CUERPO_TEMPLATE.replace("{{ transmittal_bold }}", transmittal_bold)
        .replace("{{ transmittal_link }}", transmittal_link)
        .replace("{{ referencia }}", ref)
        .replace("{{ documentos }}", documentos_bloque)
    )
    # Importante:
    # - En el UI (textarea) queremos mostrar saltos de línea como texto normal (sin etiquetas <br>).
    # - Para que el email renderice bien, convertimos \n a <br> justo antes de enviar.
    return body


@login_required
@require_POST
def extraer_transmittal_ajax(request):
    """
    AJAX: recibe un archivo (PDF/DOCX) de transmittal, extrae texto, parsea y devuelve
    asunto y cuerpo para rellenar el formulario antes de enviar.
    """
    archivo = request.FILES.get("archivo")
    if not archivo or not archivo.name:
        return JsonResponse({"ok": False, "error": "No se envió ningún archivo."}, status=400)
    try:
        text = _extract_text_from_uploaded_file(archivo)
        data = _parse_transmittal_extract(text, archivo.name)
        asunto = _build_asunto_transmittal(data)
        # Importante: pasamos `request` para que _build_cuerpo_transmittal genere HTML
        # (transmittal con <strong> y enlace <a href=...> hacia SharePoint).
        cuerpo = _build_cuerpo_transmittal(data, request=request)
        return JsonResponse({
            "ok": True,
            "asunto": asunto,
            "cuerpo": cuerpo,
            "transmittal": data.get("transmittal") or "",
            "referencia": data.get("referencia") or "",
            "documentos": data.get("documentos") or [],
        })
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


@login_required
def enviar_correo_view(request):
    """
    Formulario para enviar correo desde el sistema: destinatario(s), CC, asunto y cuerpo.
    Envío desde la cuenta configurada (EMAIL_HOST_USER). Se guarda registro en CorreoEnviado.
    """
    if request.method == 'GET':
        cc_grupos = GrupoCorreo.objects.filter(activo=True).order_by('nombre')
        ctx = {
            'email_from': getattr(settings, 'EMAIL_HOST_USER', ''),
            'cc_grupos': cc_grupos,
        }
        # Enlace desde Informar documento: ?doc=<pk>  |  desde BIM: ?rdi=<csv_id>
        doc_pk = request.GET.get('doc')
        ctx['documento_informar_id'] = ''
        ctx['rdi_informar_csv_id'] = ''
        if doc_pk:
            try:
                pref = Document.objects.select_related('folder').get(pk=int(doc_pk))
                ctx['documento_informar_id'] = str(pref.pk)
                link_doc = request.build_absolute_uri(reverse('document_detail', args=[pref.pk]))
                folder_bit = (pref.folder.code if pref.folder else '') or '—'
                ctx['destinatarios'] = ''
                ctx['asunto'] = f"{pref.code}"
                ctx['cuerpo'] = (
                    f"Documento: {pref.code}\n"
                    f"Carpeta / transmittal: {folder_bit}\n"
                    f"Ver en Condocdat: {link_doc}\n"
                )
            except (ValueError, Document.DoesNotExist):
                pass
        elif request.GET.get('rdi'):
            try:
                csv_id = int(request.GET['rdi'])
                rdi = RDIRecord.objects.select_related('last_import').get(csv_id=csv_id)
                ctx['rdi_informar_csv_id'] = str(rdi.csv_id)
                ctx['destinatarios'] = ''
                tit = (rdi.title or '').strip()
                ctx['asunto'] = f"RDI (BIM) #{rdi.csv_id}" + (f" — {tit[:120]}" if tit else "")
                ctx['cuerpo'] = _cuerpo_correo_desde_rdi_record(rdi, request)
            except (ValueError, RDIRecord.DoesNotExist):
                pass
        return render(request, 'documents/enviar_correo.html', ctx)

    # POST
    cc_grupos = GrupoCorreo.objects.filter(activo=True).order_by('nombre')
    usar_plantilla = request.POST.get('usar_plantilla_transmittal') == 'on'
    documento_informar_id = (request.POST.get('documento_informar_id') or '').strip()
    rdi_informar_csv_id = (request.POST.get('rdi_informar_csv_id') or '').strip()
    destinatarios_raw = (request.POST.get('destinatarios') or '').strip()
    copia_raw = (request.POST.get('copia') or '').strip()
    destinatario_grupos_ids = request.POST.getlist('destinatario_grupos')
    cc_grupos_ids = request.POST.getlist('cc_grupos')
    # Siempre se envía lo editado en el formulario (asunto/cuerpo). La plantilla solo rellena vía JS al elegir archivo.
    asunto = (request.POST.get('asunto') or '').strip()
    cuerpo = (request.POST.get('cuerpo') or '').strip()

    # Adjuntos para el correo (no se re-extrae transmittal en servidor: prevalecen las correcciones del usuario).
    adjuntos_list = []
    for f in request.FILES.getlist('adjuntos'):
        if f and f.name:
            contenido = f.read()
            adjuntos_list.append((f.name, contenido, getattr(f, 'content_type', None) or 'application/octet-stream'))

    to_list = list(_parse_emails(destinatarios_raw))
    for g in GrupoCorreo.objects.filter(pk__in=destinatario_grupos_ids, activo=True):
        for e in g.lista_emails():
            if e not in to_list:
                to_list.append(e)
    cc_list = list(_parse_emails(copia_raw))
    for g in GrupoCorreo.objects.filter(pk__in=cc_grupos_ids, activo=True):
        for e in g.lista_emails():
            if e not in cc_list:
                cc_list.append(e)

    if not to_list:
        messages.error(request, 'Indique al menos un destinatario (email).')
        return render(request, 'documents/enviar_correo.html', {
            'email_from': getattr(settings, 'EMAIL_HOST_USER', ''),
            'cc_grupos': cc_grupos,
            'destinatarios': destinatarios_raw,
            'copia': copia_raw,
            'asunto': asunto,
            'cuerpo': cuerpo,
            'usar_plantilla_transmittal': usar_plantilla,
            'documento_informar_id': documento_informar_id,
            'rdi_informar_csv_id': rdi_informar_csv_id,
        })
    if not asunto:
        messages.error(request, 'El asunto no puede estar vacío.')
        return render(request, 'documents/enviar_correo.html', {
            'email_from': getattr(settings, 'EMAIL_HOST_USER', ''),
            'cc_grupos': cc_grupos,
            'destinatarios': destinatarios_raw,
            'copia': copia_raw,
            'asunto': asunto,
            'cuerpo': cuerpo,
            'usar_plantilla_transmittal': usar_plantilla,
            'documento_informar_id': documento_informar_id,
            'rdi_informar_csv_id': rdi_informar_csv_id,
        })

    _email_pw = (getattr(settings, 'EMAIL_HOST_PASSWORD', None) or '').strip()
    if not _email_pw:
        messages.error(
            request,
            'No está configurada la contraseña de correo (EMAIL_HOST_PASSWORD en el entorno). '
            'Configure la variable de entorno y reinicie el servidor.'
        )
        return render(request, 'documents/enviar_correo.html', {
            'email_from': getattr(settings, 'EMAIL_HOST_USER', ''),
            'cc_grupos': cc_grupos,
            'destinatarios': destinatarios_raw,
            'copia': copia_raw,
            'asunto': asunto,
            'cuerpo': cuerpo,
            'usar_plantilla_transmittal': usar_plantilla,
            'documento_informar_id': documento_informar_id,
            'rdi_informar_csv_id': rdi_informar_csv_id,
        })

    if usar_plantilla and not adjuntos_list:
        messages.warning(request, 'Para usar la plantilla transmittal debe adjuntar al menos un archivo (PDF o DOCX).')
        return render(request, 'documents/enviar_correo.html', {
            'email_from': getattr(settings, 'EMAIL_HOST_USER', ''),
            'cc_grupos': cc_grupos,
            'destinatarios': destinatarios_raw,
            'copia': copia_raw,
            'asunto': asunto,
            'cuerpo': cuerpo,
            'usar_plantilla_transmittal': True,
            'documento_informar_id': documento_informar_id,
            'rdi_informar_csv_id': rdi_informar_csv_id,
        })

    adjuntos_nombres = [t[0] for t in adjuntos_list]

    registro = CorreoEnviado(
        destinatarios=destinatarios_raw,
        copia=copia_raw,
        asunto=asunto,
        cuerpo=cuerpo,
        adjuntos_nombres=', '.join(adjuntos_nombres) if adjuntos_nombres else '',
        enviado_por=request.user,
    )
    try:
        msg = EmailMessage(
            subject=asunto,
            body=cuerpo,
            from_email=settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER,
            to=to_list,
            cc=cc_list,
        )
        # Si el cuerpo contiene HTML (anchor/link), lo enviamos como HTML.
        # Para evitar que el usuario vea <br> en el textarea, convertimos \n -> <br>
        # únicamente en el envío (y solo si aún no hay <br>).
        if "<a href=" in cuerpo or "<br>" in cuerpo:
            msg.content_subtype = "html"
            if "<br" not in cuerpo:
                msg.body = cuerpo.replace("\n", "<br>\n")
        for name, contenido, mimetype in adjuntos_list:
            msg.attach(name, contenido, mimetype or 'application/octet-stream')
        msg.send(fail_silently=False)
        registro.enviado_ok = True
        registro.save()
        informar_note = ""
        if documento_informar_id.isdigit():
            try:
                doc_inf = Document.objects.get(pk=int(documento_informar_id))
                prev = doc_inf.informado
                if prev == Document.INFORMADO_NO:
                    doc_inf.informado = Document.INFORMADO_SI
                elif prev == Document.INFORMADO_SI:
                    doc_inf.informado = Document.INFORMADO_OTRA
                if doc_inf.informado != prev:
                    doc_inf.save(update_fields=["informado", "updated_at"])
                    informar_note += f" Estado Informar (documento): «{doc_inf.get_informado_display()}»."
            except Document.DoesNotExist:
                pass
        if rdi_informar_csv_id.isdigit():
            try:
                rdi_inf = RDIRecord.objects.get(csv_id=int(rdi_informar_csv_id))
                prev_r = rdi_inf.informado
                if prev_r == RDI_INFORMADO_NO:
                    rdi_inf.informado = RDI_INFORMADO_SI
                elif prev_r == RDI_INFORMADO_SI:
                    rdi_inf.informado = RDI_INFORMADO_OTRA
                if rdi_inf.informado != prev_r:
                    rdi_inf.save(update_fields=["informado", "updated_at"])
                    informar_note += f" Estado Informar (RDI): «{rdi_inf.get_informado_display()}»."
            except (ValueError, RDIRecord.DoesNotExist):
                pass
        messages.success(
            request,
            f'Correo enviado correctamente a {", ".join(to_list)}.{informar_note}',
        )
        return redirect('enviar_correo')
    except Exception as e:
        registro.enviado_ok = False
        registro.error_msg = str(e)[:512]
        registro.save()
        messages.error(request, f'Error al enviar: {e}')
        return render(request, 'documents/enviar_correo.html', {
            'email_from': getattr(settings, 'EMAIL_HOST_USER', ''),
            'cc_grupos': cc_grupos,
            'destinatarios': destinatarios_raw,
            'copia': copia_raw,
            'asunto': asunto,
            'cuerpo': cuerpo,
            'usar_plantilla_transmittal': usar_plantilla,
            'documento_informar_id': documento_informar_id,
            'rdi_informar_csv_id': rdi_informar_csv_id,
        })


@login_required
@require_POST
def document_upload_attachments(request, pk):
    """Añadir varios archivos adjuntos a un documento. POST con input name='files' (múltiple)."""
    doc = get_object_or_404(Document, pk=pk)
    files = request.FILES.getlist('files')
    if not files:
        messages.warning(request, 'No se seleccionó ningún archivo.')
        return redirect('document_detail', pk=pk)
    created = 0
    for f in files:
        if not f.name:
            continue
        DocumentAttachment.objects.create(document=doc, file=f)
        created += 1
    if created:
        messages.success(request, f'Se añadieron {created} archivo(s) adjunto(s).')
    return redirect('document_detail', pk=pk)


# ---------- Carpetas ----------

@login_required
@require_GET
def folder_list(request):
    """Listado de carpetas (transmittals) con cantidad de documentos y archivos."""
    folders = (
        Folder.objects
        .annotate(
            document_count=Count('documents', distinct=True),
            folder_file_count=Count('folder_files', distinct=True),
            attachment_count=Count('documents__attachments', distinct=True),
        )
        .order_by('-date', '-created_at')[:500]
    )
    return render(request, 'documents/folder_list.html', {'folder_list': folders})


@login_required
@require_GET
def folder_detail(request, pk):
    """Detalle de una carpeta: documentos y archivos que contiene."""
    folder = get_object_or_404(
        Folder.objects.prefetch_related('documents', 'folder_files'),
        pk=pk
    )
    documents = folder.documents.select_related('project', 'company', 'process', 'doc_type').order_by('-date')
    folder_files = folder.folder_files.order_by('name')
    return render(request, 'documents/folder_detail.html', {
        'folder': folder,
        'documents': documents,
        'folder_files': folder_files,
    })


@login_required
@require_POST
def folder_upload_files(request, pk):
    """Añadir varios archivos a una carpeta. POST con input name='files' (múltiple)."""
    folder = get_object_or_404(Folder, pk=pk)
    files = request.FILES.getlist('files')
    if not files:
        messages.warning(request, 'No se seleccionó ningún archivo.')
        return redirect('folder_detail', pk=pk)
    created = 0
    for f in files:
        if not f.name:
            continue
        FolderFile.objects.create(folder=folder, name=f.name, file=f)
        created += 1
    if created:
        messages.success(request, f'Se añadieron {created} archivo(s) a la carpeta.')
    return redirect('folder_detail', pk=pk)


# ---------- Búsqueda unificada (nombre + contenido) ----------

@login_required
@require_GET
def search_unified_view(request):
    """Búsqueda por nombre/código y contenido. Responde HTML o JSON. Documentos y archivos de carpeta incluyen hits con snippets (igual lógica que listado de documentos)."""
    q = request.GET.get('q', '').strip()
    result = search_unified(q, limit=200)
    terms = _normalize_terms(q)
    use_multi = len(terms) > 1

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.GET.get('format') == 'json':
        # Documentos: mismo payload que document_list (con hits/snippets)
        doc_ids = [d.id for d in result['documents']]
        qs_docs = Document.objects.filter(pk__in=doc_ids).prefetch_related('attachments').select_related('project', 'company', 'process', 'doc_type', 'folder')
        docs = []
        for d in qs_docs:
            hits = []
            if q and d.content_extract:
                if use_multi and text_matches_all_terms_as_words(d.content_extract, terms):
                    snippets = extract_snippets_multi_term(d.content_extract, terms, context_words=10, max_snippets=5)
                elif not use_multi and text_matches_single_query(d.content_extract, q):
                    snippets = extract_snippets(d.content_extract, q, context_words=10, max_snippets=5)
                else:
                    snippets = []
                if snippets:
                    file_name = d.file.name if d.file else 'Documento principal'
                    hits.append({
                        'source': 'document',
                        'file_name': file_name.split('/')[-1] if file_name else 'Documento principal',
                        'file_url': d.file.url if d.file else None,
                        'snippets': snippets,
                    })
            for att in d.attachments.all():
                if q and att.extracted_text:
                    if use_multi and text_matches_all_terms_as_words(att.extracted_text, terms):
                        snippets = extract_snippets_multi_term(att.extracted_text, terms, context_words=10, max_snippets=5)
                    elif not use_multi and text_matches_single_query(att.extracted_text, q):
                        snippets = extract_snippets(att.extracted_text, q, context_words=10, max_snippets=5)
                    else:
                        snippets = []
                    if snippets:
                        name = att.file.name.split('/')[-1] if att.file.name else 'Adjunto'
                        hits.append({
                            'source': 'attachment',
                            'file_name': name,
                            'file_url': att.file.url if att.file else None,
                            'snippets': snippets,
                        })
            docs.append({
                'id': d.id,
                'code': d.code,
                'title': d.title or '',
                'status_display': d.get_status_display(),
                'date': d.date.isoformat() if d.date else '',
                'folder_id': d.folder_id,
                'folder_code': d.folder.code if d.folder else None,
                'file_url': d.file.url if d.file else None,
                'has_file': bool(d.file),
                'hits': hits,
            })
        folders_data = [
            {'id': f.id, 'code': f.code, 'title': f.title or '', 'date': f.date.isoformat() if f.date else ''}
            for f in result['folders']
        ]
        files_data = []
        for ff in result['folder_files']:
            hits = []
            if q and ff.extracted_text:
                if use_multi and text_matches_all_terms_as_words(ff.extracted_text, terms):
                    snippets = extract_snippets_multi_term(ff.extracted_text, terms, context_words=10, max_snippets=5)
                elif not use_multi and text_matches_single_query(ff.extracted_text, q):
                    snippets = extract_snippets(ff.extracted_text, q, context_words=10, max_snippets=5)
                else:
                    snippets = []
                if snippets:
                    hits.append({
                        'source': 'folder_file',
                        'file_name': ff.name,
                        'file_url': ff.file.url if ff.file else None,
                        'snippets': snippets,
                    })
            files_data.append({
                'id': ff.id,
                'name': ff.name,
                'file_url': ff.file.url if ff.file else None,
                'folder_id': ff.folder_id,
                'folder_code': ff.folder.code if ff.folder else None,
                'folder_title': (ff.folder.title or '') if ff.folder else '',
                'hits': hits,
            })
        return JsonResponse({
            'documents': docs,
            'folders': folders_data,
            'folder_files': files_data,
        })

    return render(request, 'documents/search_unified.html', {
        'query': q,
        'documents': result['documents'],
        'folders': result['folders'],
        'folder_files': result['folder_files'],
    })


# ---------- Estatus Cartas (documentos transmittal ODATA-BUF-CM-TTAL-*; info en extracto de adjuntos) ----------

# Solo considerar "carta" si el nombre del archivo contiene CAR o CARTAS como palabra (no "carga", etc.)
_CARTA_FILENAME_RE = re.compile(r"(^|[-_./\\])(CAR|CARTAS)([-_./\\]|$)", re.IGNORECASE)


def _is_carta_filename(name):
    """True si el nombre de archivo contiene CAR o CARTAS como palabra completa (ej. -CAR- o cartas.pdf)."""
    if not name:
        return False
    return bool(_CARTA_FILENAME_RE.search(name))


def _get_cartas_status_rows():
    """Construye la lista de filas del listado Estatus Cartas (reutilizable para HTML y export)."""
    # CAR/CARTAS como palabra en la ruta del adjunto (evita que entren "carga", "carro", etc.)
    # Sin backslash en el patrón para compatibilidad con todos los backends de BD.
    carta_regex = r"(^|[-_./])(CAR|CARTAS)([-_./]|$)"
    docs = (
        Document.objects
        .filter(code__icontains="TTAL", attachments__file__iregex=carta_regex)
        .distinct()
        .select_related("project", "company", "process", "doc_type", "folder")
        .prefetch_related("attachments")
        .order_by("-date", "-created_at")[:500]
    )
    respuesta_map = _build_respuesta_map()
    rows = []
    for d in docs:
        # Primer adjunto del documento (para extraer "enviado a" cuando el nombre no comienza con ODA)
        first_att = d.attachments.order_by("pk").first()
        first_extract = (first_att.extracted_text if first_att and first_att.extracted_text else None) or d.content_extract or ""
        for att in d.attachments.all():
            nombre_archivo = os.path.basename(att.file.name) if att.file else ""
            if not _is_carta_filename(nombre_archivo):
                continue
            # Si el nombre del archivo NO comienza con ODA: es otro formato → Requiere respuesta = NO, Enviado a = después de Señor hasta ODATA
            comienza_con_oda = (nombre_archivo or "").upper().startswith("ODA")
            if not comienza_con_oda:
                requiere = "NO"
                enviado_a = _parse_enviado_a_despues_senor(first_extract)
                if not enviado_a:
                    enviado_a = _parse_atencion(first_extract)
            else:
                requiere = _parse_requiere_respuesta(att.extracted_text)
                if not requiere:
                    requiere = _parse_requiere_respuesta(d.content_extract)
                enviado_a = _parse_atencion(att.extracted_text)
                if not enviado_a:
                    enviado_a = _parse_atencion(d.content_extract)
            detalle = _parse_asunto(att.extracted_text)
            if not detalle:
                detalle = _parse_asunto(d.content_extract)
            if not detalle:
                detalle = d.description or ""
            asunto = _parse_asunto(att.extracted_text)
            if not asunto:
                asunto = _parse_asunto(d.content_extract)
            fecha_envio = _parse_fecha_envio(att.extracted_text)
            if fecha_envio is None:
                fecha_envio = _parse_fecha_envio(d.content_extract)
            if fecha_envio is None:
                fecha_envio = d.date
            fecha_respuesta = None
            respuesta = ""
            if requiere == "SI":
                car_key = _normalize_car_code(nombre_archivo)
                if car_key and car_key in respuesta_map:
                    extract_respuesta = respuesta_map[car_key]
                    fecha_respuesta = _parse_fecha_envio(extract_respuesta)
                    respuesta = _parse_saluda_atentamente(extract_respuesta)
                    if not respuesta:
                        respuesta = _parse_atencion(extract_respuesta)
            rows.append({
                "document": d,
                "transmittal": d.folder.code if d.folder else "",
                "transmittal_title": (d.folder.title if d.folder else "") or "",
                "tipo": "CAR",
                "asunto": asunto or "",
                "descripcion": d.title or d.description or nombre_archivo or "",
                "fecha_envio": fecha_envio,
                "responsable": d.company.name or d.company.code,
                "documento_archivo": nombre_archivo or d.code,
                "estado": d.get_status_display(),
                "detalle": detalle,
                "enviado_a": enviado_a,
                "requiere_respuesta": requiere,
                "respuesta": respuesta,
                "fecha_respuesta": fecha_respuesta,
                "para": "",
                "file_url": att.file.url if att.file else None,
            })
    # Orden: fecha más reciente primero, luego por ID del registro (mayor a menor)
    def _sort_key(r):
        fe = r.get("fecha_envio") or date.min
        ord_val = fe.toordinal() if hasattr(fe, "toordinal") else 0
        doc = r.get("document")
        pk = doc.pk if doc is not None else 0
        return (-ord_val, -pk)
    rows.sort(key=_sort_key)
    return rows


def _cartas_status_excel_response(rows, list_name=None):
    """Genera HttpResponse con el listado en Excel.
    Texto con ajuste de línea (wrap) en las celdas para no truncar. list_name: nombre para archivo/hoja."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    fecha_str = timezone.now().strftime("%d-%m-%Y")
    if list_name:
        safe_name = list_name.replace(" ", "-").replace("/", "-")[:40]
        filename = f"{safe_name}-{fecha_str}.xlsx"
        sheet_title = safe_name[:31]
    else:
        filename = f"Estatus-Cartas-al-{fecha_str}.xlsx"
        sheet_title = "Estatus Cartas"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    headers = [
        "Transmittal", "Asunto", "Fecha envío (emisor)", "Responsable", "Documento - Archivo",
        "Estado", "Detalle", "Enviado a", "Requiere respuesta?", "Respuesta", "Fecha respuesta",
    ]
    ws.append(headers)
    for h in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        fe = row.get("fecha_envio")
        fe_str = fe.strftime("%d/%m/%Y") if hasattr(fe, "strftime") else str(fe or "")
        fr = row.get("fecha_respuesta")
        fr_str = fr.strftime("%d/%m/%Y") if fr and hasattr(fr, "strftime") else ""
        trans = (row.get("transmittal") or "")[:25]
        ws.append([
            trans,
            row.get("asunto") or "",
            fe_str,
            row.get("responsable") or "",
            row.get("documento_archivo") or "",
            row.get("estado") or "",
            row.get("detalle") or "",
            row.get("enviado_a") or "",
            row.get("requiere_respuesta") or "",
            row.get("respuesta") or "",
            fr_str,
        ])
    # Ajuste de texto en todas las celdas de datos (sin truncar; añade líneas)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    for r in range(1, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).alignment = wrap_align
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _cartas_status_pdf_response(rows, inline=False, list_name=None):
    """Genera HttpResponse con el listado en PDF.
    Celdas con Paragraph para que el texto se ajuste en varias líneas sin truncar."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.units import cm
    from io import BytesIO
    fecha_str = timezone.now().strftime("%d-%m-%Y")
    if list_name:
        safe_name = list_name.replace(" ", "-").replace("/", "-")[:40]
        filename = f"{safe_name}-{fecha_str}.pdf"
        pagesize = A4
    else:
        filename = f"Estatus-Cartas-al-{fecha_str}.pdf"
        # Horizontal explícito (ancho, alto) = (A4[1], A4[0]) para evitar diferencias por versión de ReportLab
        pagesize = (A4[1], A4[0])
    buf = BytesIO()
    left_right_margin = 0.5 * cm
    doc = SimpleDocTemplate(buf, pagesize=pagesize, rightMargin=left_right_margin, leftMargin=left_right_margin, topMargin=1.2*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    para_cell = ParagraphStyle("Cell", parent=styles["Normal"], fontName="Helvetica", fontSize=6, leading=7, leftIndent=0, rightIndent=0, spaceBefore=0, spaceAfter=0)
    para_header = ParagraphStyle("Header", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7, leading=8)

    def safe_para(text, style=para_cell):
        if text is None:
            text = ""
        s = str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")
        return Paragraph(s or " ", style)

    headers = [
        "Transmittal", "Asunto", "F.Envio", "Responsable", "Documento - Archivo",
        "Estado", "Detalle", "Enviado a", "Req.Resp", "Respuesta", "F.Resp",
    ]
    data = [[safe_para(h, para_header) for h in headers]]
    for row in rows:
        fe = row.get("fecha_envio")
        fe_str = fe.strftime("%d/%m/%Y") if hasattr(fe, "strftime") else str(fe or "")
        fr = row.get("fecha_respuesta")
        fr_str = fr.strftime("%d/%m/%Y") if fr and hasattr(fr, "strftime") else ""
        data.append([
            safe_para((row.get("transmittal") or "")[:25]),
            safe_para(row.get("asunto") or ""),
            safe_para(fe_str),
            safe_para(row.get("responsable") or ""),
            safe_para(row.get("documento_archivo") or ""),
            safe_para(row.get("estado") or ""),
            safe_para(row.get("detalle") or ""),
            safe_para(row.get("enviado_a") or ""),
            safe_para(row.get("requiere_respuesta") or ""),
            safe_para(row.get("respuesta") or ""),
            safe_para(fr_str),
        ])
    col_widths = [2.8*cm, 5*cm, 1.6*cm, 2.8*cm, 3.2*cm, 1.5*cm, 3.2*cm, 2.5*cm, 1.2*cm, 2.8*cm, 1.6*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E7E6E6")]),
    ]))
    doc.build([t])
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    if inline:
        response["Content-Disposition"] = f'inline; filename="{filename}"'
    else:
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_GET
def cartas_status(request):
    """
    Lista solo cartas: documentos transmittal (TTAL) que tengan al menos un adjunto tipo CAR.
    Si ?format=excel o ?format=pdf, devuelve el archivo para descarga (Estatus-Cartas-al-DD-MM-YYYY).
    """
    rows = _get_cartas_status_rows()

    if request.GET.get("format") == "excel":
        return _cartas_status_excel_response(rows)
    if request.GET.get("format") == "pdf":
        open_inline = request.GET.get("open") == "1"
        return _cartas_status_pdf_response(rows, inline=open_inline)

    n_si = sum(1 for r in rows if r.get("requiere_respuesta") == "SI")
    n_no = sum(1 for r in rows if r.get("requiere_respuesta") == "NO")
    n_vacio = sum(1 for r in rows if not r.get("requiere_respuesta"))
    print(f"[cartas_status] RESUMEN: filas(adjuntos)={len(rows)} | SI={n_si} | NO={n_no} | vacío={n_vacio}")

    return render(request, "documents/cartas_status.html", {"rows": rows})


def _extract_referencia_from_text(text):
    """
    Extrae el texto que sigue a 'Referencia:' en el extracto.
    Incluye todo el párrafo (varias líneas) hasta un doble salto de línea,
    para capturar por ejemplo "Detalle documentos Adjuntos:" y las líneas siguientes
    (ej. "- Workshop Compilado comentarios"). Máximo 2000 caracteres.
    """
    if not text or not isinstance(text, str):
        return ""
    t = text
    ref_lower = "referencia"
    idx = t.lower().find(ref_lower)
    if idx < 0:
        return ""
    start = idx + len(ref_lower)
    while start < len(t) and (t[start] == ":" or t[start].isspace()):
        start += 1
    rest = t[start:].strip()
    # Tomar todo el párrafo hasta doble salto (no cortar en la primera línea)
    for sep in ("\n\n", "\r\n\r\n"):
        if sep in rest:
            rest = rest.split(sep)[0].strip()
            break
    return rest[:2000] if len(rest) > 2000 else rest


def _extract_unidad_emisora_from_text(text):
    """
    Extrae el texto que sigue a 'Unidad emisora' en el extracto (ej. "BUFFER Pablo Andrés Soto Calderón").
    Toma una línea o hasta doble salto. Máximo 500 caracteres.
    """
    if not text or not isinstance(text, str):
        return ""
    t = text
    label = "unidad emisora"
    idx = t.lower().find(label)
    if idx < 0:
        return ""
    start = idx + len(label)
    while start < len(t) and (t[start] == ":" or t[start].isspace()):
        start += 1
    rest = t[start:].strip()
    for sep in ("\n\n", "\r\n\r\n", "\n", "\r\n"):
        if sep in rest:
            rest = rest.split(sep)[0].strip()
            break
    return rest[:500] if len(rest) > 500 else rest


def _extract_status_from_text(text):
    """
    Extrae el valor de 'Status:' o 'Estatus:' desde el extracto del documento principal.
    Se usa para la columna Estado en logs TRN / ODATA (Propamat↔Odata).
    Devuelve texto corto normalizado (ej. 'Aprobado', 'Rechazado', u otro valor).
    """
    if not text or not isinstance(text, str):
        return ""
    t = text
    lower = t.lower()
    labels = ["status", "estatus"]
    idx = -1
    label_len = 0
    for lab in labels:
        i = lower.find(lab)
        if i >= 0 and (idx < 0 or i < idx):
            idx = i
            label_len = len(lab)
    if idx < 0:
        return ""
    start = idx + label_len
    while start < len(t) and (t[start] == ":" or t[start].isspace()):
        start += 1
    rest = t[start:].strip()
    for sep in ("\n\n", "\r\n\r\n", "\n", "\r\n"):
        if sep in rest:
            rest = rest.split(sep)[0].strip()
            break
    if not rest:
        return ""
    val = re.sub(r"\s+", " ", rest).strip()
    vlow = val.lower()
    if "aprob" in vlow or "approved" in vlow:
        return "Aprobado"
    if "rechaz" in vlow or "rejected" in vlow:
        return "Rechazado"
    if val.isupper() or val.islower():
        return val.capitalize()
    return val[:80]


def _extract_respuesta_from_extracted_text(text, max_chars=1200):
    """
    Extrae la sección agregada por la extracción de XLS/XLSX:
      "Respuesta:" + valores de esa columna.
    Devuelve texto (limitado) o '' si no existe.
    """
    if not text or not isinstance(text, str):
        return ""
    lower = text.lower()
    marker = "respuesta:"
    idx = lower.rfind(marker)
    if idx < 0:
        return ""
    rest = text[idx + len(marker):].strip()
    if not rest:
        return ""

    # Si vienen pares "Pregunta:" / "Respuesta:" múltiples, cortar en la próxima "Pregunta:"
    rest_lower = rest.lower()
    next_q = rest_lower.find("pregunta:")
    if next_q >= 0:
        rest = rest[:next_q].strip()

    rest = re.sub(r"\n{3,}", "\n\n", rest).strip()
    if len(rest) > max_chars:
        rest = rest[:max_chars].rsplit("\n", 1)[0].rstrip() + " …"
    return rest


def _extract_respuesta_from_excel_pairs(text, terms=None, query=None, max_chars=1200):
    """
    Para archivos Excel extraídos como pares:
      "Pregunta: ...\\nRespuesta: ..."

    Devuelve la "Respuesta" cuya "Pregunta" matchee con:
    - si query (frase exacta) viene: substring match en la pregunta
    - si no: todos los términos aparecen en la pregunta
    """
    if not text or not isinstance(text, str):
        return ""
    terms = terms or []
    query = (query or "").strip()
    pattern = re.compile(
        r"Pregunta:\s*(?P<preg>.*?)\s*Respuesta:\s*(?P<resp>.*?)(?=Pregunta:|$)",
        flags=re.IGNORECASE | re.DOTALL,
    )
    q_lower = query.lower()
    for m in pattern.finditer(text):
        preg = (m.group("preg") or "").strip()
        resp = (m.group("resp") or "").strip()
        if not resp:
            continue
        preg_l = preg.lower()
        if q_lower:
            if q_lower in preg_l:
                result = resp
                break
        else:
            if terms and all((t or "").lower() in preg_l for t in terms):
                result = resp
                break
    else:
        return ""

    result = result.replace("\r\n", "\n").replace("\r", "\n")
    result = re.sub(r"[ \t]+", " ", result)
    result = re.sub(r"\n{3,}", "\n\n", result).strip()
    if not result:
        return ""
    if len(result) > max_chars:
        result = result[:max_chars].rsplit(" ", 1)[0].rstrip() + " …"
    return result


def _extract_qa_from_delimiters(text, query=None, terms=None):
    """
    Extrae pares QA desde texto indexado con formato fijo:
      ¿Pregunta?@Respuesta@

    - pregunta: incluye delimitadores ¿...?
    - respuesta: contenido ENTRE @...@ (sin @)
    """
    if not text or not isinstance(text, str):
        return {"pregunta": "", "respuesta": "", "especialidad": "", "gc": ""}

    pattern_meta = re.compile(
        r"GC:\s*(?P<gc>[^\r\n]*)[\r\n]+\s*Especialidad:\s*(?P<esp>[^\r\n]*)"
        r"(?:[\r\n]+\s*N:\s*(?P<num>[^\r\n]*))?"
        r"[\r\n]+(?P<preg>¿.*?\?)[ \t\r\n]*@(?P<resp>.*?)@",
        flags=re.DOTALL | re.IGNORECASE,
    )
    pattern_plain = re.compile(
        r"(?P<preg>¿.*?\?)[ \t\r\n]*@(?P<resp>.*?)@",
        flags=re.DOTALL | re.IGNORECASE,
    )

    matches = list(pattern_meta.finditer(text))
    if not matches:
        matches = list(pattern_plain.finditer(text))
    if not matches:
        return {"pregunta": "", "respuesta": "", "especialidad": "", "gc": ""}

    query = (query or "").strip()
    terms = terms or []

    def norm(s: str) -> str:
        return (s or "").strip().replace("\r\n", "\n").replace("\r", "\n")

    # 1) Prioridad: si viene query, match en pregunta o respuesta
    if query:
        ql = query.lower()
        for m in matches:
            preg = norm(m.group("preg"))
            resp = norm(m.group("resp"))
            gc = norm(m.groupdict().get("gc", ""))
            especialidad = norm(m.groupdict().get("esp", ""))
            numero = norm(m.groupdict().get("num", ""))
            pl = preg.lower()
            rl = resp.lower()
            if ql in pl or ql in rl:
                return {
                    "pregunta": preg,
                    "respuesta": resp,
                    "especialidad": especialidad,
                    "gc": gc,
                    "numero": numero,
                }

    # 2) Multi-término: cada término debe aparecer en pregunta o respuesta
    if terms:
        terms_l = [(t or "").strip().lower() for t in terms if (t or "").strip()]
        if terms_l:
            for m in matches:
                preg = norm(m.group("preg"))
                resp = norm(m.group("resp"))
                gc = norm(m.groupdict().get("gc", ""))
                especialidad = norm(m.groupdict().get("esp", ""))
                numero = norm(m.groupdict().get("num", ""))
                pl = preg.lower()
                rl = resp.lower()
                if all((t in pl) or (t in rl) for t in terms_l):
                    return {
                        "pregunta": preg,
                        "respuesta": resp,
                        "especialidad": especialidad,
                        "gc": gc,
                        "numero": numero,
                    }

    # 3) fallback: primer match
    m = matches[0]
    return {
        "pregunta": norm(m.group("preg")),
        "respuesta": norm(m.group("resp")),
        "especialidad": norm(m.groupdict().get("esp", "")),
        "gc": norm(m.groupdict().get("gc", "")),
        "numero": norm(m.groupdict().get("num", "")),
    }


def _extract_all_qa_from_delimiters(text, query=None, terms=None, max_pairs=8):
    """
    Igual que _extract_qa_from_delimiters pero devuelve una lista:
      ¿Pregunta?@Respuesta@

    - si query viene: retorna pares donde query aparece en pregunta o respuesta
    - si terms viene: retorna pares donde CADA término aparece en pregunta o respuesta
    """
    if not text or not isinstance(text, str):
        return []

    pattern_meta = re.compile(
        r"GC:\s*(?P<gc>[^\r\n]*)[\r\n]+\s*Especialidad:\s*(?P<esp>[^\r\n]*)"
        r"(?:[\r\n]+\s*N:\s*(?P<num>[^\r\n]*))?"
        r"[\r\n]+(?P<preg>¿.*?\?)[ \t\r\n]*@(?P<resp>.*?)@",
        flags=re.DOTALL | re.IGNORECASE,
    )
    pattern_plain = re.compile(
        r"(?P<preg>¿.*?\?)[ \t\r\n]*@(?P<resp>.*?)@",
        flags=re.DOTALL | re.IGNORECASE,
    )

    matches = list(pattern_meta.finditer(text))
    if not matches:
        matches = list(pattern_plain.finditer(text))
    if not matches:
        return []

    query = (query or "").strip().lower()
    terms_l = [(t or "").strip().lower() for t in (terms or []) if (t or "").strip()]
    # Evita que términos repetidos afecten el conteo de coincidencias.
    terms_l = list(dict.fromkeys(terms_l))

    def norm(s: str) -> str:
        return (s or "").strip().replace("\r\n", "\n").replace("\r", "\n")

    out = []
    for m in matches:
        preg = norm(m.group("preg"))
        resp = norm(m.group("resp"))
        gc = norm(m.groupdict().get("gc", ""))
        especialidad = norm(m.groupdict().get("esp", ""))
        numero = norm(m.groupdict().get("num", ""))
        if not preg or not resp:
            continue

        pl = preg.lower()
        rl = resp.lower()

        ok = False
        if query:
            ok = (query in pl) or (query in rl)
        elif terms_l:
            # Búsqueda extendida: aceptar SOLO si TODOS los términos aparecen
            # en la pregunta o en la respuesta (AND).
            matched = sum(1 for t in terms_l if (t in pl) or (t in rl))
            ok = matched == len(terms_l)
            if ok:
                out.append(
                    (
                        matched,
                        m.start(),
                        {
                            "pregunta": preg,
                            "respuesta": resp,
                            "especialidad": especialidad,
                            "gc": gc,
                            "numero": numero,
                        },
                    )
                )
            if ok:
                # continuamos para no duplicar el append en el bloque final
                continue
        else:
            ok = True

        if ok:
            # Sin scoring cuando no usamos términos
            out.append(
                (
                    1,
                    m.start(),
                    {
                        "pregunta": preg,
                        "respuesta": resp,
                        "especialidad": especialidad,
                        "gc": gc,
                        "numero": numero,
                    },
                )
            )
            if len(out) >= max_pairs:
                break
    # Si usamos términos, out viene con tuplas (matched, start, qa)
    # Ordenar por mayor cantidad de términos, y por aparición (start) en caso de empate.
    if terms_l:
        out_sorted = sorted(out, key=lambda x: (-x[0], x[1]))
        return [item[2] for item in out_sorted[:max_pairs]]
    return [item[2] for item in out]


def _extract_pregunta_y_respuesta_from_excel_pairs(text, terms=None, query=None, max_chars=1200):
    """
    Para archivos Excel extraídos como pares:
      "Pregunta: ...\\nRespuesta: ..."

    Devuelve dict {pregunta, respuesta, especialidad} para el par cuya Pregunta matchee:
    - si query viene: substring match en la pregunta
    - si no: todos los términos aparecen en la pregunta
    """
    if not text or not isinstance(text, str):
        return {"pregunta": "", "respuesta": "", "especialidad": ""}
    terms = terms or []
    query = (query or "").strip()

    pattern = re.compile(
        r"Pregunta:\s*(?P<preg>.*?)\s*Respuesta:\s*(?P<resp>.*?)(?=Pregunta:|$)",
        flags=re.IGNORECASE | re.DOTALL,
    )

    q_lower = query.lower()
    for m in pattern.finditer(text):
        preg = (m.group("preg") or "").strip()
        resp = (m.group("resp") or "").strip()
        if not resp:
            continue

        preg_l = preg.lower()
        resp_l = resp.lower()
        if q_lower:
            if q_lower in preg_l or q_lower in resp_l:
                break
        else:
            if terms and all((t or "").lower() in preg_l for t in terms):
                break
            if terms and all((t or "").lower() in resp_l for t in terms):
                break
    else:
        return {"pregunta": "", "respuesta": "", "especialidad": ""}

    # Especialidad: tomada desde la misma "fila/registro" que contiene Pregunta/Respuesta.
    # En el texto indexado del Excel suele venir como:
    #   Especialidad: XYZ
    #   Pregunta: ...
    #
    # Por eso buscamos "especialidad:" lo más cerca posible antes del match.
    especialidad = ""
    try:
        start_pos = m.start()
        window_start = max(0, start_pos - 500)  # suficiente para cubrir el bloque "Especialidad: ...\n"
        before_window = text[window_start:start_pos]
        # Importante: no usar `\s*` porque puede consumir el salto de línea cuando el valor está vacío,
        # y terminar capturando el siguiente label (ej. "Pregunta: ...") como "especialidad".
        esp_re = re.search(
            r"especialidad:[ \t]*(?P<esp>[^\r\n]*)",
            before_window,
            flags=re.IGNORECASE,
        )
        if esp_re:
            especialidad = (esp_re.group("esp") or "").strip()
    except Exception:
        especialidad = ""

    preg_norm = re.sub(r"\s+", " ", preg).strip()
    resp_norm = resp.replace("\r\n", "\n").replace("\r", "\n")
    resp_norm = re.sub(r"[ \t]+", " ", resp_norm)
    resp_norm = re.sub(r"\n{3,}", "\n\n", resp_norm).strip()
    especialidad_norm = re.sub(r"\s+", " ", especialidad).strip()
    if len(especialidad_norm) > 120:
        especialidad_norm = especialidad_norm[:120].rsplit(" ", 1)[0].strip()
    if len(resp_norm) > max_chars:
        resp_norm = resp_norm[:max_chars].rsplit(" ", 1)[0].rstrip() + " …"

    return {"pregunta": preg_norm, "respuesta": resp_norm, "especialidad": especialidad_norm}


def _get_main_extract_for_folder(folder):
    """
    Devuelve el extracto donde se buscan Referencia, Responsable y Documento-Archivo en los logs.
    Ese texto está en el DOCUMENTO PRINCIPAL de la carpeta (content_extract), no en los PDF adjuntos.
    Orden: primer documento con content_extract (documento maestro, antes de los adjuntos),
    luego primer archivo de carpeta (FolderFile). No se usan extractos de adjuntos.
    """
    for doc in folder.documents.all():
        if doc.content_extract and doc.content_extract.strip():
            return doc.content_extract
    for ff in folder.folder_files.all():
        if ff.extracted_text and ff.extracted_text.strip():
            return ff.extracted_text
    return ""


def _get_main_document_for_folder(folder):
    """
    Devuelve el documento principal asociado a la carpeta para usar valores de BD (ej. status).
    Heurística: primer documento con content_extract; si no, el primer documento de la carpeta.
    """
    docs = list(folder.documents.all())
    for doc in docs:
        if doc.content_extract and doc.content_extract.strip():
            return doc
    return docs[0] if docs else None


def _extract_detalle_documentos_adjuntos_from_text(text):
    """
    Para listado TRN (Propamat→Odata): extrae el fragmento entre 'Detalle documentos Adjuntos:'
    y 'Destinatario:'. Se usa la ÚLTIMA aparición de esa etiqueta antes de 'Destinatario'
    para evitar tomar una tabla anterior (p. ej. que suelte "Rev.00"). Solo se incluyen
    líneas que parecen nombres de documentos; se deja de incluir al detectar cabeceras
    de tabla o secciones (Unidad revisora/emisora).
    """
    if not text or not isinstance(text, str):
        return ""
    t = text
    t_lower = t.lower()
    pos_dest = t_lower.find("destinatario")
    if pos_dest < 0:
        pos_dest = len(t)
    # Última aparición de "detalle documentos adjuntos" ANTES de "Destinatario"
    label = "detalle documentos adjuntos"
    last_start = -1
    pos = 0
    while pos < pos_dest:
        idx = t_lower.find(label, pos)
        if idx < 0 or idx >= pos_dest:
            break
        start = idx + len(label)
        while start < len(t) and (t[start] in ":\n\r\t" or t[start].isspace()):
            start += 1
        last_start = start
        pos = idx + 1
    if last_start < 0:
        return ""
    block = t[last_start:pos_dest].strip()

    # Dejar de añadir líneas al llegar a cabeceras de tabla o siguiente sección
    stop_markers = (
        "ítem",
        "n° documento",
        "nº documento",
        "rev.",
        "título / descripción documento",
        "titulo / descripcion documento",
        "comentarios:",
        "unidad revisora",
        "unidad emisora",
    )

    def should_stop(line):
        low = line.strip().lower()
        if not low:
            return True
        for m in stop_markers:
            if low == m:
                return True
            if (m.endswith(":") or m.endswith(" ")) and low.startswith(m):
                return True
            if not m.endswith(":") and (low.startswith(m + " ") or low.startswith(m + ":")):
                return True
        if low.isdigit():
            return True
        if len(low) <= 4 and low.startswith("-") and low[1:].replace(" ", "").isdigit():
            return True
        # No incluir líneas que son solo revisión (Rev.00, Rev.01, N/A)
        if re.match(r"^rev\.?\s*\d*$", low) or low == "n/a":
            return True
        return False

    lines = []
    for ln in block.splitlines():
        ln = ln.strip()
        if not ln:
            continue
        if should_stop(ln):
            break
        lines.append(ln)
    if not lines:
        return ""
    formatted = "\n".join("• " + ln for ln in lines)
    return formatted[:1500] if len(formatted) > 1500 else formatted


def _extract_after_referencia_from_text(text):
    """
    Extrae y lista lo que está después de "Referencia:" en el extracto.
    Se corta al encontrar "Detalle documentos Adjuntos:" o "Destinatario:".
    Ejemplo: después de "Referencia:" puede venir texto y luego líneas como
    "Flujo Financiero_Histograma DC ODATA R0", "Listado de equipos criticos...", etc.
    """
    if not text or not isinstance(text, str):
        return ""
    t_lower = text.lower()
    idx = t_lower.find("referencia:")
    if idx < 0:
        return ""
    start = idx + len("referencia:")
    while start < len(text) and (text[start] in ":\n\r\t" or text[start].isspace()):
        start += 1
    rest = text[start:]
    # Cortar en la siguiente sección
    for stop in ("detalle documentos adjuntos", "destinatario"):
        pos = rest.lower().find(stop)
        if pos >= 0:
            rest = rest[:pos]
            break
    rest = rest.strip()
    lines = [ln.strip() for ln in rest.splitlines() if ln.strip()]
    if not lines:
        return ""
    return "\n".join("• " + ln for ln in lines)[:1500]


def _extract_detalle_documentos_adjuntos_odata_from_text(text):
    """
    Extrae para logs Odata→Propamat el bloque entre "(4)" y "Comentarios:" del extracto
    del documento maestro (content_extract). Ese bloque son las filas de la tabla, p. ej.:
      Ítem Rev. (1) (2) (3) (4)
      1 N/A OT 11 VI 1
      2 N/A OT 11 VI 1
      3 N/A OT 11 VI 1
      4 N/A OT 11 VI 1
      Comentarios:
    Se devuelve la lista de líneas después de "(4)" y antes de "Comentarios:" (con viñeta).
    """
    if not text or not isinstance(text, str):
        return ""
    idx_coment = text.lower().find("comentarios:")
    if idx_coment < 0:
        return ""
    before_comentarios = text[:idx_coment]
    # Última aparición de "(4)" (cabecera suele ser "Ítem Rev. (1) (2) (3) (4)")
    marker = "(4)"
    last_pos = -1
    pos = 0
    while True:
        found = before_comentarios.find(marker, pos)
        if found < 0:
            break
        last_pos = found
        pos = found + 1
    if last_pos < 0:
        return ""
    # Justo después de "(4)" hasta Comentarios: = filas de la tabla (1 N/A OT 11 VI 1, etc.)
    start = last_pos + len(marker)
    rest = before_comentarios[start:].strip()
    # No incluir líneas que ya sean la leyenda (por si el orden del PDF variara)
    stop_markers = (
        "(1) tipo", "(2) identificación", "(3) tipo de copia", "(4) número de copias",
        "unidad revisora", "unidad emisora", "document type", "submission identification",
        "copy type", "number of copies",
    )
    lines_out = []
    for ln in rest.splitlines():
        ln = ln.strip()
        if not ln:
            continue
        low = ln.lower()
        if any(low.startswith(m) for m in stop_markers):
            break
        lines_out.append(ln)
    if not lines_out:
        return ""
    formatted = "\n".join("• " + ln for ln in lines_out)
    return formatted[:1500] if len(formatted) > 1500 else formatted


def _extract_titulo_descripcion_documento_from_text(text):
    """
    Extrae el texto que sigue al encabezado 'Título / Descripción Documento' en el extracto.
    Ese contenido suele ser la descripción del documento en la tabla del PDF.
    Si la tabla está como imagen en el PDF, el extracto no lo contendrá (limitación de la capa de texto).
    """
    if not text or not isinstance(text, str):
        return ""
    t = text
    # Variantes del encabezado (con/sin tilde, con espacio o barra)
    for label in ("título / descripción documento", "titulo / descripcion documento",
                  "título/descripción documento", "titulo/descripcion documento",
                  "título - descripción documento", "descripción documento"):
        idx = t.lower().find(label)
        if idx >= 0:
            start = idx + len(label)
            while start < len(t) and (t[start] in ":-\n\r\t" or t[start].isspace()):
                start += 1
            rest = t[start:].strip()
            # Tomar hasta doble salto o hasta otra cabecera típica
            for sep in ("\n\n", "\r\n\r\n"):
                if sep in rest:
                    rest = rest.split(sep)[0].strip()
                    break
            # Una sola línea si hay muchas (primera línea suele ser la descripción)
            if "\n" in rest:
                rest = rest.split("\n")[0].strip()
            return rest[:500] if len(rest) > 500 else rest
    return ""


def _get_logs_folder_rows(code_filter, order_by="-date", then_by="-code"):
    """
    Construye filas para los listados de logs por tipo de carpeta.
    Misma lógica para ambos: filtrar carpetas por código, tomar el primer extracto del documento
    asociado a la carpeta y extraer de ahí Referencia, Responsable y Documento-Archivo.

    - Odata a Propamat (funciona 100%): code_filter 'TRN' → code__icontains='TRN'.
    - Propamat a Odata: code_filter 'Odata' → carpetas con formato ODATA-ST01-F5-TTAL-PPT-?????
      (code__icontains='ODATA-ST01-F5-TTAL-PPT'); el documento asociado tiene ese mismo formato
      y ahí está el primer extracto con la lista de archivos (entre (4) y Comentarios:).
    """
    if code_filter and code_filter.lower() == "odata":
        folders = (
            Folder.objects.filter(code__icontains="ODATA-ST01-F5-TTAL-PPT")
            .prefetch_related("folder_files", "documents__attachments")
            .order_by(order_by, then_by)[:500]
        )
    else:
        folders = (
            Folder.objects.filter(code__icontains=code_filter)
            .prefetch_related("folder_files", "documents__attachments")
            .order_by(order_by, then_by)[:500]
        )
    def _clean_doc_arch(text):
        """
        Limpia el texto de Documento-Archivo en los logs,
        eliminando líneas inútiles como:
        - 'Rev'
        - '.'
        - 'www.propamat.cl'
        - 'Rev . www.propamat.cl'
        que a veces se cuelan desde el extracto del PDF.
        """
        if not text:
            return ""
        out_lines = []
        for raw in str(text).splitlines():
            ln = raw.strip()
            if not ln:
                continue
            # Quitar viñetas y puntos sueltos para comparar
            norm = ln.lstrip("•-*").strip().lower()
            norm = norm.replace(" ", " ")  # espacio no separable
            norm = re.sub(r"\s+", " ", norm)
            if norm in ("rev", ".", "www.propamat.cl", "rev . www.propamat.cl"):
                continue
            out_lines.append(ln)
        return "\n".join(out_lines).strip()

    rows = []
    is_odata_list = "odata" in (code_filter or "").lower()
    for f in folders:
        main_extract = _get_main_extract_for_folder(f)
        main_doc = _get_main_document_for_folder(f)
        if is_odata_list:
            # Misma lógica que el envío de correo transmittal: referencia = texto del asunto (antes de nota/ítem);
            # documento_archivo = lista de documentos como en el cuerpo (antes de Saludos cordiales, sin Emitido).
            parsed = _parse_transmittal_extract(main_extract, f.code)
            ref_raw = parsed.get("referencia") or ""
            referencia = _referencia_solo_para_asunto(_limpiar_referencia_para_asunto(ref_raw))
            docs = parsed.get("documentos") or []
            # Excluir cualquier ítem que sea solo "Emitido" (en cualquier posición)
            docs = [d for d in docs if re.sub(r"[,.\s]+$", "", (d or "").lower()).strip() != "emitido"]
            doc_arch = "\n".join("• " + d for d in docs) if docs else ""
            if not doc_arch:
                doc_arch = _extract_after_referencia_from_text(main_extract)
                if not doc_arch:
                    doc_arch = _extract_detalle_documentos_adjuntos_odata_from_text(main_extract)
                # Quitar líneas con "emitido" y cortar en "Saludos cordiales"
                if doc_arch:
                    lineas = []
                    for ln in doc_arch.replace("\r", "\n").splitlines():
                        ln = ln.strip()
                        if not ln:
                            continue
                        if "saludos cordiales" in ln.lower():
                            break
                        if "emitido" in ln.lower():
                            continue
                        lineas.append(ln)
                    doc_arch = "\n".join(lineas) if lineas else ""
        else:
            doc_arch = _extract_detalle_documentos_adjuntos_from_text(main_extract)
        if not doc_arch:
            doc_arch = _extract_titulo_descripcion_documento_from_text(main_extract)
            # No mostrar fallback si es solo revisión (Rev.00) o celda de tabla
            if doc_arch and not is_odata_list and re.match(r"^Rev\.?\s*\d*$", doc_arch.strip(), re.IGNORECASE):
                doc_arch = ""
        doc_arch = _clean_doc_arch(doc_arch)
        if is_odata_list and not referencia:
            referencia = _referencia_solo_para_asunto(_limpiar_referencia_para_asunto(_extract_referencia_from_text(main_extract)))
        # Estado desde base de datos (Document.status), no desde el extracto del PDF/DOCX
        if main_doc and getattr(main_doc, "status", None):
            try:
                estado = main_doc.get_status_display()
            except Exception:
                estado = str(main_doc.status)
        else:
            estado = ""
        # Etiqueta del campo Document.informado (p. ej. plantilla Informar); no sustituye "estado" (status del doc).
        if main_doc:
            try:
                informado_display = main_doc.get_informado_display()
            except Exception:
                informado_display = getattr(main_doc, "informado", "") or "—"
        else:
            informado_display = "—"
        # Transmittal para listado/Excel/PDF: solo el código (ej. ODATA-ST01-F5-TTAL-PPT-00068), sin " — PROPAMAT-A-ODATA-XX"
        code = (f.code or "").strip()
        transmittal_display = code.split(" — ")[0].strip() if code else ""
        rows.append({
            "folder": f,
            "document": main_doc,
            "transmittal": code,
            "transmittal_title": f.title or "",
            "transmittal_display": transmittal_display,
            "descripcion": f.description or "",
            "referencia": referencia if is_odata_list else _extract_referencia_from_text(main_extract),
            "fecha_envio": f.date,
            "responsable": _extract_unidad_emisora_from_text(main_extract),
            "documento_archivo": doc_arch,
            "estado": estado,
            "informado_display": informado_display,
            "detalle": "",
            "enviado_a": "",
            "requiere_respuesta": "",
            "respuesta": "",
            "fecha_respuesta": None,
            "para": "",
            "file_url": None,
        })
    return rows


def _format_log_row_for_export(row, request=None):
    """
    Formatea una fila de log con la misma lógica que los templates:
    Transmittal = solo el código (parte antes de " — "), sin " — PROPAMAT-A-ODATA-XX". Igual en Excel y PDF.
    """
    trans = row.get("transmittal_display")
    if trans is None:
        trans = (row.get("transmittal") or "").strip()
        trans = trans.split(" — ")[0].strip() if trans else ""
    fe = row.get("fecha_envio")
    fecha_str = fe.strftime("%d/%m/%Y") if fe and hasattr(fe, "strftime") else "—"
    link = "Ver"
    if request:
        if row.get("folder"):
            link = request.build_absolute_uri(reverse("folder_detail", args=[row["folder"].pk]))
        elif row.get("document"):
            link = request.build_absolute_uri(reverse("document_detail", args=[row["document"].pk]))
    return {
        "transmittal": trans,
        "referencia": (row.get("referencia") or "—").strip(),
        "fecha_envio": fecha_str,
        "responsable": (row.get("responsable") or "—").strip(),
        "documento_archivo": (row.get("documento_archivo") or "—").strip(),
        "estado": (row.get("estado") or "—").strip(),
        "link": link,
    }


def _logs_excel_response(rows, list_name, request=None):
    """
    Excel para listados de logs (Propamat↔Odata) con las mismas columnas que el template.
    Celdas con wrap, márgenes pequeños, filtros automáticos en la fila de encabezado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.worksheet.page import PageMargins
    from io import BytesIO
    fecha_str = timezone.now().strftime("%d-%m-%Y")
    safe_name = (list_name or "Logs").replace(" ", "-").replace("/", "-")[:40]
    filename = f"{safe_name}-{fecha_str}.xlsx"
    sheet_title = safe_name[:31]
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    headers = [
        "Transmittal", "Referencia", "Fecha envío (emisor)",
        "Responsable", "Documento - Archivo", "Estado",
    ]
    ws.append(headers)
    thin = Side(style="thin", color="4472C4")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for row in rows:
        d = _format_log_row_for_export(row, request)
        ws.append([
            d["transmittal"],
            d["referencia"],
            d["fecha_envio"],
            d["responsable"],
            d["documento_archivo"],
            d["estado"],
        ])
    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    light = Side(style="thin", color="CCCCCC")
    for r in range(1, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.alignment = wrap_align
            if r > 1:
                cell.border = Border(top=light, left=light, right=light, bottom=light)
    # Filtros automáticos en la primera fila (encabezados)
    n_data = len(rows) + 1
    ws.auto_filter.ref = f"A1:F{n_data}"
    # Márgenes pequeños para más espacio de tabla
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.4, header=0.2, footer=0.2)
    # Anchos de columna razonables
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 12
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _logs_pdf_response(rows, list_name, inline=False):
    """
    PDF para listados de logs: mismas columnas que el template.
    Orientación horizontal, márgenes pequeños, celdas multilínea (Paragraph), estilo elegante.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.units import cm
    from io import BytesIO
    fecha_str = timezone.now().strftime("%d-%m-%Y")
    safe_name = (list_name or "Logs").replace(" ", "-").replace("/", "-")[:40]
    filename = f"{safe_name}-{fecha_str}.pdf"
    # Horizontal (landscape)
    pagesize = (A4[1], A4[0])
    buf = BytesIO()
    left_right = 0.4 * cm
    top_bottom = 0.6 * cm
    doc = SimpleDocTemplate(buf, pagesize=pagesize, leftMargin=left_right, rightMargin=left_right, topMargin=top_bottom, bottomMargin=top_bottom)
    styles = getSampleStyleSheet()
    para_cell = ParagraphStyle("LogCell", parent=styles["Normal"], fontName="Helvetica", fontSize=6, leading=7.5, leftIndent=0, rightIndent=0, spaceBefore=0, spaceAfter=0)
    para_header = ParagraphStyle("LogHeader", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=7, leading=8)

    def truncate_for_pdf_cell(text, max_chars=900, max_lines=35):
        """Evita que una celda tenga tanto texto que la fila supere la altura de la página (LayoutError)."""
        if not text:
            return ""
        s = str(text).strip()
        lines = s.splitlines()
        if len(lines) > max_lines:
            lines = lines[:max_lines]
            s = "\n".join(lines) + " …"
        if len(s) > max_chars:
            s = s[:max_chars].rstrip() + " …"
        return s

    def safe_para(text, style=para_cell):
        if text is None:
            text = ""
        t = truncate_for_pdf_cell(str(text))
        s = t.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")
        return Paragraph(s or " ", style)

    headers = [
        "Transmittal", "Referencia", "Fecha envío (emisor)",
        "Responsable", "Documento - Archivo", "Estado",
    ]
    data = [[safe_para(h, para_header) for h in headers]]
    for row in rows:
        d = _format_log_row_for_export(row, request=None)
        data.append([
            safe_para(d["transmittal"]),
            safe_para(d["referencia"]),
            safe_para(d["fecha_envio"]),
            safe_para(d["responsable"]),
            safe_para(d["documento_archivo"]),
            safe_para(d["estado"]),
        ])
    # Anchos repartidos para landscape A4 (~27.7 cm útil), 6 columnas
    col_widths = [3.5*cm, 4.0*cm, 2.6*cm, 3.2*cm, 9.8*cm, 2.6*cm]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5090")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B0BEC5")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ECEFF1")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    doc.build([t])
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type="application/pdf")
    if inline:
        response["Content-Disposition"] = f'inline; filename="{filename}"'
    else:
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_GET
def logs_propamat_odata(request):
    """
    Logs Propamat a Odata: carpetas cuyo código contiene TRN.
    """
    rows = _get_logs_folder_rows("TRN")
    if request.GET.get("format") == "excel":
        # Hoja/archivo claramente identificados como TRN (Propamat → Odata)
        return _logs_excel_response(rows, list_name="Logs TRN Propamat a Odata", request=request)
    if request.GET.get("format") == "pdf":
        open_inline = request.GET.get("open") == "1"
        return _logs_pdf_response(rows, list_name="Logs TRN Propamat a Odata", inline=open_inline)
    return render(
        request,
        "documents/logs_propamat_odata.html",
        {"rows": rows, "page_title": "Logs de Odata a Propamat", "page_subtitle": "Carpetas de Propamat."},
    )


@login_required
@require_GET
def logs_odata_propamat(request):
    """
    Logs Odata a Propamat: carpetas cuyo código contiene Odata.
    """
    rows = _get_logs_folder_rows("Odata")
    if request.GET.get("format") == "excel":
        # Hoja/archivo claramente identificados como ODATA-ST01 (Odata → Propamat)
        return _logs_excel_response(rows, list_name="Logs ODATA-ST01 Odata a Propamat", request=request)
    if request.GET.get("format") == "pdf":
        open_inline = request.GET.get("open") == "1"
        return _logs_pdf_response(rows, list_name="Logs ODATA-ST01 Odata a Propamat", inline=open_inline)
    return render(
        request,
        "documents/logs_odata_propamat.html",
        {"rows": rows, "page_title": "Propamat a Odata", "page_subtitle": "Carpetas Propamat."},
    )


@login_required
@require_GET
def informar_list(request):
    """
    Informar: misma data, columnas y exportación que «Logs Odata a Propamat»
    (_get_logs_folder_rows("Odata")).
    """
    rows = _get_logs_folder_rows("Odata")
    if request.GET.get("format") == "excel":
        return _logs_excel_response(rows, list_name="Logs ODATA-ST01 Odata a Propamat", request=request)
    if request.GET.get("format") == "pdf":
        open_inline = request.GET.get("open") == "1"
        return _logs_pdf_response(rows, list_name="Logs ODATA-ST01 Odata a Propamat", inline=open_inline)
    return render(
        request,
        "documents/informar.html",
        {"rows": rows, "page_title": "Informar de Odata", "page_subtitle": "Carpetas Propamat."},
    )

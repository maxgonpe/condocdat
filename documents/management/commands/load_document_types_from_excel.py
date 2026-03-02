"""
Carga tipos de documento (code, name) desde los Excel en doc/
hacia la tabla DocumentType en SQLite.

Fuentes (prioridad):
1. Matriz de codificación .xls — hoja Estructura, columnas TIP (code) y nombre (~118 tipos).
2. ESTATUS DOCUMENTAL PROPAMAT.xlsx, hoja PROCEDIMIENTOS.
3. Lista fija por defecto.

Uso:
  python manage.py load_document_types_from_excel
  python manage.py load_document_types_from_excel --path /ruta/a/doc
  python manage.py load_document_types_from_excel --clear
"""
from pathlib import Path

from django.core.management.base import BaseCommand
from django.conf import settings

from documents.models import DocumentType


DEFAULT_TYPES = [
    ("MAT", "Matriz de codificación"),
    ("DWG", "Plano"),
    ("TTAL", "Transmittal"),
    ("PPT", "Presentación"),
    ("CTA", "Carta"),
    ("NC", "No Conformidad"),
    ("INST", "Instructivo"),
    ("PL", "Plan"),
    ("PRO", "Procedimiento"),
    ("MC", "Memoria de cálculo"),
]


def load_from_matriz_xls(doc_dir):
    """Lee el .xls Matriz de codificación. Hoja Estructura, fila 40+ col 10=code col 11=name."""
    pairs = []
    for f in doc_dir.glob("*.xls"):
        if "Matriz" not in f.name and "matriz" not in f.name:
            continue
        try:
            import xlrd
        except ImportError:
            return []
        wb = xlrd.open_workbook(str(f))
        if "Estructura" not in wb.sheet_names():
            return []
        sh = wb.sheet_by_name("Estructura")
        for i in range(40, sh.nrows):
            row = sh.row_values(i)
            if len(row) < 12:
                continue
            code_cell = row[10]
            name_cell = row[11]
            code = str(code_cell).strip() if code_cell else ""
            name = str(name_cell).strip() if name_cell else ""
            if not code and not name:
                continue
            if not code and name:
                code = name.replace(" ", "")[:10]
            if not name:
                name = code
            code = code[:10].upper()
            if code and code.isalnum():
                pairs.append((code, name[:255]))
        break
    return pairs


def get_doc_dir(path=None):
    if path:
        return Path(path)
    return Path(settings.BASE_DIR) / "doc"


def extract_code_from_document_number(doc_number):
    """Extrae el código de tipo desde un número de documento."""
    if not doc_number:
        return None
    s = str(doc_number).strip()
    if "-" not in s:
        return s[:10] if s.isalnum() else None
    parts = [p.strip() for p in s.split("-") if p.strip()]
    # ODA-BUF-QA-MAT-00001 → MAT (parte 4)
    # FL-FO-CA-IIEE-PL-02 → PL (parte 5, la que es alfabética antes del número)
    if len(parts) >= 5 and parts[4].replace(".", "").isalnum():
        cand = parts[4][:10]
        if cand.isalpha():
            return cand
    if len(parts) >= 4 and parts[3].isalpha():
        return parts[3][:10]
    return None


def load_from_procedimientos_excel(doc_dir):
    """Lee ESTATUS DOCUMENTAL PROPAMAT.xlsx, hoja PROCEDIMIENTOS."""
    try:
        import openpyxl
    except ImportError:
        return []
    path = doc_dir / "ESTATUS DOCUMENTAL PROPAMAT.xlsx"
    if not path.exists():
        return []
    pairs = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "PROCEDIMIENTOS" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["PROCEDIMIENTOS"]
    for row in ws.iter_rows(min_row=10, max_row=600, values_only=True):
        if not row or len(row) < 2:
            continue
        codigo_celda = row[0]
        tipo_doc = row[1]
        if not tipo_doc or not isinstance(tipo_doc, str):
            continue
        name = str(tipo_doc).strip()
        if len(name) < 2 or name.upper() in ("TIPO DOCUMENTO", "NONE", "N/A"):
            continue
        code = extract_code_from_document_number(codigo_celda)
        if not code:
            code = name[:10].upper() if name.isalpha() else name.replace(" ", "")[:10]
        code = str(code).upper()[:10]
        pairs.append((code, name))
    wb.close()
    return pairs


def load_defaults():
    """Lista (code, name) por defecto."""
    return list(DEFAULT_TYPES)


class Command(BaseCommand):
    help = "Carga tipos de documento desde los Excel en doc/ a la tabla DocumentType."

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            default=None,
            help="Carpeta donde están los Excel (por defecto: <proyecto>/doc)",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Vaciar la tabla DocumentType antes de cargar (opcional).",
        )

    def handle(self, *args, **options):
        doc_dir = get_doc_dir(options.get("path"))
        if not doc_dir.exists():
            self.stdout.write(self.style.WARNING(f"No existe la carpeta: {doc_dir}"))
            self.stdout.write("Se cargarán solo los tipos por defecto.")

        if options.get("clear"):
            n = DocumentType.objects.count()
            DocumentType.objects.all().delete()
            self.stdout.write(f"Eliminados {n} tipos de documento.")

        # 1) Matriz .xls (lista completa ~118 tipos)
        matriz_pairs = load_from_matriz_xls(doc_dir) if doc_dir.exists() else []
        # 2) PROCEDIMIENTOS (complementa o sobrescribe nombre)
        excel_pairs = load_from_procedimientos_excel(doc_dir) if doc_dir.exists() else []
        # 3) Por defecto
        default_pairs = load_defaults()

        by_code = {}
        for code, name in matriz_pairs:
            code = str(code).upper()[:10]
            by_code[code] = name
        for code, name in excel_pairs:
            code = str(code).upper()[:10]
            by_code[code] = name
        for code, name in default_pairs:
            code = str(code).upper()[:10]
            if code not in by_code:
                by_code[code] = name

        created = 0
        updated = 0
        for code, name in by_code.items():
            obj, was_created = DocumentType.objects.get_or_create(
                code=code,
                defaults={"name": name},
            )
            if was_created:
                created += 1
                self.stdout.write(f"  Creado: {code} — {name}")
            else:
                if obj.name != name:
                    obj.name = name
                    obj.save()
                    updated += 1
                    self.stdout.write(f"  Actualizado: {code} — {name}")

        self.stdout.write(
            self.style.SUCCESS(
                f"Listo: {created} creados, {updated} actualizados. Total DocumentType: {DocumentType.objects.count()}"
            )
        )

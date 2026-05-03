from __future__ import annotations

import io
import re
import subprocess
import tempfile
from datetime import date, datetime
from pathlib import Path

from django.conf import settings
from django.core.files.base import ContentFile
from django.db.models import Max
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel as xl_from_excel

from .models import Transmital, TransmitalFolderConfig, TransmitalFolderLog

BASE_TEMPLATE = Path(
    getattr(
        settings,
        "TRANSMITAL_TEMPLATE_PATH",
        settings.BASE_DIR / "doc" / "ODATA-ST01-F5-TTAL-PPT-00293.xlsx",
    )
)
SHEET_NAME = "TTAL 013"

# Fecha fija en carátula (celda L4), requerida por operación.
FECHA_CARATULA_DEFAULT = date(2026, 1, 28)


def _bump_folder_config_to(consecutivo: int) -> None:
    """Mantiene alineado el consecutivo del creador de carpetas con los transmitales."""
    cfg = TransmitalFolderConfig.objects.order_by("id").first()
    if cfg is None:
        TransmitalFolderConfig.objects.create(
            base_path=str(BASE_TEMPLATE.parent),
            current_number=consecutivo,
        )
        return
    if cfg.current_number < consecutivo:
        cfg.current_number = consecutivo
        cfg.save(update_fields=["current_number", "updated_at"])


def _cell_to_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            dt = xl_from_excel(float(v))
            return dt.date() if isinstance(dt, datetime) else None
        except Exception:
            return None
    return None


def _date_to_excel(v: date | None):
    if not v:
        return None
    # Escribir solo fecha (sin hora) para evitar desborde visual en celdas.
    return v


def _next_consecutivo() -> int:
    mx_db = Transmital.objects.aggregate(mx=Max("consecutivo"))["mx"] or 0
    mx_doc = 0
    docs_dir = BASE_TEMPLATE.parent
    for p in docs_dir.glob("ODATA-ST01-F5-TTAL-PPT-*.xls*"):
        m = re.search(r"ODATA-ST01-F5-TTAL-PPT-(\d{5})", p.name, flags=re.IGNORECASE)
        if not m:
            continue
        mx_doc = max(mx_doc, int(m.group(1)))
    mx_cfg = 0
    cfg = TransmitalFolderConfig.objects.order_by("id").first()
    if cfg:
        mx_cfg = int(cfg.current_number or 0)
    mx_log = TransmitalFolderLog.objects.aggregate(mx=Max("sequence_number"))["mx"] or 0
    return max(mx_db, mx_doc, mx_cfg, mx_log) + 1


def _codigo_from_consecutivo(n: int) -> str:
    return f"ODATA-ST01-F5-TTAL-PPT-{n:05d}"


def _safe_int(v, default=0):
    try:
        return int(v)
    except Exception:
        return default


def create_transmital_from_template() -> Transmital:
    if not BASE_TEMPLATE.is_file():
        raise FileNotFoundError(f"No existe plantilla: {BASE_TEMPLATE}")

    consecutivo = _next_consecutivo()
    codigo = _codigo_from_consecutivo(consecutivo)
    wb = load_workbook(BASE_TEMPLATE)
    ws = wb[SHEET_NAME]
    ws["I7"] = codigo
    ws["L4"] = _date_to_excel(FECHA_CARATULA_DEFAULT)
    ws["H10"] = _date_to_excel(timezone.localdate())

    revision = str(ws["L3"].value or "")
    fecha_envio = _cell_to_date(ws["H10"].value)
    numero_paginas = _safe_int(ws["J10"].value, 1)
    destinatario = str(ws["C11"].value or "").strip()
    empresa = str(ws["E11"].value or "").replace("Empresa:", "").strip()
    referencia = str(ws["A12"].value or "").strip()
    emision = str(ws["B32"].value or "").strip()
    unidad_revisora = str(ws["A55"].value or "").strip()
    unidad_emisora = str(ws["G55"].value or "").strip()

    item_snapshots = []
    for i in range(1, Transmital.ITEM_COUNT + 1):
        row = 13 + i
        item_snapshots.append(
            (
                str(ws[f"B{row}"].value or "").strip(),
                str(ws[f"F{row}"].value or "").strip(),
                str(ws[f"G{row}"].value or "").strip(),
                str(ws[f"K{row}"].value or "").strip(),
                str(ws[f"L{row}"].value or "").strip(),
            )
        )

    bio = io.BytesIO()
    wb.save(bio)
    wb.close()
    bio.seek(0)

    obj = Transmital(
        consecutivo=consecutivo,
        codigo_transmital=codigo,
        revision=revision,
        fecha_caratula=FECHA_CARATULA_DEFAULT,
        fecha_envio=fecha_envio,
        numero_paginas=numero_paginas,
        destinatario=destinatario,
        empresa=empresa,
        referencia=referencia,
        emision=emision,
        unidad_revisora=unidad_revisora,
        unidad_emisora=unidad_emisora,
    )
    obj.file.save(f"{codigo}.xlsx", ContentFile(bio.getvalue()), save=False)

    for i, snap in enumerate(item_snapshots, start=1):
        setattr(obj, f"item_{i:02d}_documento", snap[0])
        setattr(obj, f"item_{i:02d}_rev_documento", snap[1])
        setattr(obj, f"item_{i:02d}_titulo", snap[2])
        setattr(obj, f"item_{i:02d}_rev_emisor", snap[3])
        setattr(obj, f"item_{i:02d}_estatus", snap[4])

    obj.save()
    _bump_folder_config_to(consecutivo)
    return obj


def sync_transmital_to_excel(obj: Transmital) -> None:
    # Blindaje de emisión: carátula siempre fija al valor operativo.
    if obj.fecha_caratula != FECHA_CARATULA_DEFAULT:
        obj.fecha_caratula = FECHA_CARATULA_DEFAULT
        Transmital.objects.filter(pk=obj.pk).update(fecha_caratula=FECHA_CARATULA_DEFAULT)

    if not obj.fecha_envio:
        obj.fecha_envio = timezone.localdate()
        Transmital.objects.filter(pk=obj.pk).update(fecha_envio=obj.fecha_envio)

    wb = load_workbook(obj.file.path)
    ws = wb[SHEET_NAME]
    codigo = (obj.codigo_transmital or "").strip()
    ws["I7"] = codigo
    ws["L3"] = obj.revision
    ws["L4"] = _date_to_excel(FECHA_CARATULA_DEFAULT)
    ws["H10"] = _date_to_excel(obj.fecha_envio)
    ws["L4"].number_format = "dd-mm-yyyy"
    ws["H10"].number_format = "dd-mm-yyyy"
    ws["J10"] = obj.numero_paginas
    ws["C11"] = obj.destinatario
    ws["E11"] = f"Empresa: {obj.empresa}" if obj.empresa else ""
    ws["A12"] = obj.referencia
    ws["B32"] = obj.emision
    ws["A55"] = obj.unidad_revisora
    ws["G55"] = obj.unidad_emisora

    for i in range(1, Transmital.ITEM_COUNT + 1):
        row = 13 + i
        ws[f"A{row}"] = i
        doc = (getattr(obj, f"item_{i:02d}_documento") or "").strip()
        ws[f"B{row}"] = doc or None
        ws[f"F{row}"] = getattr(obj, f"item_{i:02d}_rev_documento")
        ws[f"G{row}"] = getattr(obj, f"item_{i:02d}_titulo")
        ws[f"K{row}"] = getattr(obj, f"item_{i:02d}_rev_emisor")
        ws[f"L{row}"] = getattr(obj, f"item_{i:02d}_estatus")

    wb.save(obj.file.path)
    wb.close()
    Transmital.objects.filter(pk=obj.pk).update(updated_at=timezone.now())


def transmital_download_filename(obj: Transmital) -> str:
    return f"{obj.codigo_transmital}.xlsx"


def transmital_pdf_filename(obj: Transmital) -> str:
    return f"{obj.codigo_transmital}.pdf"


def build_transmital_pdf_buffer(obj: Transmital) -> io.BytesIO:
    with tempfile.TemporaryDirectory(prefix="transmital_pdf_") as td:
        tmp_dir = Path(td)
        tmp_xlsx = tmp_dir / transmital_download_filename(obj)
        tmp_pdf = tmp_dir / transmital_pdf_filename(obj)
        tmp_xlsx.write_bytes(Path(obj.file.path).read_bytes())

        cmd = [
            "soffice",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(tmp_dir),
            str(tmp_xlsx),
        ]
        proc = subprocess.run(cmd, capture_output=True, text=True)
        if proc.returncode != 0 or not tmp_pdf.exists():
            detail = re.sub(r"\s+", " ", (proc.stderr or proc.stdout or "")).strip()
            raise RuntimeError(f"No se pudo convertir a PDF con LibreOffice: {detail}")

        return io.BytesIO(tmp_pdf.read_bytes())
